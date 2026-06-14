"""
LitManager CLI — 图书管理 AI 命令行工具
===============================================
用法:
  litman init                        初始化数据库
  litman model list                  列出所有模型
  litman model add                   添加 LLM 模型（交互式）
  litman model remove <id>           删除模型
  litman model test <id>             测试模型连接
  litman cat list                    列出图书分类
  litman cat add <name>              添加分类
  litman book list [--q KEYWORD]     列出/搜索图书
  litman book show <id>              查看图书详情
  litman book add                    添加图书（交互式）
  litman book edit <id>              编辑图书（交互式）
  litman book delete <id>            删除图书
  litman reader list                 列出读者
  litman reader add                  添加读者（交互式）
  litman copy list <book_id>         列出某图书的所有单册
  litman agent <指令>                自然语言数据库代理
  litman agent                       进入交互式 agent 模式
"""

import argparse
import asyncio
import json
import os
import sys
from datetime import date
from typing import Optional

# ── handle PyInstaller bundle ──
if getattr(sys, "frozen", False):
    sys.path.insert(0, sys._MEIPASS)
    DATA_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    DATA_DIR = os.path.dirname(os.path.abspath(__file__))
    if DATA_DIR not in sys.path:
        sys.path.insert(0, DATA_DIR)
os.environ.setdefault("LITMAN_DB_PATH", os.path.join(DATA_DIR, "llm_manager.db"))

from contextlib import asynccontextmanager
from app.database import get_db as _get_db_raw


@asynccontextmanager
async def get_db():
    async for session in _get_db_raw():
        yield session


# ═══════════════════════════════════════════════════════════
#  async helpers
# ═══════════════════════════════════════════════════════════

async def _init_db():
    from app.database import init_db
    import sys
    if getattr(sys, "frozen", False):
        # Use synchronous init for frozen mode (asyncio may hang)
        import sqlite3
        from app.models.model import Base
        from app.database import engine
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
        print("[OK] 数据库初始化完成。")
    else:
        await init_db()
        print("[OK] 数据库初始化完成。")


# ── Model ────────────────────────────────────────────────

async def _model_list():
    from app.models.model import LLMModel
    from sqlalchemy import select
    async with get_db() as db:
        result = await db.execute(select(LLMModel).order_by(LLMModel.id))
        models = result.scalars().all()
    if not models:
        print("(没有配置模型)")
        return
    from app.services.model_service import _resolve_api_base

    print(f"{'ID':<4} {'名称':<20} {'提供商':<12} {'模型':<24} {'API地址':<35}")
    print("-" * 95)
    for m in models:
        base = _resolve_api_base(m.provider, m.api_base) or "-"
        print(f"{m.id:<4} {m.name:<20} {m.provider:<12} {m.model_name:<24} {base:<35}")


async def _model_add():
    from app.services.model_service import PROVIDER_DEFAULTS

    provider = _input(
        "提供商 (openai/anthropic/azure/ollama/google/deepseek/groq/mistral/openrouter/github/other): ",
        default="openai",
    ).strip().lower()
    provider = provider.lstrip("\ufeff")

    defaults = PROVIDER_DEFAULTS.get(provider, {})
    default_base = defaults.get("api_base", "")

    name = _input("显示名称: ")
    model_name = _input(f"模型名 (例如 {_model_example(provider)}): ")
    api_base = _input("API Base URL:", default=default_base)
    api_key = _input("API Key:") if provider not in ("ollama",) else ""
    temperature = _input("温度:", default="0.7")
    max_tokens = _input("最大Token:", default="4096")

    from app.models.model import LLMModel
    async with get_db() as db:
        m = LLMModel(
            name=name, provider=provider, model_name=model_name,
            api_base=api_base or None, api_key=api_key or None,
            temperature=float(temperature), max_tokens=int(max_tokens),
        )
        db.add(m)
        await db.commit()
        await db.refresh(m)
    print(f"[OK] 模型已添加 (id={m.id})")

    if _input("立即测试连接? (y/n): ", default="y").lower() == "y":
        await _model_test(m.id)


def _model_example(provider: str) -> str:
    examples = {
        "openai": "gpt-4o",
        "anthropic": "claude-sonnet-4-20250514",
        "azure": "my-gpt4-deployment",
        "ollama": "gemma4:latest",
        "google": "gemini-2.0-flash",
        "deepseek": "deepseek-chat",
        "groq": "llama3-70b-8192",
        "together": "mistralai/Mixtral-8x7B-Instruct-v0.1",
        "mistral": "mistral-large-latest",
        "openrouter": "anthropic/claude-3.5-sonnet",
        "github": "gpt-4o",
    }
    return examples.get(provider, "model-name")


async def _model_test(model_id: int):
    from app.models.model import LLMModel
    from app.services.model_service import LLMClient

    async with get_db() as db:
        m = await db.get(LLMModel, model_id)
        if not m:
            print(f"[Error] 模型 {model_id} 不存在")
            return

    print(f"测试 {m.name} ({m.provider}/{m.model_name})...")
    client = LLMClient(
        provider=m.provider, api_base=m.api_base, api_key=m.api_key or "",
        model_name=m.model_name, temperature=0.7, max_tokens=256,
    )

    try:
        data = await client.chat_completion([
            {"role": "user", "content": "Reply with exactly 'OK' and nothing else."},
        ])
        reply = data["choices"][0]["message"]["content"].strip()
        usage = data.get("usage", {})
        print(f"  响应: {reply}")
        if usage.get("prompt_tokens"):
            print(f"  Token: {usage.get('prompt_tokens', '?')} 入 / {usage.get('completion_tokens', '?')} 出")
        print("[OK] 连接成功!")
    except Exception as e:
        print(f"[FAILED] {e}")
        print("  提示:")
        print("    - Ollama: 确保 'ollama serve' 已运行且模型已拉取")
        print("    - OpenAI/Anthropic: 检查 API Key 和模型名")
        print("    - Azure: 检查资源名、部署名和 API 版本")


async def _model_remove(model_id: int):
    from app.models.model import LLMModel
    async with get_db() as db:
        m = await db.get(LLMModel, model_id)
        if not m:
            print(f"[Error] 模型 {model_id} 不存在")
            return
        await db.delete(m)
        await db.commit()
    print(f"[OK] 模型 {model_id} 已删除")


# ── Category ─────────────────────────────────────────────

async def _cat_list():
    from app.models.model import Category
    from sqlalchemy import select
    async with get_db() as db:
        result = await db.execute(select(Category).order_by(Category.id))
        cats = result.scalars().all()
    if not cats:
        print("(没有分类)")
        return
    print(f"{'ID':<4} {'分类编码':<12} {'分类名称':<24} {'父分类ID':<10}")
    print("-" * 50)
    for c in cats:
        print(f"{c.id:<4} {(c.code or '-'):<12} {c.name:<24} {(str(c.parent_id) if c.parent_id else '-'):<10}")


async def _cat_add(name: str, code: Optional[str] = None):
    from app.models.model import Category
    if code is None:
        code = _input("分类编码: ", default="")
    async with get_db() as db:
        cat = Category(name=name, code=code or None)
        db.add(cat)
        await db.commit()
        await db.refresh(cat)
    print(f"[OK] 分类已添加 (id={cat.id})")


# ── Book ─────────────────────────────────────────────────

async def _book_list(keyword: Optional[str] = None):
    from app.models.model import Book
    from sqlalchemy import select, or_
    async with get_db() as db:
        query = select(Book).order_by(Book.id.desc()).limit(50)
        if keyword:
            like = f"%{keyword}%"
            query = query.where(or_(
                Book.title.ilike(like), Book.authors.ilike(like),
                Book.publisher.ilike(like), Book.isbn.ilike(like),
            ))
        result = await db.execute(query)
        items = result.scalars().all()
    if not items:
        print("(没有图书)")
        return
    h = f"{'ID':<5} {'书名':<52} {'作者':<24} {'出版社':<20} {'年份':<6} {'总册':<5} {'可借':<5}"
    sep = "-" * 117
    print(h)
    print(sep)
    for b in items:
        title = b.title[:48] + ".." if len(b.title) > 50 else b.title
        auth = (b.authors or "-")[:22] + ".." if b.authors and len(b.authors) > 24 else (b.authors or "-")
        pub = (b.publisher or "-")[:18] + ".." if b.publisher and len(b.publisher) > 20 else (b.publisher or "-")
        yr = str(b.publication_year) if b.publication_year else "-"
        print(f"{b.id:<5} {title:<52} {auth:<24} {pub:<20} {yr:<6} {(b.total_copies or 0):<5} {(b.available_copies or 0):<5}")


async def _book_show(book_id: int):
    from app.models.model import Book
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    async with get_db() as db:
        result = await db.execute(
            select(Book).options(selectinload(Book.category)).where(Book.id == book_id)
        )
        b = result.scalar_one_or_none()
    if not b:
        print(f"[Error] 图书 {book_id} 不存在")
        return
    cat_name = b.category.name if b.category else "-"
    print(f"\n{'='*60}")
    print(f"  图书 #{b.id}")
    print(f"{'='*60}")
    print(f"  索书号    : {b.call_number or '-'}")
    print(f"  书名      : {b.title}")
    print(f"  副书名    : {b.subtitle or '-'}")
    print(f"  作者      : {b.authors or '-'}")
    print(f"  出版社    : {b.publisher or '-'}")
    print(f"  出版年份  : {b.publication_year or '-'}")
    print(f"  版本      : {b.edition or '-'}")
    print(f"  ISBN      : {b.isbn or '-'}")
    print(f"  分类      : {cat_name} (ID:{b.category_id or '-'})")
    print(f"  语种      : {b.language or '-'}")
    print(f"  页数      : {b.pages or '-'}")
    print(f"  价格      : {b.price or '-'}")
    print(f"  馆藏位置  : {b.location or '-'}")
    print(f"  总册数    : {b.total_copies or 0}")
    print(f"  可借数量  : {b.available_copies or 0}")
    print(f"  图书状态  : {b.status or '-'}")
    print(f"  备注      : {b.notes or '-'}")


async def _book_add():
    title = _input("书名: ", required=True)
    call_number = _input("索书号: ")
    subtitle = _input("副书名: ")
    authors = _input("作者: ")
    publisher = _input("出版社: ")
    pub_year = _input("出版年份: ")
    edition = _input("版本: ")
    isbn = _input("ISBN: ")
    category_id = _input("分类ID (0=无): ", default="0")
    language = _input("语种: ")
    pages = _input("页数: ")
    price = _input("价格: ")
    location = _input("馆藏位置: ")
    total_copies = _input("总册数: ", default="1")
    available_copies = _input("可借数量: ", default="1")
    notes = _input("备注: ")

    from app.models.model import Book
    async with get_db() as db:
        b = Book(
            title=title,
            call_number=call_number or None,
            subtitle=subtitle or None,
            authors=authors or None,
            publisher=publisher or None,
            publication_year=int(pub_year) if pub_year else None,
            edition=edition or None,
            isbn=isbn or None,
            category_id=int(category_id) if category_id and int(category_id) > 0 else None,
            language=language or None,
            pages=int(pages) if pages else None,
            price=float(price) if price else None,
            location=location or None,
            total_copies=int(total_copies) if total_copies else 1,
            available_copies=int(available_copies) if available_copies else 1,
            notes=notes or None,
        )
        db.add(b)
        await db.commit()
        await db.refresh(b)

        # Auto-create copies if total_copies > 0
        total = b.total_copies or 1
        if hasattr(b, 'copies') and total > 0:
            from app.models.model import BookCopy
            for i in range(total):
                copy = BookCopy(
                    book_id=b.id,
                    asset_number=f"{b.id}-{i+1:03d}",
                    status="在馆",
                )
                db.add(copy)
            await db.commit()

    print(f"[OK] 图书已添加 (id={b.id}, 共 {b.total_copies or 1} 册)")


async def _book_edit(book_id: int):
    from app.models.model import Book
    async with get_db() as db:
        b = await db.get(Book, book_id)
        if not b:
            print(f"[Error] 图书 {book_id} 不存在")
            return

        b.title = _input("书名: ", default=b.title)
        b.call_number = _input("索书号: ", default=b.call_number or "")
        b.subtitle = _input("副书名: ", default=b.subtitle or "")
        b.authors = _input("作者: ", default=b.authors or "")
        b.publisher = _input("出版社: ", default=b.publisher or "")
        b.isbn = _input("ISBN: ", default=b.isbn or "")
        b.language = _input("语种: ", default=b.language or "")
        b.location = _input("馆藏位置: ", default=b.location or "")
        b.notes = _input("备注: ", default=b.notes or "")

        await db.commit()
    print(f"[OK] 图书 {book_id} 已更新")


async def _book_delete(book_id: int):
    from app.models.model import Book
    async with get_db() as db:
        b = await db.get(Book, book_id)
        if not b:
            print(f"[Error] 图书 {book_id} 不存在")
            return
        await db.delete(b)
        await db.commit()
    print(f"[OK] 图书 {book_id} 已删除")


# ── Reader ───────────────────────────────────────────────

async def _reader_list():
    from app.models.model import Reader
    from sqlalchemy import select
    async with get_db() as db:
        result = await db.execute(select(Reader).order_by(Reader.id))
        readers = result.scalars().all()
    if not readers:
        print("(没有读者)")
        return
    h = f"{'ID':<5} {'卡号':<16} {'姓名':<16} {'类型':<12} {'电话':<16} {'卡状态':<10} {'可借':<5}"
    sep = "-" * 80
    print(h)
    print(sep)
    for r in readers:
        print(f"{r.id:<5} {(r.card_number or '-'):<16} {(r.name or '-'):<16} {(r.identity_type or '-'):<12} {(r.phone or '-'):<16} {(r.card_status or '-'):<10} {(r.max_borrow or 0):<5}")


async def _reader_add():
    card_number = _input("读者卡号: ", required=True)
    name = _input("姓名: ", required=True)
    identity_type = _input("身份类型 (学生/教师/职工): ", default="学生")
    phone = _input("联系电话: ")
    max_borrow = _input("可借上限: ", default="10")

    from app.models.model import Reader
    async with get_db() as db:
        r = Reader(
            card_number=card_number,
            name=name,
            identity_type=identity_type or None,
            phone=phone or None,
            card_status="正常",
            max_borrow=int(max_borrow) if max_borrow else 10,
            register_date=date.today(),
        )
        db.add(r)
        await db.commit()
        await db.refresh(r)
    print(f"[OK] 读者已添加 (id={r.id})")


# ── BookCopy ─────────────────────────────────────────────

async def _copy_list(book_id: int):
    from app.models.model import BookCopy
    from sqlalchemy import select
    async with get_db() as db:
        result = await db.execute(
            select(BookCopy).where(BookCopy.book_id == book_id).order_by(BookCopy.id)
        )
        copies = result.scalars().all()
    if not copies:
        print(f"(图书 #{book_id} 没有单册记录)")
        return
    h = f"{'单册ID':<8} {'资产编号':<16} {'书籍状态':<12} {'存放书架':<16}"
    sep = "-" * 52
    print(h)
    print(sep)
    for c in copies:
        print(f"{c.id:<8} {(c.asset_number or '-'):<16} {(c.status or '-'):<12} {(c.shelf or '-'):<16}")


# ── Agent ────────────────────────────────────────────────

async def _agent_cmd(model_id: int, instruction: Optional[str] = None):
    from app.services.db_agent import DBAgent

    if instruction:
        async with get_db() as db:
            agent = DBAgent(model_id=model_id, db=db)
            result = await agent.execute(instruction)
        print(result)
        return

    print(f"\n{'='*60}")
    print("  LitManager Agent — 交互模式")
    print("  用自然语言描述你要做什么。")
    print("  示例:")
    print("    - '搜索机器学习的图书'")
    print("    - '添加一本书: 书名 三体, 作者 刘慈欣'")
    print("    - '显示编号5的图书'")
    print("    - '统计分类统计'")
    print("    - '列出所有读者'")
    print("  输入 'exit' 或 'quit' 退出。")
    print(f"{'='*60}\n")

    while True:
        try:
            instr = input("You: ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break
        if not instr:
            continue
        if instr.lower() in ("exit", "quit"):
            break

        async with get_db() as db:
            agent = DBAgent(model_id=model_id, db=db)
            try:
                result = await agent.execute(instr)
                print(f"\n{result}\n")
            except Exception as e:
                print(f"\n[Error] {e}\n")


# ═══════════════════════════════════════════════════════════
#  utility
# ═══════════════════════════════════════════════════════════



async def _book_import(filepath: str):
    """Import books from Excel file via API or directly."""
    import openpyxl
    from app.models.model import Book, BookCopy, Category
    from sqlalchemy import select

    async with get_db() as db:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]
        
        if 'title' not in headers:
            print("[Error] Excel 缺少 'title' 列")
            return
        
        cats = (await db.execute(select(Category))).scalars().all()
        cat_map = {c.name: c.id for c in cats}
        
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            rd = {}
            for i, v in enumerate(row):
                if i < len(headers) and headers[i]:
                    rd[headers[i]] = v
            title = rd.get('title')
            if not title:
                continue
            
            year = rd.get('publication_year')
            if year and isinstance(year, float):
                year = int(year)
            pages = rd.get('pages')
            if pages and isinstance(pages, float):
                pages = int(pages)
            copies = rd.get('total_copies')
            if copies and isinstance(copies, float):
                copies = int(copies)
            
            cat_name = rd.get('category', '')
            cat_id = cat_map.get(cat_name) if cat_name else None
            
            book = Book(
                title=str(title),
                call_number=str(rd.get('call_number', '')) or None,
                authors=str(rd.get('authors', '')) or None,
                publisher=str(rd.get('publisher', '')) or None,
                publication_year=year,
                isbn=str(rd.get('isbn', '')) or None,
                category_id=cat_id,
                language=str(rd.get('language', '')) or None,
                pages=pages,
                location=str(rd.get('location', '')) or None,
                total_copies=copies or 1,
                available_copies=copies or 1,
                notes=str(rd.get('notes', '')) or None,
            )
            db.add(book)
            await db.commit()
            await db.refresh(book)
            total = book.total_copies or 1
            for i in range(total):
                cp = BookCopy(book_id=book.id, asset_number=f"{book.id}-{i+1:03d}", status="\u5728\u9986")
                db.add(cp)
            await db.commit()
            imported += 1
        
        print(f"[OK] 成功导入 {imported} 本图书")


def _input(prompt_text: str, default: str = "", required: bool = False) -> str:
    full_prompt = prompt_text
    if default:
        full_prompt += f" [{default}]"
    full_prompt += " "
    try:
        val = input(full_prompt).strip()
    except EOFError:
        val = ""
    if not val and default:
        val = default
    if required and not val:
        print("[Error] 此项必填。")
        sys.exit(1)
    return val


def _get_model_id(args) -> int:
    if args.model:
        return args.model
    asyncio.run(_model_list())
    try:
        return int(input("选择模型 ID: ").strip())
    except (ValueError, EOFError):
        print("无效的模型ID")
        sys.exit(1)


# ═══════════════════════════════════════════════════════════
#  argparse
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        prog="litman",
        description="LitManager — 图书 AI 管理工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    sub = parser.add_subparsers(dest="command")

    # init
    sub.add_parser("init", help="初始化数据库")

    # model
    model_p = sub.add_parser("model", help="管理 LLM 模型")
    model_sub = model_p.add_subparsers(dest="sub", required=True)
    model_sub.add_parser("list", help="列出所有模型")
    model_sub.add_parser("add", help="添加模型（交互式）")
    model_rm_p = model_sub.add_parser("remove", help="删除模型")
    model_rm_p.add_argument("id", type=int)
    model_test_p = model_sub.add_parser("test", help="测试模型连接")
    model_test_p.add_argument("id", type=int)

    # category
    cat_p = sub.add_parser("cat", help="管理图书分类")
    cat_sub = cat_p.add_subparsers(dest="sub", required=True)
    cat_sub.add_parser("list", help="列出分类")
    cat_add_p =     cat_add_p = cat_sub.add_parser("add", help="添加分类")
    cat_add_p.add_argument("name", help="分类名称")
    cat_add_p.add_argument("--code", help="分类编码")

    # book (replaces literature)
    book_p = sub.add_parser("book", help="管理图书")
    book_sub = book_p.add_subparsers(dest="sub", required=True)
    book_list_p = book_sub.add_parser("list", help="列出/搜索图书")
    book_list_p.add_argument("--q", help="搜索关键词")
    book_show_p = book_sub.add_parser("show", help="查看图书详情")
    book_show_p.add_argument("id", type=int)
    book_sub.add_parser("add", help="添加图书（交互式）")
    book_edit_p = book_sub.add_parser("edit", help="编辑图书（交互式）")
    book_edit_p.add_argument("id", type=int)
    book_del_p = book_sub.add_parser("delete", help="删除图书")
    book_del_p.add_argument("id", type=int)
    book_import_p = book_sub.add_parser("import", help="从 Excel 导入图书")
    book_import_p.add_argument("file", help="Excel 文件路径 (.xlsx)")

    # reader
    reader_p = sub.add_parser("reader", help="管理读者")
    reader_sub = reader_p.add_subparsers(dest="sub", required=True)
    reader_sub.add_parser("list", help="列出所有读者")
    reader_sub.add_parser("add", help="添加读者（交互式）")

    # copy
    copy_p = sub.add_parser("copy", help="管理单册资产")
    copy_sub = copy_p.add_subparsers(dest="sub", required=True)
    copy_list_p = copy_sub.add_parser("list", help="列出某图书的所有单册")
    copy_list_p.add_argument("book_id", type=int)

    # agent
    agent_p = sub.add_parser("agent", help="自然语言数据库代理")
    agent_p.add_argument("instruction", nargs="*", help="自然语言指令（省略则进入交互模式）")
    agent_p.add_argument("--model", type=int, help="模型 ID（默认让选择）")

    args = parser.parse_args()

    try:
        if args.command is None:
            # Default to agent interactive mode when double-clicked
            _model_id = _get_model_id(args)
            asyncio.run(_agent_cmd(model_id, None))
        elif args.command == "init":
            asyncio.run(_init_db())

        elif args.command == "model":
            if args.sub == "list":
                asyncio.run(_model_list())
            elif args.sub == "add":
                asyncio.run(_model_add())
            elif args.sub == "remove":
                asyncio.run(_model_remove(args.id))
            elif args.sub == "test":
                asyncio.run(_model_test(args.id))

        elif args.command == "cat":
            if args.sub == "list":
                asyncio.run(_cat_list())
            elif args.sub == "add":
                asyncio.run(_cat_add(args.name, args.code))

        elif args.command == "book":
            if args.sub == "list":
                asyncio.run(_book_list(args.q))
            elif args.sub == "import":
                asyncio.run(_book_import(args.file))
            elif args.sub == "show":
                asyncio.run(_book_show(args.id))
            elif args.sub == "add":
                asyncio.run(_book_add())
            elif args.sub == "edit":
                asyncio.run(_book_edit(args.id))
            elif args.sub == "delete":
                asyncio.run(_book_delete(args.id))

        elif args.command == "reader":
            if args.sub == "list":
                asyncio.run(_reader_list())
            elif args.sub == "add":
                asyncio.run(_reader_add())

        elif args.command == "copy":
            if args.sub == "list":
                asyncio.run(_copy_list(args.book_id))

        elif args.command == "agent":
            model_id = _get_model_id(args)
            instruction = " ".join(args.instruction) if args.instruction else None
            asyncio.run(_agent_cmd(model_id, instruction))

    except KeyboardInterrupt:
        print("\n[Interrupted]")
    except Exception as e:
        print(f"[Fatal Error] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
