# -*- coding: utf-8 -*-
"""Export router. Excel export endpoints for books, readers, and borrowing records."""

import io
from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import openpyxl

from app.database import get_db
from app.models.model import Book, Reader, BorrowingRecord, BookCopy
from app.security import get_current_admin

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/books")
async def export_books(
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Export books to Excel."""
    result = await db.execute(
        select(Book).options(selectinload(Book.category))
    )
    books = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "图书列表"

    headers = ["ID", "书名", "作者", "出版社", "出版年份", "ISBN", "分类", "总册数", "可借册数", "状态"]
    ws.append(headers)

    for b in books:
        ws.append([
            b.id, b.title, b.authors or "", b.publisher or "",
            b.publication_year, b.isbn or "",
            b.category.name if b.category else "",
            b.total_copies or 0, b.available_copies or 0, b.status or ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=books.xlsx"}
    )


@router.get("/readers")
async def export_readers(
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Export readers to Excel."""
    result = await db.execute(select(Reader))
    readers = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "读者列表"

    headers = ["ID", "卡号", "姓名", "身份类型", "电话", "注册日期", "卡状态", "最大借阅数"]
    ws.append(headers)

    for r in readers:
        ws.append([
            r.id, r.card_number, r.name, r.identity_type or "",
            str(r.register_date) if r.register_date else "",
            r.card_status or "", r.max_borrow or 10
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=readers.xlsx"}
    )


@router.get("/borrowings")
async def export_borrowings(
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Export borrowing records to Excel."""
    result = await db.execute(
        select(BorrowingRecord)
        .options(
            selectinload(BorrowingRecord.reader),
            selectinload(BorrowingRecord.copy).selectinload(BookCopy.book)
        )
    )
    records = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "借阅记录"

    headers = ["ID", "读者", "图书", "借阅日期", "应还日期", "归还日期", "逾期天数", "状态"]
    ws.append(headers)

    for r in records:
        reader_name = r.reader.name if r.reader else ""
        book_title = r.copy.book.title if r.copy and r.copy.book else ""
        ws.append([
            r.id, reader_name, book_title,
            str(r.borrow_date) if r.borrow_date else "",
            str(r.due_date) if r.due_date else "",
            str(r.return_date) if r.return_date else "",
            r.overdue_days or 0, r.status or ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=borrowings.xlsx"}
    )
