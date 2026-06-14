"""Reader router. English response keys, added PUT endpoint."""

import datetime
import io
import openpyxl

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.model import Reader, BorrowingRecord, BookCopy, Book
from app.schemas.model import ReaderCreate, ReaderUpdate
from app.utils import classify_borrowing_status, log_audit
from app.security import get_current_admin, create_access_token, get_current_reader


router = APIRouter(prefix="/api/readers", tags=["readers"])


class ReaderLoginRequest(BaseModel):
    card_number: str
    password: str


@router.post("/login")
async def reader_login(body: ReaderLoginRequest, db: AsyncSession = Depends(get_db)):
    """Reader login endpoint."""
    result = await db.execute(
        select(Reader).where(Reader.card_number == body.card_number)
    )
    reader = result.scalar_one_or_none()
    
    if not reader:
        raise HTTPException(status_code=401, detail="读者不存在")
    
    if not reader.password:
        raise HTTPException(status_code=401, detail="该读者未设置密码")
    
    if not reader.check_password(body.password):
        raise HTTPException(status_code=401, detail="密码错误")
    
    # Commit password migration if check_password re-hashed from SHA256 to bcrypt
    await db.commit()
    
    if reader.card_status != "正常":
        raise HTTPException(status_code=403, detail="读者卡状态异常")
    
    # Generate JWT token
    token = create_access_token(
        data={
            "sub": str(reader.id),
            "card_number": reader.card_number,
            "role": "reader",
            "type": "reader"
        }
    )
    
    return {
        "token": token,
        "id": reader.id,
        "card_number": reader.card_number,
        "name": reader.name,
        "role": "reader"
    }


@router.get("/")
async def list_readers(
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    result = await db.execute(select(Reader).order_by(Reader.id))
    rows = result.scalars().all()
    return {"data": [{"id": r.id, "card_number": r.card_number, "name": r.name,
                       "identity_type": r.identity_type, "phone": r.phone,
                       "register_date": str(r.register_date) if r.register_date else "",
                       "card_status": r.card_status, "max_borrow": r.max_borrow} for r in rows],
            "total": len(rows)}


@router.post("/")
async def create_reader(
    body: ReaderCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    exists = await db.execute(select(Reader).where(Reader.card_number == body.card_number))
    if exists.scalar_one_or_none():
        raise HTTPException(400, "读者卡号已存在")
    reader = Reader(
        card_number=body.card_number,
        name=body.name,
        identity_type=body.identity_type,
        phone=body.phone,
        register_date=datetime.date.today() if not body.register_date else datetime.date.fromisoformat(str(body.register_date)),
        card_status=body.card_status or "正常",
        max_borrow=body.max_borrow or 10,
    )
    # Set initial password to card_number (hashed with bcrypt)
    reader.set_password(body.card_number)
    db.add(reader)
    await db.commit()
    await db.refresh(reader)

    # Log audit
    await log_audit(
        db=db,
        user_id=current_admin["id"],
        user_type="admin",
        action="create",
        resource_type="reader",
        resource_id=reader.id,
        details={"card_number": reader.card_number, "name": reader.name},
        ip_address=request.client.host if request.client else None
    )

    return {"id": reader.id, "message": "添加成功"}


@router.put("/{reader_id}")
async def update_reader(
    reader_id: int,
    body: ReaderUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(reader, k, v)
    await db.commit()

    # Log audit
    await log_audit(
        db=db,
        user_id=current_admin["id"],
        user_type="admin",
        action="update",
        resource_type="reader",
        resource_id=reader.id,
        details={"card_number": reader.card_number, "name": reader.name, "changes": body.model_dump(exclude_unset=True)},
        ip_address=request.client.host if request.client else None
    )

    return {"message": "更新成功"}


@router.delete("/{reader_id}")
async def delete_reader(
    reader_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    from app.models.model import BorrowingRecord
    active = await db.execute(
        select(BorrowingRecord).where(
            BorrowingRecord.reader_id == reader_id,
            BorrowingRecord.status == "借出"
        )
    )
    if active.scalar_one_or_none():
        raise HTTPException(400, "该读者有未归还的借阅，无法删除")
    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    # Log audit before deletion
    await log_audit(
        db=db,
        user_id=current_admin["id"],
        user_type="admin",
        action="delete",
        resource_type="reader",
        resource_id=reader.id,
        details={"card_number": reader.card_number, "name": reader.name},
        ip_address=request.client.host if request.client else None
    )

    await db.delete(reader)
    await db.commit()
    return {"message": "删除成功"}


@router.post("/import", status_code=201)
async def import_readers(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Import readers from Excel file.

    Expected columns (case-insensitive):
    - card_number (required): reader card number
    - name (required): reader name
    - identity_type: ID card / passport / etc.
    - phone: phone number
    - card_status: 正常 / 挂失 / 注销 (default: 正常)
    - max_borrow: max books allowed (default: 10)
    """
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception:
        raise HTTPException(400, "无法读取 Excel 文件")

    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    # Map expected columns
    col_map = {}
    for i, h in enumerate(headers):
        if h in ("card_number", "卡号", "读者卡号"):
            col_map["card_number"] = i
        elif h in ("name", "姓名", "读者姓名"):
            col_map["name"] = i
        elif h in ("identity_type", "证件类型", "身份证类型"):
            col_map["identity_type"] = i
        elif h in ("phone", "电话", "手机号"):
            col_map["phone"] = i
        elif h in ("card_status", "卡状态", "状态"):
            col_map["card_status"] = i
        elif h in ("max_borrow", "最大借阅数", "借阅上限"):
            col_map["max_borrow"] = i

    if "card_number" not in col_map or "name" not in col_map:
        raise HTTPException(400, "Excel 文件缺少必需列：card_number（卡号）和 name（姓名）")

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue

        card_number = str(row[col_map["card_number"]]).strip() if row[col_map["card_number"]] else ""
        name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""

        if not card_number or not name:
            errors.append({"row": row_idx, "error": "卡号或姓名为空"})
            skipped += 1
            continue

        # Check duplicate
        exists = await db.execute(select(Reader).where(Reader.card_number == card_number))
        if exists.scalar_one_or_none():
            errors.append({"row": row_idx, "error": f"卡号 {card_number} 已存在"})
            skipped += 1
            continue

        reader = Reader(
            card_number=card_number,
            name=name,
            identity_type=str(row[col_map["identity_type"]]).strip() if "identity_type" in col_map and row[col_map["identity_type"]] else None,
            phone=str(row[col_map["phone"]]).strip() if "phone" in col_map and row[col_map["phone"]] else None,
            register_date=datetime.date.today(),
            card_status=str(row[col_map["card_status"]]).strip() if "card_status" in col_map and row[col_map["card_status"]] else "正常",
            max_borrow=int(row[col_map["max_borrow"]]) if "max_borrow" in col_map and row[col_map["max_borrow"]] else 10,
        )
        db.add(reader)
        imported += 1

    await db.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "total_rows": imported + skipped,
        "errors": errors[:20],  # Limit error list
    }


# ── User Profile ────────────────────────────────────────

@router.get("/{reader_id}/profile")
async def get_reader_profile(
    reader_id: int,
    db: AsyncSession = Depends(get_db),
    current_reader: dict = Depends(get_current_reader)
):
    """Get reader profile with borrowing statistics."""
    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    # Count borrowings
    result = await db.execute(
        select(BorrowingRecord).where(BorrowingRecord.reader_id == reader_id)
    )
    records = result.scalars().all()

    active = sum(1 for r in records if r.status == "借出")
    returned = sum(1 for r in records if r.status == "已归还")
    overdue = sum(1 for r in records if r.status == "逾期未还")

    return {
        "id": reader.id,
        "name": reader.name,
        "card_number": reader.card_number,
        "identity_type": reader.identity_type or "",
        "phone": reader.phone or "",
        "register_date": str(reader.register_date) if reader.register_date else "",
        "card_status": reader.card_status,
        "max_borrow": reader.max_borrow,
        "stats": {
            "total_borrowed": len(records),
            "active": active,
            "returned": returned,
            "overdue": overdue,
        }
    }


@router.get("/{reader_id}/borrowings")
async def get_reader_borrowings(
    reader_id: int,
    db: AsyncSession = Depends(get_db),
    current_reader: dict = Depends(get_current_reader)
):
    """Get all borrowing records for a reader."""
    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    result = await db.execute(
        select(BorrowingRecord)
        .options(selectinload(BorrowingRecord.copy).selectinload(BookCopy.book))
        .where(BorrowingRecord.reader_id == reader_id)
        .order_by(BorrowingRecord.borrow_date.desc())
    )
    records = result.scalars().all()

    items = []
    for r in records:
        book_title = r.copy.book.title if r.copy and r.copy.book else ""

        items.append({
            "id": r.id,
            "book_title": book_title,
            "copy_id": r.copy_id,
            "borrow_date": str(r.borrow_date) if r.borrow_date else "",
            "due_date": str(r.due_date) if r.due_date else "",
            "return_date": str(r.return_date) if r.return_date else "",
            "overdue_days": r.overdue_days or 0,
            "status": r.status,
            "classified_status": _classify_status(r),
            "renewed": r.renewed if hasattr(r, "renewed") else 0,
        })

    return {"data": items, "total": len(items)}


@router.get("/{reader_id}/overdue")
async def get_reader_overdue(
    reader_id: int,
    db: AsyncSession = Depends(get_db),
    current_reader: dict = Depends(get_current_reader)
):
    """Get overdue/active borrowing records for a reader."""
    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    result = await db.execute(
        select(BorrowingRecord)
        .options(selectinload(BorrowingRecord.copy).selectinload(BookCopy.book))
        .where(
            BorrowingRecord.reader_id == reader_id,
            BorrowingRecord.status.in_(["借出", "临期未还", "逾期未还"])
        )
        .order_by(BorrowingRecord.due_date.asc())
    )
    records = result.scalars().all()

    items = []
    for r in records:
        book_title = r.copy.book.title if r.copy and r.copy.book else ""

        items.append({
            "id": r.id,
            "book_title": book_title,
            "copy_id": r.copy_id,
            "borrow_date": str(r.borrow_date) if r.borrow_date else "",
            "due_date": str(r.due_date) if r.due_date else "",
            "overdue_days": r.overdue_days or 0,
            "status": r.status,
            "classified_status": _classify_status(r),
            "renewed": r.renewed if hasattr(r, "renewed") else 0,
        })

    return {"data": items, "total": len(items)}


@router.post("/{reader_id}/avatar")
async def upload_avatar(
    reader_id: int,
    avatar: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload avatar image for a reader. Saves to ./avatars/ directory."""
    import os

    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    # Validate file type
    allowed = {"image/png", "image/jpeg", "image/jpg", "image/gif", "image/webp", "image/bmp"}
    if avatar.content_type not in allowed:
        raise HTTPException(400, f"不支持的图片格式: {avatar.content_type}")

    # Save avatar
    avatars_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "avatars")
    os.makedirs(avatars_dir, exist_ok=True)

    ext = avatar.filename.rsplit(".", 1)[-1].lower() if "." in avatar.filename else "png"
    filename = f"reader_{reader_id}.{ext}"
    filepath = os.path.join(avatars_dir, filename)

    with open(filepath, "wb") as f:
        content = await avatar.read()
        f.write(content)

    return {"message": "头像上传成功", "path": filepath}


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@router.post("/{reader_id}/change-password")
async def change_reader_password(
    reader_id: int,
    body: ChangePasswordRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_reader: dict = Depends(get_current_reader)
):
    """Reader changes own password."""
    # Verify the reader is changing their own password
    if current_reader["id"] != reader_id:
        raise HTTPException(403, "只能修改自己的密码")

    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    # Verify old password
    if not reader.check_password(body.old_password):
        raise HTTPException(400, "原密码错误")

    # Validate new password
    if len(body.new_password) < 6:
        raise HTTPException(400, "新密码长度至少6位")

    # Set new password
    reader.set_password(body.new_password)
    await db.commit()

    # Log audit
    await log_audit(
        db=db,
        user_id=reader_id,
        user_type="reader",
        action="change_password",
        resource_type="reader",
        resource_id=reader_id,
        details={"card_number": reader.card_number, "name": reader.name},
        ip_address=request.client.host if request.client else None
    )

    return {"message": "密码修改成功"}


class ResetPasswordRequest(BaseModel):
    new_password: str


@router.post("/{reader_id}/reset-password")
async def reset_reader_password(
    reader_id: int,
    body: ResetPasswordRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Admin resets reader password."""
    reader = await db.get(Reader, reader_id)
    if not reader:
        raise HTTPException(404, "读者不存在")

    reader.set_password(body.new_password)
    await db.commit()

    # Log audit
    await log_audit(
        db=db,
        user_id=current_admin["id"],
        user_type="admin",
        action="reset_password",
        resource_type="reader",
        resource_id=reader.id,
        details={"card_number": reader.card_number, "name": reader.name},
        ip_address=request.client.host if request.client else None
    )

    return {"message": "密码重置成功"}


@router.get("/{reader_id}/fines")
async def get_reader_fines(
    reader_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Get all fines for a reader."""
    from app.models.model import Fine
    result = await db.execute(
        select(Fine).where(Fine.reader_id == reader_id).order_by(Fine.id.desc())
    )
    fines = result.scalars().all()
    data = [{
        "id": f.id,
        "borrowing_id": f.borrowing_id,
        "amount": f.amount,
        "reason": f.reason,
        "is_paid": f.is_paid,
        "created_at": str(f.created_at) if f.created_at else ""
    } for f in fines]
    return {"data": data, "total": len(data)}
