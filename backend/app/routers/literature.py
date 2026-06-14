"""Books & categories router. Uses clean English keys in responses."""

import json
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import func as sa_func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.model import Book, BookCopy, Category
from app.schemas.model import BookCreate, BookUpdate, CategoryCreate, CategoryResponse
from app.utils import log_audit
from app.security import get_current_admin

import io
import openpyxl
from fastapi import UploadFile, File

router = APIRouter(prefix="/api/books", tags=["books"])


# ── Statistics ───────────────────────────────────────

@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_db)):
    """Get library statistics."""
    total_books = (await db.execute(select(sa_func.count(Book.id)))).scalar() or 0
    total_copies = (await db.execute(select(sa_func.sum(Book.total_copies)))).scalar() or 0
    available_copies = (await db.execute(select(sa_func.sum(Book.available_copies)))).scalar() or 0
    
    return {
        "total_books": total_books,
        "total_copies": total_copies,
        "available_copies": available_copies,
    }


# ── Categories ───────────────────────────────────────

@router.get("/categories")
async def list_categories(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Category).order_by(Category.id))
    cats = result.scalars().all()
    return [{"id": c.id, "code": c.code, "name": c.name, "parent_id": c.parent_id,
             "children": []} for c in cats]


@router.post("/categories", status_code=201)
async def create_category(
    data: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    cat = Category(**data.model_dump())
    db.add(cat)
    await db.commit()
    await db.refresh(cat)
    return cat


@router.put("/categories/{cat_id}")
async def update_category(
    cat_id: int,
    data: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    cat = await db.get(Category, cat_id)
    if not cat:
        raise HTTPException(404, "分类不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(cat, k, v)
    await db.commit()
    await db.refresh(cat)
    return cat


@router.delete("/categories/{cat_id}", status_code=204)
async def delete_category(
    cat_id: int,
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    cat = await db.get(Category, cat_id)
    if not cat:
        raise HTTPException(404, "分类不存在")
    # check for books under this category
    book_count = (await db.execute(select(sa_func.count(Book.id)).where(Book.category_id == cat_id))).scalar()
    if book_count:
        raise HTTPException(400, f"该分类下有 {book_count} 本图书，无法删除")
    await db.delete(cat)
    await db.commit()


# ── Books ────────────────────────────────────────────

def _book_dict(b: Book, cat_name: str = "") -> dict:
    return {
        "id": b.id,
        "call_number": b.call_number or "",
        "title": b.title,
        "subtitle": b.subtitle or "",
        "authors": b.authors or "",
        "publisher": b.publisher or "",
        "publication_year": b.publication_year,
        "edition": b.edition or "",
        "isbn": b.isbn or "",
        "category_id": b.category_id,
        "category_name": cat_name,
        "language": b.language or "",
        "pages": b.pages,
        "price": b.price,
        "location": b.location or "",
        "total_copies": b.total_copies or 0,
        "available_copies": b.available_copies or 0,
        "status": b.status or "",
        "notes": b.notes or "",
        "created_at": str(b.created_at) if b.created_at else "",
    }


@router.get("")
async def list_books(
    q: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    query = select(Book)
    if q:
        like = f"%{q}%"
        query = query.where(or_(
            Book.title.ilike(like), Book.authors.ilike(like),
            Book.publisher.ilike(like), Book.isbn.ilike(like),
        ))
    if category_id is not None:
        query = query.where(Book.category_id == category_id)

    count_q = select(sa_func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    query = query.order_by(Book.id.desc()).offset((page - 1) * page_size).limit(page_size)
    query = query.options(selectinload(Book.category))
    items = (await db.execute(query)).scalars().all()

    result = []
    for b in items:
        cat_name = b.category.name if b.category else ""
        result.append(_book_dict(b, cat_name))
    return {"items": result, "total": total}


@router.post("/search-ai")
async def search_books_ai(data: dict, db: AsyncSession = Depends(get_db)):
    """Smart search with LLM. Accepts natural language description, LLM generates SQL-like filters."""
    from app.models.model import LLMModel
    from app.services.model_service import LLMClient
    
    query_text = data.get("query", "").strip()
    if not query_text:
        raise HTTPException(400, "请输入搜索描述")
    
    # Check LLM availability
    model = await db.get(LLMModel, 1)
    if not model:
        # Fallback to basic search if no LLM
        query = select(Book)
        like = f"%{query_text}%"
        query = query.where(or_(
            Book.title.ilike(like), Book.authors.ilike(like),
            Book.publisher.ilike(like), Book.isbn.ilike(like),
        ))
        items = (await db.execute(query.limit(100))).scalars().all()
        result = [_book_dict(b) for b in items]
        return {"items": result, "total": len(result), "llm_used": False}
    
    # Build category context
    cats_result = await db.execute(select(Category).order_by(Category.id))
    cats = cats_result.scalars().all()
    cat_ctx = "Available categories:\n" + "\n".join(
        f"  [{c.id}] {c.name} (code: {c.code or '-'})" for c in cats
    ) if cats else "No categories exist."
    
    client = LLMClient(
        provider=model.provider, api_base=model.api_base,
        api_key=model.api_key or "",
        model_name=model.model_name, temperature=0.1, max_tokens=512,
    )
    
    prompt = f"""You are a book search assistant. Given a user's natural language description, extract search criteria.

{cat_ctx}

User description: "{query_text}"

Output ONLY a JSON object with these fields (all optional, use null if not applicable):
{{
  "title": "<exact or partial title>",
  "author": "<author name>",
  "publisher": "<publisher name>",
  "isbn": "<ISBN>",
  "category_id": <category id or null>,
  "year_from": <year or null>,
  "year_to": <year or null>,
  "language": "<language or null>",
  "description": "<brief explanation of your interpretation>"
}}

Rules:
- Use partial match for title/author/publisher (no wildcards needed, backend adds %)
- If user says "about AI" or "related to machine learning", set title to "AI" or "machine learning"
- If user mentions a specific book name, use exact title
- If no clear criteria, return empty object {{}}
"""
    
    try:
        resp = await client.chat_completion([
            {"role": "system", "content": "You are a book search assistant. Output only valid JSON."},
            {"role": "user", "content": prompt},
        ])
        reply = resp["choices"][0]["message"]["content"].strip()
        
        # Extract JSON
        obj = None
        try:
            start = reply.index("{")
            depth = 0
            for i in range(start, len(reply)):
                if reply[i] == "{":
                    depth += 1
                elif reply[i] == "}":
                    depth -= 1
                    if depth == 0:
                        obj = json.loads(reply[start:i+1])
                        break
        except (ValueError, json.JSONDecodeError):
            pass
        
        if not obj:
            obj = {}
        
    except Exception as e:
        obj = {}
    
    # Build SQL query from LLM output
    query = select(Book)
    conditions = []
    
    if obj.get("title"):
        like = f"%{obj['title']}%"
        conditions.append(Book.title.ilike(like))
    if obj.get("author"):
        like = f"%{obj['author']}%"
        conditions.append(Book.authors.ilike(like))
    if obj.get("publisher"):
        like = f"%{obj['publisher']}%"
        conditions.append(Book.publisher.ilike(like))
    if obj.get("isbn"):
        conditions.append(Book.isbn == obj["isbn"])
    if obj.get("category_id") is not None:
        conditions.append(Book.category_id == obj["category_id"])
    if obj.get("year_from"):
        conditions.append(Book.publication_year >= obj["year_from"])
    if obj.get("year_to"):
        conditions.append(Book.publication_year <= obj["year_to"])
    if obj.get("language"):
        conditions.append(Book.language == obj["language"])
    
    # If LLM extracted no conditions, fall back to basic fuzzy search
    if not conditions:
        like = f"%{query_text}%"
        conditions.append(Book.title.ilike(like))
        conditions.append(Book.authors.ilike(like))
        conditions.append(Book.publisher.ilike(like))
        conditions.append(Book.isbn.ilike(like))
    
    query = query.where(or_(*conditions))
    query = query.order_by(Book.id.desc()).limit(100)
    items = (await db.execute(query)).scalars().all()
    
    result = [_book_dict(b) for b in items]
    return {
        "items": result,
        "total": len(result),
        "llm_used": True,
        "criteria": obj.get("description", ""),
    }


@router.get("/isbn-lookup")
async def isbn_lookup(isbn: str, db: AsyncSession = Depends(get_db)):
    """Look up book info by ISBN."""
    from app.services.isbn_lookup import lookup_isbn
    result = await lookup_isbn(isbn)
    if not result:
        raise HTTPException(404, "未找到该ISBN的图书信息")
    return result


@router.get("/template")
async def download_template():
    """Download Excel import template."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Book Import Template"
    
    headers = ["title", "authors", "publisher", "publication_year", "isbn",
               "category", "language", "pages", "price", "location",
               "total_copies", "call_number", "notes"]
    
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    # Example row
    example = ["Introduction to Algorithms", "Thomas Cormen", "MIT Press",
               2009, "978-0262033848", "Computer Science", "English",
               1312, 89.99, "A1-001", 3, "TP301.6/C813", "Third edition"]
    for col, v in enumerate(example, 1):
        ws.cell(row=2, column=col, value=v)
    
    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=book_import_template.xlsx"}
    )

@router.post("/import", status_code=201)
async def import_books(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Import books from Excel (.xlsx) file.
    
    Expected columns: title, authors, publisher, publication_year, isbn,
                      category (name), language, pages, price, location,
                      total_copies, notes
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的文件")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # Read header row
    headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]
    
    # Validate required columns
    if 'title' not in headers:
        raise HTTPException(400, "Excel 缺少 'title' 列")
    
    # Build category name -> id mapping
    cats_result = await db.execute(select(Category))
    cat_map = {c.name: c.id for c in cats_result.scalars().all()}
    
    imported = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue  # skip empty rows
        
        row_data = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = val
        
        title = row_data.get('title')
        if not title:
            errors.append(f"第 {row_idx} 行缺少 title")
            continue
        
        # Resolve category (accept both 'category' and 'category_name')
        cat_name = row_data.get('category') or row_data.get('category_name', '')
        if cat_name and cat_name not in cat_map:
            # Auto-create category if it doesn't exist
            new_cat = Category(name=cat_name)
            db.add(new_cat)
            await db.flush()
            cat_map[cat_name] = new_cat.id
        category_id = cat_map.get(cat_name) if cat_name else None
        
        year = row_data.get('publication_year')
        if year and isinstance(year, float):
            year = int(year)
        
        pages = row_data.get('pages')
        if pages and isinstance(pages, float):
            pages = int(pages)
        
        price = row_data.get('price')
        if price is not None:
            price = float(price)
        
        copies = row_data.get('total_copies')
        if copies and isinstance(copies, float):
            copies = int(copies)
        
        book = Book(
            title=str(title),
            call_number=str(row_data.get('call_number', '')) or None,
            authors=str(row_data.get('authors', '')) or None,
            publisher=str(row_data.get('publisher', '')) or None,
            publication_year=year,
            isbn=str(row_data.get('isbn', '')) or None,
            category_id=category_id,
            language=str(row_data.get('language', '')) or None,
            pages=pages,
            location=str(row_data.get('location', '')) or None,
            total_copies=copies or 1,
            available_copies=copies or 1,
            price=price,
            notes=str(row_data.get('notes', '')) or None,
        )
        db.add(book)
        await db.commit()
        await db.refresh(book)
        
        # Create copies
        total = book.total_copies or 1
        for i in range(total):
            copy = BookCopy(
                book_id=book.id,
                asset_number=f"{book.id}-{i+1:03d}",
                status="在馆",
            )
            db.add(copy)
        await db.commit()
        imported += 1
    
    return {
        "imported": imported,
        "errors": errors,
        "message": f"成功导入 {imported} 本图书" + (f"，{len(errors)} 个错误" if errors else "")
    }


@router.post("/import-ai", status_code=201)
async def import_books_ai(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Import books from Excel with LLM auto-classification.

    Expected columns: title, authors, publisher, publication_year, isbn,
                      language, pages, price, location, total_copies, notes
    Category and call_number are auto-generated by LLM.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的文件")

    # Check LLM model availability
    from app.models.model import LLMModel
    from app.services.model_service import LLMClient

    model = await db.get(LLMModel, 1)
    if not model:
        raise HTTPException(400, "请先在设置中配置 LLM 模型（ID=1）")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Read header row
    headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]

    # Validate required columns
    if 'title' not in headers:
        raise HTTPException(400, "Excel 缺少 'title' 列")

    # Build existing category context for LLM
    cats_result = await db.execute(select(Category).order_by(Category.id))
    cats = cats_result.scalars().all()
    if cats:
        cat_ctx = "Available categories:\n" + "\n".join(
            f"  [{c.id}] {c.name} (code: {c.code or '-'})" for c in cats
        )
    else:
        cat_ctx = "No categories exist yet. Suggest a new category name."

    client = LLMClient(
        provider=model.provider, api_base=model.api_base,
        api_key=model.api_key or "",
        model_name=model.model_name, temperature=0.1, max_tokens=512,
    )

    imported = 0
    errors = []
    classified = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        row_data = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = val

        title = row_data.get('title')
        if not title:
            errors.append({"row": row_idx, "error": "缺少 title"})
            continue

        authors = str(row_data.get('authors', '') or '')
        publisher = str(row_data.get('publisher', '') or '')

        # Call LLM to classify this book
        prompt = f"""You are a library book classifier. Classify the following book into the best category.

{cat_ctx}

Book information:
- Title: {title}
- Authors: {authors}
- Publisher: {publisher}

Output ONLY a JSON object (no explanation):
{{"category_id": <existing category id or null>, "category_name": "<suggested category name if new>", "call_number": "<CLC code / author mark>", "language": "<guessed language: Chinese/English/other>"}}"""

        category_id = None
        call_number = None
        language = None

        try:
            resp = await client.chat_completion([
                {"role": "system", "content": "You are a book classifier. Output only valid JSON."},
                {"role": "user", "content": prompt},
            ])
            reply = resp["choices"][0]["message"]["content"].strip()

            # Extract JSON from reply
            obj = None
            try:
                start = reply.index("{")
                depth = 0
                for i in range(start, len(reply)):
                    if reply[i] == "{":
                        depth += 1
                    elif reply[i] == "}":
                        depth -= 1
                        if depth == 0:
                            obj = json.loads(reply[start:i+1])
                            break
            except (ValueError, json.JSONDecodeError):
                pass

            if obj:
                cat_id = obj.get("category_id")
                if cat_id is not None:
                    # Verify category exists
                    existing = await db.get(Category, cat_id)
                    if existing:
                        category_id = cat_id
                call_number = obj.get("call_number")
                language = obj.get("language")
        except Exception as e:
            errors.append({"row": row_idx, "error": f"LLM 分类失败: {e}"})

        # Parse numeric fields
        year = row_data.get('publication_year')
        if year and isinstance(year, float):
            year = int(year)

        pages = row_data.get('pages')
        if pages and isinstance(pages, float):
            pages = int(pages)

        price = row_data.get('price')
        if price is not None:
            price = float(price)

        copies = row_data.get('total_copies')
        if copies and isinstance(copies, float):
            copies = int(copies)

        book = Book(
            title=str(title),
            call_number=str(call_number or '') or None,
            authors=authors or None,
            publisher=publisher or None,
            publication_year=year,
            isbn=str(row_data.get('isbn', '') or '') or None,
            category_id=category_id,
            language=str(language or row_data.get('language', '') or '') or None,
            pages=pages,
            location=str(row_data.get('location', '') or '') or None,
            total_copies=copies or 1,
            available_copies=copies or 1,
            price=price,
            notes=str(row_data.get('notes', '') or '') or None,
        )
        db.add(book)
        await db.commit()
        await db.refresh(book)

        # Create copies
        total = book.total_copies or 1
        for i in range(total):
            copy = BookCopy(
                book_id=book.id,
                asset_number=f"{book.id}-{i+1:03d}",
                status="在馆",
            )
            db.add(copy)
        await db.commit()

        classified.append({
            "row": row_idx,
            "title": str(title),
            "category_id": category_id,
            "call_number": call_number,
            "language": language,
        })
        imported += 1

    return {
        "imported": imported,
        "errors": errors,
        "classified": classified,
        "message": f"AI 分类并导入 {imported} 本图书" + (f"，{len(errors)} 个错误" if errors else "")
    }


# ─ Uncategorized Books & Batch Classification ──────────

@router.get("/uncategorized")
async def list_uncategorized_books(
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """List books that have no category assigned."""
    result = await db.execute(
        select(Book).where(Book.category_id.is_(None)).order_by(Book.id)
    )
    books = result.scalars().all()
    items = []
    for b in books:
        items.append({
            "id": b.id,
            "title": b.title,
            "authors": b.authors or "",
            "publisher": b.publisher or "",
            "isbn": b.isbn or "",
            "call_number": b.call_number or "",
        })
    return {"count": len(items), "books": items}


@router.post("/classify")
async def classify_uncategorized_books(
    db: AsyncSession = Depends(get_db),
    current_admin: dict = Depends(get_current_admin)
):
    """Batch classify uncategorized books using web search + LLM.

    For each uncategorized book:
    1. Search the web for book info
    2. Use LLM to determine best category match
    3. Update the book's category_id and call_number
    """
    from app.models.model import LLMModel
    from app.services.model_service import LLMClient
    from app.services.web_search import search_book_info

    model = await db.get(LLMModel, 1)
    if not model:
        raise HTTPException(400, "请先在设置中配置 LLM 模型（ID=1）")

    # Get uncategorized books
    result = await db.execute(
        select(Book).where(Book.category_id.is_(None)).order_by(Book.id)
    )
    books = result.scalars().all()

    if not books:
        return {"message": "没有未分类的图书", "classified": 0, "failed": 0, "results": []}

    # Build category context
    cats_result = await db.execute(select(Category).order_by(Category.id))
    cats = cats_result.scalars().all()
    if cats:
        cat_ctx = "Available categories:\n" + "\n".join(
            f"  [{c.id}] {c.name} (code: {c.code or '-'})" for c in cats
        )
    else:
        cat_ctx = "No categories exist yet. Suggest a new category name."

    client = LLMClient(
        provider=model.provider, api_base=model.api_base,
        api_key=model.api_key or "",
        model_name=model.model_name, temperature=0.1, max_tokens=1024,
    )

    classified = 0
    failed = 0
    results = []

    for book in books:
        # Step 1: Web search for book info
        web_info = await search_book_info(book.title, book.authors or "")

        # Step 2: LLM classification with web context
        prompt = f"""You are a professional librarian. Classify this book into the best matching category.

Available categories in the library:
{cat_ctx}

Book to classify:
- Title: {book.title}
- Authors: {book.authors or 'Unknown'}
- Publisher: {book.publisher or 'Unknown'}
- ISBN: {book.isbn or 'Unknown'}

Web search results about this book:
{web_info}

Analyze the web search results to understand the book's subject, genre, and academic field.
Then match it to the most appropriate category from the library's category list.

Output ONLY JSON:
{{"category_id": <id or null if no match>, "category_name": "<matched or suggested name>",
  "call_number": "<suggested call number>", "confidence": "<high/medium/low>",
  "reason": "<brief reason>"}}"""

        try:
            resp = await client.chat_completion([
                {"role": "system", "content": "You are a professional librarian. Output only JSON."},
                {"role": "user", "content": prompt},
            ])
            reply = resp["choices"][0]["message"]["content"].strip()

            # Extract JSON
            obj = None
            try:
                start = reply.index("{")
                depth = 0
                for i in range(start, len(reply)):
                    if reply[i] == "{":
                        depth += 1
                    elif reply[i] == "}":
                        depth -= 1
                        if depth == 0:
                            obj = json.loads(reply[start:i+1])
                            break
            except (ValueError, json.JSONDecodeError):
                pass

            if obj and obj.get("category_id") is not None:
                cat_id = obj["category_id"]
                existing = await db.get(Category, cat_id)
                if existing:
                    book.category_id = cat_id
                    if obj.get("call_number"):
                        book.call_number = obj["call_number"]
                    await db.commit()
                    classified += 1
                    results.append({
                        "id": book.id,
                        "title": book.title,
                        "category_id": cat_id,
                        "category_name": existing.name,
                        "call_number": obj.get("call_number", ""),
                        "confidence": obj.get("confidence", ""),
                        "reason": obj.get("reason", ""),
                    })
                else:
                    # Category ID doesn't exist, try to create by name
                    cat_name = obj.get("category_name", "")
                    if cat_name:
                        # Check if category with this name exists
                        cat_result = await db.execute(
                            select(Category).where(Category.name == cat_name)
                        )
                        existing_cat = cat_result.scalar_one_or_none()
                        if existing_cat:
                            book.category_id = existing_cat.id
                            if obj.get("call_number"):
                                book.call_number = obj["call_number"]
                            await db.commit()
                            classified += 1
                            results.append({
                                "id": book.id,
                                "title": book.title,
                                "category_id": existing_cat.id,
                                "category_name": existing_cat.name,
                                "call_number": obj.get("call_number", ""),
                                "confidence": obj.get("confidence", ""),
                                "reason": obj.get("reason", ""),
                            })
                        else:
                            # Create new category
                            new_cat = Category(name=cat_name, code=f"CAT-{cat_name[:4].upper()}")
                            db.add(new_cat)
                            await db.commit()
                            await db.refresh(new_cat)
                            book.category_id = new_cat.id
                            if obj.get("call_number"):
                                book.call_number = obj["call_number"]
                            await db.commit()
                            classified += 1
                            results.append({
                                "id": book.id,
                                "title": book.title,
                                "category_id": new_cat.id,
                                "category_name": new_cat.name,
                                "call_number": obj.get("call_number", ""),
                                "confidence": obj.get("confidence", ""),
                                "reason": obj.get("reason", ""),
                            })
                    else:
                        failed += 1
                        results.append({"id": book.id, "title": book.title, "error": f"Category #{cat_id} not found and no name provided"})
            else:
                # No category_id, try to create by category_name
                cat_name = obj.get("category_name", "") if obj else ""
                if cat_name:
                    cat_result = await db.execute(
                        select(Category).where(Category.name == cat_name)
                    )
                    existing_cat = cat_result.scalar_one_or_none()
                    if existing_cat:
                        book.category_id = existing_cat.id
                        if obj.get("call_number"):
                            book.call_number = obj["call_number"]
                        await db.commit()
                        classified += 1
                        results.append({
                            "id": book.id,
                            "title": book.title,
                            "category_id": existing_cat.id,
                            "category_name": existing_cat.name,
                            "call_number": obj.get("call_number", ""),
                            "confidence": obj.get("confidence", ""),
                            "reason": obj.get("reason", ""),
                        })
                    else:
                        # Create new category
                        new_cat = Category(name=cat_name, code=f"CAT-{cat_name[:4].upper()}")
                        db.add(new_cat)
                        await db.commit()
                        await db.refresh(new_cat)
                        book.category_id = new_cat.id
                        if obj.get("call_number"):
                            book.call_number = obj["call_number"]
                        await db.commit()
                        classified += 1
                        results.append({
                            "id": book.id,
                            "title": book.title,
                            "category_id": new_cat.id,
                            "category_name": new_cat.name,
                            "call_number": obj.get("call_number", ""),
                            "confidence": obj.get("confidence", ""),
                            "reason": obj.get("reason", ""),
                        })
                else:
                    failed += 1
                    results.append({"id": book.id, "title": book.title, "error": "LLM did not return a valid category"})

        except Exception as e:
            failed += 1
            results.append({"id": book.id, "title": book.title, "error": str(e)})

    return {
        "classified": classified,
        "failed": failed,
        "total": len(books),
        "results": results,
        "message": f"分类完成：{classified} 本成功，{failed} 本失败，共 {len(books)} 本"
    }
@router.get("/{book_id}")
async def get_book(book_id: int, db: AsyncSession = Depends(get_db)):
    b = await db.get(Book, book_id)
    if not b:
        raise HTTPException(404, "图书不存在")
    cat_name = ""
    if b.category_id:
        cat_obj = await db.get(Category, b.category_id)
        cat_name = cat_obj.name if cat_obj else ""
    return _book_dict(b, cat_name)


@router.post("", status_code=201)
async def create_book(data: BookCreate, request: Request, db: AsyncSession = Depends(get_db), current_admin: dict = Depends(get_current_admin)):
    book = Book(**data.model_dump())
    db.add(book)
    await db.commit()
    await db.refresh(book)
    # Auto-create copies
    total = book.total_copies or 1
    from app.models.model import BookCopy
    for i in range(total):
        copy = BookCopy(book_id=book.id, asset_number=f"{book.id}-{i+1:03d}", status="在馆")
        db.add(copy)
    await db.commit()

    # Log audit
    await log_audit(
        db=db,
        user_id=None,  # TODO: get from auth token
        user_type="admin",
        action="create",
        resource_type="book",
        resource_id=book.id,
        details={"title": book.title, "isbn": book.isbn},
        ip_address=request.client.host if request.client else None
    )

    return {"id": book.id, "message": "添加成功"}


@router.put("/{book_id}")
async def update_book(book_id: int, data: BookUpdate, request: Request, db: AsyncSession = Depends(get_db), current_admin: dict = Depends(get_current_admin)):
    book = await db.get(Book, book_id)
    if not book:
        raise HTTPException(404, "图书不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(book, k, v)
    await db.commit()

    # Log audit
    await log_audit(
        db=db,
        user_id=None,  # TODO: get from auth token
        user_type="admin",
        action="update",
        resource_type="book",
        resource_id=book.id,
        details={"title": book.title, "changes": data.model_dump(exclude_unset=True)},
        ip_address=request.client.host if request.client else None
    )

    return {"message": "更新成功"}


@router.delete("/{book_id}", status_code=204)
async def delete_book(book_id: int, request: Request, db: AsyncSession = Depends(get_db), current_admin: dict = Depends(get_current_admin)):
    book = await db.get(Book, book_id)
    if not book:
        raise HTTPException(404, "图书不存在")

    # Log audit before deletion
    await log_audit(
        db=db,
        user_id=None,  # TODO: get from auth token
        user_type="admin",
        action="delete",
        resource_type="book",
        resource_id=book.id,
        details={"title": book.title, "isbn": book.isbn},
        ip_address=request.client.host if request.client else None
    )

    await db.delete(book)
    await db.commit()


class BatchDeleteRequest(BaseModel):
    book_ids: list[int]


@router.post("/batch-delete", status_code=204)
async def batch_delete_books(body: BatchDeleteRequest, db: AsyncSession = Depends(get_db)):
    """Delete multiple books by their IDs."""
    if not body.book_ids:
        raise HTTPException(400, "请提供要删除的图书ID列表")

    books = await db.execute(select(Book).where(Book.id.in_(body.book_ids)))
    books_to_delete = books.scalars().all()

    if not books_to_delete:
        raise HTTPException(404, "未找到要删除的图书")

    for book in books_to_delete:
        await db.delete(book)
    await db.commit()




