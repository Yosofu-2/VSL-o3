# LitManager Modern GUI - Inspired by TRAE Interface
import sys
import os
import json
import threading
import tkinter as tk
import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
from typing import Optional
from api_client import APIClient
from i18n import tr, set_language, get_language
from gui_utils import ToastNotification, LoadingOverlay, add_hover_effect
from statistics_view import StatisticsView

ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

API = APIClient("http://127.0.0.1:8000")

# TRAE-Inspired Modern Color Scheme
C_SIDEBAR = "#f5f5f7"       # Light gray sidebar (TRAE style)
C_SIDEBAR_HOVER = "#e8e8ed" # Subtle hover
C_SIDEBAR_ACTIVE = "#d4d4db"# Active state (gray)
C_SIDEBAR_TEXT = "#1d1d1f"  # Dark text for sidebar
C_HEADER = "#ffffff"        # Pure white header
C_CONTENT = "#ffffff"       # Pure white content bg
C_CARD = "#ffffff"          # White card
C_ACCENT = "#d4d4db"        # Gray accent (matches sidebar active)
C_ACCENT_HOVER = "#c0c0c8"  # Darker gray on hover (click animation)
C_SUCCESS = "#a8d5ba"       # Soft green
C_SUCCESS_HOVER = "#96c4a8" # Darker green hover
C_WARN = "#f5d76e"          # Soft yellow
C_DANGER = "#f4a0a0"        # Soft red
C_DANGER_HOVER = "#e89090"  # Darker red hover
C_TEXT = "#1d1d1f"          # Near black text
C_TEXT_SEC = "#86868b"      # Gray secondary text
C_BORDER = "#e5e5e7"        # Very light border
C_INPUT_BG = "#fafafa"      # Input background
C_BUTTON_BG = "#ffffff"     # Button background
C_BUTTON_BORDER = "#d2d2d7" # Button border


class ModernApp(ctk.CTk):
    """Main application window with modern TRAE-style interface."""

    def __init__(self, admin_info=None):
        super().__init__()
        self.title("LitManager")
        self.geometry("1200x760")
        self.minsize(1000, 600)

        self.admin_info = admin_info or {}

        self._build_layout()

    def _build_layout(self):
        """Build the main layout with sidebar, header, and content area."""
        
        # ── Sidebar ───────────────────────────────────────
        self.sidebar = ctk.CTkFrame(self, fg_color=C_SIDEBAR, width=200, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        
        self._build_sidebar()
        
        # ── Main Content Area ─────────────────────────────
        self.main_area = ctk.CTkFrame(self, fg_color=C_CONTENT)
        self.main_area.pack(side="left", fill="both", expand=True)
        
        # Header
        self.header = ctk.CTkFrame(self.main_area, fg_color=C_HEADER, height=60, corner_radius=0)
        self.header.pack(side="top", fill="x")
        self.header.pack_propagate(False)
        self._build_header()
        
        # Content Frame
        self.content_frame = ctk.CTkFrame(self.main_area, fg_color=C_CONTENT)
        self.content_frame.pack(side="top", fill="both", expand=True)
        
        # ── Views ─────────────────────────────────────────
        self.views = {}
        self.current_view = None
        
        # Initialize views
        self._init_views()
        
        # Defer initial view switch to after mainloop starts
        self.after(100, lambda: self.switch_view("dashboard"))

    def _build_sidebar(self):
        """Build the navigation sidebar."""
        
        # Logo/Title
        logo_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent", height=60)
        logo_frame.pack(fill="x", pady=(16, 8))
        logo_frame.pack_propagate(False)
        
        logo_label = ctk.CTkLabel(
            logo_frame, 
            text="LitManager", 
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=C_SIDEBAR_TEXT
        )
        logo_label.place(relx=0.5, rely=0.5, anchor="center")
        
        # Navigation Items
        nav_items = [
            ("book_search", tr("Book Search"), "🔍"),
            ("dashboard", tr("Dashboard"), ""),
            ("statistics", tr("Statistics"), "📈"),
            ("books", tr("Books"), ""),
            ("readers", tr("Readers"), "👥"),
            ("borrowing", tr("Borrowing"), "📖"),
            ("audit_log", "审计日志", "📋"),
            ("models", tr("Models"), "🤖"),
            ("ai", tr("AI Assistant"), "💬"),
            ("settings", tr("Settings"), "⚙️"),
        ]
        
        self.nav_buttons = []
        for view_id, label, icon in nav_items:
            btn = ctk.CTkButton(
                self.sidebar,
                text=f"{icon}  {label}",
                font=ctk.CTkFont(size=13),
                fg_color="transparent",
                text_color=C_SIDEBAR_TEXT,
                hover_color=C_SIDEBAR_HOVER,
                anchor="w",
                height=44,
                corner_radius=8,
                command=lambda vid=view_id: self.switch_view(vid)
            )
            btn.pack(fill="x", padx=8, pady=2)
            self.nav_buttons.append((view_id, btn))
            
            # Add hover effect for visual feedback
            add_hover_effect(btn, hover_color=C_SIDEBAR_HOVER, original_color="transparent")
        
        # Bottom section
        bottom_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        bottom_frame.pack(side="bottom", fill="x", pady=(0, 16))
        
        ctk.CTkButton(
            bottom_frame,
            text=f"❓ {tr('Help')}",
            font=ctk.CTkFont(size=12),
            fg_color="transparent",
            text_color=C_SIDEBAR_TEXT,
            hover_color=C_SIDEBAR_HOVER,
            anchor="w",
            height=36,
            corner_radius=8,
            command=self._show_help
        ).pack(fill="x")

    def _build_header(self):
        """Build the top header bar - 无搜索框，仅标题 + 通知铃铛 + 头像"""
        
        # Center: Page Title
        self.page_title = ctk.CTkLabel(
            self.header,
            text=tr("Dashboard"),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=C_TEXT
        )
        self.page_title.place(relx=0.5, rely=0.5, anchor="center")
        
        # Right: Actions
        actions_frame = ctk.CTkFrame(self.header, fg_color="transparent")
        actions_frame.pack(side="right", fill="y", padx=16, pady=8)
        
        # Notification bell with badge
        self.notif_btn = ctk.CTkButton(
            actions_frame,
            text="🔔",
            font=ctk.CTkFont(size=18),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            width=40,
            height=40,
            corner_radius=12,
            command=self._show_notifications
        )
        self.notif_btn.pack(side="right", padx=(0, 8))
        
        self.notif_badge = ctk.CTkLabel(
            actions_frame, text="",
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color="#ffffff", fg_color="#ef4444",
            width=18, height=18, corner_radius=9
        )
        self.notif_badge.place(relx=0.0, rely=0.0, anchor="ne", x=-4, y=4)
        self.notif_badge.lower(self.notif_btn)
        
        # User avatar
        ctk.CTkButton(
            actions_frame,
            text="👤",
            font=ctk.CTkFont(size=18),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=40,
            height=40,
            corner_radius=12,
            command=self._show_user_menu
        ).pack(side="right")

        # Load unread count (defer to after mainloop starts)
        self.after(50, self._refresh_notif_badge)

    def _init_views(self):
        """Initialize all views."""
        self.views["book_search"] = BookSearchView(self.content_frame)
        self.views["dashboard"] = DashboardView(self.content_frame)
        self.views["statistics"] = StatisticsView(self.content_frame)
        self.views["books"] = BooksView(self.content_frame)
        self.views["excel_import"] = ExcelImportView(self.content_frame)
        self.views["readers"] = ReadersView(self.content_frame)
        self.views["reader_import"] = ReaderImportView(self.content_frame)
        self.views["borrowing"] = BorrowingView(self.content_frame)
        self.views["audit_log"] = AuditLogView(self.content_frame)
        self.views["models"] = ModelsView(self.content_frame)
        self.views["model_edit"] = ModelEditView(self.content_frame)
        self.views["ai"] = AIAssistantView(self.content_frame)
        self.views["settings"] = SettingsView(self.content_frame, app_ref=self)
        self.views["user_profile"] = UserProfileView(self.content_frame, reader_id=1)

    def switch_view(self, view_id, **kwargs):
        """Switch to a different view with smooth transition."""
        # Hide current view
        if self.current_view:
            self.current_view.pack_forget()
        
        # Show new view
        self.current_view = self.views.get(view_id)
        if self.current_view:
            self.current_view.pack(fill="both", expand=True)
            
            # Smooth fade-in effect: raise widget after brief delay
            self.after(10, lambda: self.current_view.tkraise() if self.current_view else None)
            
            # Defer refresh to after mainloop starts to avoid "main thread is not in main loop"
            def _do_refresh():
                if self.current_view:
                    if kwargs:
                        self.current_view.refresh(**kwargs)
                    else:
                        self.current_view.refresh()
            self.after(50, _do_refresh)
            
            # Update page title
            titles = {
                "book_search": tr("Book Search"),
                "dashboard": tr("Dashboard"),
                "statistics": tr("Statistics"),
                "books": tr("Books Management"),
                "excel_import": tr("Excel Import"),
                "readers": tr("Readers Management"),
                "reader_import": tr("Reader Import"),
                "borrowing": tr("Borrowing Records"),
                "audit_log": "审计日志",
                "models": tr("LLM Models"),
                "model_edit": tr("Add Model"),
                "ai": tr("AI Assistant"),
                "settings": tr("Settings"),
                "user_profile": tr("User Profile")
            }
            self.page_title.configure(text=titles.get(view_id, view_id))
        
        # Update nav button states
        for vid, btn in self.nav_buttons:
            if vid == view_id:
                btn.configure(fg_color=C_SIDEBAR_ACTIVE, text_color=C_SIDEBAR_TEXT)
            else:
                btn.configure(fg_color="transparent", text_color=C_SIDEBAR_TEXT)

    def _show_settings(self):
        self.switch_view("settings")

    def _show_help(self):
        messagebox.showinfo(tr("Help"), tr("Help documentation coming soon!"))

    def _refresh_notif_badge(self):
        """Fetch unread notification count and update badge."""
        def _fetch():
            try:
                uid = self.admin_info.get("id", 1)
                data = API.get_unread_notification_count(uid, "admin")
                count = data.get("count", 0) if isinstance(data, dict) else 0
                self.after(0, lambda: self._update_badge(count))
            except Exception:
                self.after(0, lambda: self._update_badge(0))

        threading.Thread(target=_fetch, daemon=True).start()

    def _update_badge(self, count):
        """Update notification badge display."""
        if count > 0:
            self.notif_badge.configure(text=str(count) if count < 100 else "99+")
            self.notif_badge.lift(self.notif_btn)
        else:
            self.notif_badge.configure(text="")
            self.notif_badge.lower(self.notif_btn)

    def _show_notifications(self):
        """Show notifications panel."""
        try:
            uid = self.admin_info.get("id", 1)
            data = API.list_notifications(uid, "admin", page_size=20)
            notifs = data.get("data", []) if isinstance(data, dict) else []
            if not notifs:
                messagebox.showinfo(tr("Notifications"), tr("No new notifications."))
                return
            msg = "\n".join(
                f"• {n.get('title', '')}: {n.get('content', '')}"
                for n in notifs[:10]
            )
            messagebox.showinfo(tr("Notifications"), msg)
            # Mark all as read
            try:
                API.mark_all_notifications_read(uid, "admin")
            except Exception:
                pass
            self._refresh_notif_badge()
        except Exception:
            messagebox.showinfo(tr("Notifications"), tr("No new notifications."))

    def _show_user_menu(self):
        """Navigate to user profile page."""
        self.switch_view("user_profile")

    def rebuild_all_views(self):
        """Destroy and recreate all views and sidebar to reflect language changes."""
        # Save current view
        current_view = None
        for vid, v in self.views.items():
            if v == self.current_view:
                current_view = vid
                break

        # Destroy sidebar
        try:
            self.sidebar.destroy()
        except Exception:
            pass

        # Destroy main area (which contains header and content_frame)
        try:
            self.main_area.destroy()
        except Exception:
            pass

        # Recreate sidebar FIRST so it stays on the left
        self.sidebar = ctk.CTkFrame(self, fg_color=C_SIDEBAR, width=200, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        # Recreate main area (which contains header and content_frame)
        self.main_area = ctk.CTkFrame(self, fg_color=C_CONTENT)
        self.main_area.pack(side="left", fill="both", expand=True)

        # Recreate header
        self.header = ctk.CTkFrame(self.main_area, fg_color=C_HEADER, height=60, corner_radius=0)
        self.header.pack(side="top", fill="x")
        self.header.pack_propagate(False)
        self._build_header()

        # Recreate content frame
        self.content_frame = ctk.CTkFrame(self.main_area, fg_color=C_CONTENT)
        self.content_frame.pack(side="top", fill="both", expand=True)

        # Recreate views
        self.views = {}
        self._init_views()

        # Restore current view
        if current_view:
            self.switch_view(current_view)
        else:
            self.switch_view("dashboard")


class StatCard(ctk.CTkFrame):
    """Modern statistics card."""
    
    def __init__(self, master, title, value, icon="📊", color=C_ACCENT, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        
        # Top accent strip
        accent_frame = ctk.CTkFrame(self, fg_color=color, height=4, corner_radius=0)
        accent_frame.pack(fill="x")
        
        # Content
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20, pady=16)
        
        # Icon and value
        top_row = ctk.CTkFrame(content, fg_color="transparent")
        top_row.pack(fill="x")
        
        ctk.CTkLabel(
            top_row,
            text=icon,
            font=ctk.CTkFont(size=28)
        ).pack(side="left")
        
        self.value_label = ctk.CTkLabel(
            top_row,
            text=str(value),
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=C_TEXT
        )
        self.value_label.pack(side="right")
        
        # Title
        ctk.CTkLabel(
            content,
            text=title,
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        ).pack(anchor="w", pady=(8, 0))
    
    def set_value(self, value):
        """Update the displayed value."""
        self.value_label.configure(text=str(value))


class DashboardView(ctk.CTkFrame):
    """Dashboard view with statistics cards and quick actions."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self._build()
    
    def _build(self):
        scroll_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True)
        
        # Stats Grid
        stats_grid = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        stats_grid.pack(fill="x", padx=24, pady=(24, 16))
        
        # Configure grid
        stats_grid.grid_columnconfigure(0, weight=1)
        stats_grid.grid_columnconfigure(1, weight=1)
        stats_grid.grid_columnconfigure(2, weight=1)
        stats_grid.grid_columnconfigure(3, weight=1)
        
        self.card_books = StatCard(stats_grid, tr("Total Books"), "—", "📚", C_ACCENT)
        self.card_books.grid(row=0, column=0, padx=8, pady=8, sticky="nsew")
        
        self.card_readers = StatCard(stats_grid, tr("Total Readers"), "—", "👥", C_SUCCESS)
        self.card_readers.grid(row=0, column=1, padx=8, pady=8, sticky="nsew")
        
        self.card_borrowed = StatCard(stats_grid, tr("Borrowed"), "—", "📖", C_WARN)
        self.card_borrowed.grid(row=0, column=2, padx=8, pady=8, sticky="nsew")
        
        self.card_available = StatCard(stats_grid, tr("Available"), "—", "✅", C_ACCENT)
        self.card_available.grid(row=0, column=3, padx=8, pady=8, sticky="nsew")
        
        # Quick Actions Section
        actions_section = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        actions_section.pack(fill="x", padx=24, pady=16)
        
        ctk.CTkLabel(
            actions_section,
            text=tr("Quick Actions"),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=C_TEXT
        ).pack(anchor="w", pady=(0, 12))
        
        actions_grid = ctk.CTkFrame(actions_section, fg_color="transparent")
        actions_grid.pack(fill="x")
        
        actions_grid.grid_columnconfigure(0, weight=1)
        actions_grid.grid_columnconfigure(1, weight=1)
        actions_grid.grid_columnconfigure(2, weight=1)
        
        actions = [
            (tr("Add Book"), "📚", C_ACCENT, lambda: self._navigate("books")),
            (tr("Add Reader"), "👤", C_SUCCESS, lambda: self._navigate("readers")),
            (tr("New Borrow"), "📖", C_WARN, lambda: self._navigate("borrowing")),
        ]
        
        for i, (label, icon, color, cmd) in enumerate(actions):
            btn = ctk.CTkButton(
                actions_grid,
                text=f"{icon} {label}",
                font=ctk.CTkFont(size=14, weight="bold"),
                fg_color=C_CARD,
                text_color=C_TEXT,
                hover_color=C_INPUT_BG,
                border_width=1,
                border_color=C_BORDER,
                height=56,
                corner_radius=12,
                command=cmd
            )
            btn.grid(row=0, column=i, padx=8, sticky="nsew")
        
        # Recent Activity
        activity_section = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        activity_section.pack(fill="x", padx=24, pady=16)
        
        ctk.CTkLabel(
            activity_section,
            text=tr("Recent Activity"),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=C_TEXT
        ).pack(anchor="w", pady=(0, 12))
        
        self.activity_list = ctk.CTkFrame(activity_section, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        self.activity_list.pack(fill="both", expand=True)
        
        # Placeholder content
        placeholder = ctk.CTkLabel(
            self.activity_list,
            text=tr("Loading recent activity..."),
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        )
        placeholder.pack(pady=24)
        self._placeholder = placeholder

    def _navigate(self, view_id):
        parent = self.master
        while parent and not hasattr(parent, "switch_view"):
            parent = parent.master
        if parent:
            parent.switch_view(view_id)

    def refresh(self):
        """Refresh dashboard data from API (non-blocking)."""
        def _do_refresh():
            try:
                stats = API.get_book_stats()
                readers = API.list_readers()
                borrowing = API.list_borrowing()
                
                borrow_list = borrowing.get("data", [])
                borrowed = sum(1 for b in borrow_list if b.get("status") == "借出")
                
                # Update UI on main thread
                self.after(0, lambda: self._update_dashboard(stats, readers, borrowed, borrow_list))
            except Exception as e:
                print(f"Dashboard refresh error: {e}")

        threading.Thread(target=_do_refresh, daemon=True).start()

    def _update_dashboard(self, stats, readers, borrowed, borrow_list):
        """Update dashboard UI elements (must be called on main thread)."""
        self.card_books.set_value(stats.get("total_copies", 0))
        self.card_readers.set_value(readers.get("total", 0))
        self.card_borrowed.set_value(borrowed)
        self.card_available.set_value(stats.get("available_copies", 0))
        self._refresh_activity(borrow_list)
    
    def _refresh_activity(self, borrow_list):
        """Update the recent activity list with borrowing records."""
        # Clear existing activity items (keep the frame)
        for w in self.activity_list.winfo_children():
            w.destroy()
        
        recent = borrow_list[:8]
        
        if not recent:
            ctk.CTkLabel(
                self.activity_list,
                text=tr("No recent activity"),
                font=ctk.CTkFont(size=13),
                text_color=C_TEXT_SEC
            ).pack(pady=24)
            return
        
        for rec in recent:
            reader_name = rec.get("reader_name", tr("Unknown"))
            book_title = rec.get("book_title", tr("Unknown"))
            status = rec.get("status", "")
            borrow_date = rec.get("borrow_date", "")
            due_date = rec.get("due_date", "")
            
            # Status icon
            if status == "借出":
                status_icon = ""
            elif status == "已归还":
                status_icon = "✅"
            elif status == "逾期":
                status_icon = "️"
            else:
                status_icon = "📋"
            
            # Build activity line
            date_str = borrow_date or ""
            if due_date and status == "借出":
                date_str = f"{borrow_date} / {tr('Due')} {due_date}"
            
            line = f"{status_icon}  {reader_name}  —  {book_title}"
            sub_line = f"      {tr('Borrow Date')}: {date_str}    {tr('Status')}: {status}"
            
            ctk.CTkLabel(
                self.activity_list,
                text=line,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color=C_TEXT,
                anchor="w"
            ).pack(fill="x", padx=16, pady=(8, 0))
            ctk.CTkLabel(
                self.activity_list,
                text=sub_line,
                font=ctk.CTkFont(size=12),
                text_color=C_TEXT_SEC,
                anchor="w"
            ).pack(fill="x", padx=16, pady=(0, 8))


class BookSearchView(ctk.CTkFrame):
    """Book search view with AI-powered natural language search."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.books = []
        self._build()
    
    def _build(self):
        # Title
        ctk.CTkLabel(
            self,
            text=tr("Book Search"),
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=C_TEXT
        ).pack(pady=(24, 8))
        
        # Description
        ctk.CTkLabel(
            self,
            text=tr("Search books with AI"),
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        ).pack(pady=(0, 20))
        
        # Search bar
        search_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        search_frame.pack(fill="x", padx=24, pady=10)
        
        search_inner = ctk.CTkFrame(search_frame, fg_color="transparent")
        search_inner.pack(fill="x", padx=20, pady=16)
        
        self.search_entry = ctk.CTkEntry(
            search_inner,
            placeholder_text=tr("Enter book name, author, topic, or description..."),
            font=ctk.CTkFont(size=14),
            fg_color=C_INPUT_BG,
            border_width=0,
            corner_radius=10,
            height=40
        )
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(0, 12))
        self.search_entry.bind("<Return>", lambda e: self._do_search())
        
        self.search_btn = ctk.CTkButton(
            search_inner,
            text=tr("Search"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=100,
            height=40,
            corner_radius=10,
            command=self._do_search
        )
        self.search_btn.pack(side="right")
        
        # Status bar
        self.status_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        )
        self.status_label.pack(pady=(10, 0))
        
        # Results table
        table_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=16)
        
        columns = ("id", "title", "authors", "publisher", "year", "isbn", "copies", "available")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        headings = {"id": "ID", "title": tr("Title"), "authors": tr("Author"),
                    "publisher": tr("Publisher"), "year": tr("Year"), "isbn": "ISBN",
                    "copies": tr("Total"), "available": tr("Avail")}
        widths = {"id": 60, "title": 280, "authors": 150, "publisher": 140,
                  "year": 60, "isbn": 120, "copies": 60, "available": 60}
        
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col])
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
    
    def refresh(self):
        """Reset the view when navigated to."""
        pass
    
    def _do_search(self):
        query = self.search_entry.get().strip()
        if not query:
            return
        
        self.search_btn.configure(state="disabled", text=tr("Searching..."))
        self.status_label.configure(text=tr("Searching..."), text_color=C_TEXT_SEC)
        
        # Show loading overlay
        self.loading = LoadingOverlay(self, tr("Searching..."))
        self.loading.show()
        
        threading.Thread(target=self._search_thread, args=(query,), daemon=True).start()
    
    def _search_thread(self, query):
        try:
            result = API.search_books_ai(query)
            self.books = result.get("items", [])
            total = result.get("total", 0)
            llm_used = result.get("llm_used", False)
            criteria = result.get("criteria", "")
            
            self.after(0, lambda: self._show_results(total, llm_used, criteria))
        except Exception as e:
            self.after(0, lambda: self._show_error(str(e)))
        finally:
            self.after(0, self._hide_loading)
    
    def _hide_loading(self):
        """Hide loading overlay."""
        if hasattr(self, 'loading'):
            self.loading.hide()
    
    def _show_results(self, total, llm_used, criteria):
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Insert results
        for b in self.books:
            self.tree.insert("", "end", values=(
                b.get("id", ""),
                b.get("title", ""),
                b.get("authors", ""),
                b.get("publisher", ""),
                b.get("publication_year", ""),
                b.get("isbn", ""),
                b.get("total_copies", 0),
                b.get("available_copies", 0),
            ))
        
        # Update status
        status_text = f"{total} {tr('results found')}"
        if llm_used and criteria:
            status_text += f" | {tr('AI interpretation')}: {criteria}"
        self.status_label.configure(text=status_text, text_color=C_SUCCESS)
        
        self.search_btn.configure(state="normal", text=tr("Search"))
    
    def _show_error(self, error):
        self.status_label.configure(text=f"{tr('Error')}: {error}", text_color="#ef4444")
        self.search_btn.configure(state="normal", text=tr("Search"))


class BooksView(ctk.CTkFrame):
    """Books management view with search and table."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.books = []
        self.selected_ids = set()  # Track selected book IDs
        self._build()
    
    def _build(self):
        # Toolbar
        toolbar = ctk.CTkFrame(self, fg_color="transparent")
        toolbar.pack(fill="x", padx=24, pady=(20, 0))
        
        # Search
        search_frame = ctk.CTkFrame(toolbar, fg_color=C_INPUT_BG, corner_radius=10)
        search_frame.pack(side="left", fill="y")
        
        self.search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text=tr("Search books..."),
            fg_color="transparent",
            border_width=0,
            font=ctk.CTkFont(size=13),
            width=300
        )
        self.search_entry.pack(side="left", fill="y", padx=12)
        self.search_entry.bind("<Return>", lambda e: self.refresh())
        
        # Selection counter (hidden by default)
        self.selection_label = ctk.CTkLabel(
            toolbar,
            text="",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_ACCENT
        )
        self.selection_label.pack(side="left", padx=16)
        
        # Actions
        ctk.CTkButton(
            toolbar,
            text=tr("+ Add Book"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._add_book
        ).pack(side="right", padx=(0, 8))

        self.classify_btn = ctk.CTkButton(
            toolbar,
            text=tr("Classify Uncategorized"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#f59e0b",
            hover_color="#d97706",
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._classify_uncategorized
        )
        self.classify_btn.pack(side="right", padx=(0, 8))

        # Batch Delete button
        self.batch_delete_btn = ctk.CTkButton(
            toolbar,
            text=tr("Batch Delete"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_DANGER,
            hover_color=C_DANGER_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._batch_delete,
            state="disabled"
        )
        self.batch_delete_btn.pack(side="right", padx=(0, 8))

        # Export button
        ctk.CTkButton(
            toolbar,
            text=tr("Export"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_SUCCESS,
            hover_color=C_SUCCESS_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._export_books
        ).pack(side="right", padx=(0, 8))

        # Edit button
        ctk.CTkButton(
            toolbar,
            text=tr("Edit"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._edit_book
        ).pack(side="right", padx=(0, 8))
        
        # Table Frame
        table_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=(16, 24))
        
        # Treeview with checkbox column
        columns = ("selected", "id", "title", "authors", "publisher", "year", "isbn", "copies", "available")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        headings = {
            "selected": "☐",
            "id": "ID",
            "title": tr("Title"),
            "authors": tr("Author"),
            "publisher": tr("Publisher"),
            "year": tr("Year"),
            "isbn": "ISBN",
            "copies": tr("Total"),
            "available": tr("Avail")
        }
        widths = {
            "selected": 40,
            "id": 60,
            "title": 280,
            "authors": 150,
            "publisher": 140,
            "year": 60,
            "isbn": 120,
            "copies": 60,
            "available": 60
        }
        
        for col in columns:
            if col == "selected":
                # Checkbox column - click to toggle
                self.tree.heading(col, text=headings[col], command=lambda: self._toggle_select_all())
            else:
                self.tree.heading(col, text=headings[col], command=lambda c=col: self._sort(c))
            self.tree.column(col, width=widths[col], anchor="center" if col == "selected" else "w")
        
        # Scrollbar
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        
        # Bind click event for checkbox toggle
        self.tree.bind("<Button-1>", self._on_tree_click)
        
        # Pagination
        pagination = ctk.CTkFrame(self, fg_color="transparent")
        pagination.pack(fill="x", padx=24, pady=(0, 24))
        
        self.page_label = ctk.CTkLabel(pagination, text=tr("Page 1"), font=ctk.CTkFont(size=12))
        self.page_label.pack(side="right")
        
        ctk.CTkButton(
            pagination,
            text=tr("Next"),
            font=ctk.CTkFont(size=12),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            width=60,
            height=28,
            corner_radius=8,
            command=self._next_page
        ).pack(side="right", padx=4)
        
        ctk.CTkButton(
            pagination,
            text=tr("Prev"),
            font=ctk.CTkFont(size=12),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            width=60,
            height=28,
            corner_radius=8,
            command=self._prev_page
        ).pack(side="right", padx=4)
        
        self.page = 1
        self.total = 0
        self._sort_col = None
        self._sort_rev = False
        self._select_all = False  # Track select all state
    
    def _sort(self, col):
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = False
        self._refresh_tree()
    
    def _next_page(self):
        if self.page * 20 < self.total:
            self.page += 1
            self.refresh()
    
    def _prev_page(self):
        if self.page > 1:
            self.page -= 1
            self.refresh()
    
    def refresh(self):
        try:
            q = self.search_entry.get().strip() or None
            data = API.list_books(q=q, page=self.page, page_size=20)
            self.books = data.get("items", [])
            self.total = data.get("total", 0)
            self._refresh_tree()
            max_p = max(1, (self.total + 19) // 20)
            self.page_label.configure(text=f"{tr('Page')} {self.page}/{max_p} ({self.total} {tr('total')})")
        except Exception as e:
            print(f"Books refresh error: {e}")
    
    def _refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        items = list(self.books)
        if self._sort_col:
            items.sort(key=lambda b: (b.get(self._sort_col) or ""), reverse=self._sort_rev)
        for b in items:
            book_id = b.get("id", "")
            is_selected = "☑" if book_id in self.selected_ids else "☐"
            self.tree.insert("", "end", values=(
                is_selected,
                book_id,
                b.get("title", "")[:60],
                (b.get("authors", "") or "")[:30],
                (b.get("publisher", "") or "")[:25],
                b.get("publication_year", "") or "",
                b.get("isbn", "") or "",
                b.get("total_copies", 0),
                b.get("available_copies", 0),
            ))
        self._update_selection_ui()
    
    def _on_tree_click(self, event):
        """Handle click on tree to toggle checkbox."""
        # Get the item under the cursor
        item = self.tree.identify_row(event.y)
        if not item:
            return
        
        # Check if click is on the checkbox column (first column)
        column = self.tree.identify_column(event.x)
        if column != "#1":  # Not the checkbox column
            return
        
        # Toggle selection
        values = self.tree.item(item)["values"]
        book_id = values[1]  # ID is in second column
        
        if book_id in self.selected_ids:
            self.selected_ids.remove(book_id)
            self.tree.set(item, "selected", "☐")
        else:
            self.selected_ids.add(book_id)
            self.tree.set(item, "selected", "☑")
        
        self._update_selection_ui()
    
    def _toggle_select_all(self):
        """Toggle select all/deselect all."""
        self._select_all = not self._select_all
        
        for item in self.tree.get_children():
            values = self.tree.item(item)["values"]
            book_id = values[1]
            
            if self._select_all:
                self.selected_ids.add(book_id)
                self.tree.set(item, "selected", "☑")
            else:
                self.selected_ids.discard(book_id)
                self.tree.set(item, "selected", "")
        
        self._update_selection_ui()
    
    def _update_selection_ui(self):
        """Update UI elements based on selection."""
        count = len(self.selected_ids)
        
        # Update selection counter
        if count > 0:
            self.selection_label.configure(text=f"{tr('Selected')}: {count}")
            self.batch_delete_btn.configure(state="normal")
        else:
            self.selection_label.configure(text="")
            self.batch_delete_btn.configure(state="disabled")
        
        # Update header checkbox
        all_selected = count > 0 and count == len(self.books)
        self.tree.heading("selected", text="☑" if all_selected else "☐")
    
    def _add_book(self):
        """Switch to Excel import page view."""
        # Traverse up to find ModernApp instance
        app = self.winfo_toplevel()
        app.switch_view("excel_import")

    def _edit_book(self):
        """Edit selected book."""
        if not self.selected_ids:
            messagebox.showinfo(tr("Edit Book"), tr("Please select a book to edit"))
            return
        
        if len(self.selected_ids) > 1:
            messagebox.showinfo(tr("Edit Book"), tr("Please select only one book to edit"))
            return
        
        book_id = list(self.selected_ids)[0]
        
        # Find the book data
        book = None
        for b in self.books:
            if b.get("id") == book_id:
                book = b
                break
        
        if not book:
            messagebox.showerror(tr("Error"), tr("Book not found"))
            return
        
        # Show edit dialog
        self._show_edit_dialog(book)

    def _show_edit_dialog(self, book):
        """Show book edit dialog."""
        dialog = ctk.CTkToplevel(self)
        dialog.title(tr("Edit Book"))
        dialog.geometry("600x500")
        dialog.transient(self)
        dialog.grab_set()
        
        # Form fields
        fields = [
            ("title", tr("Title"), book.get("title", "")),
            ("authors", tr("Author"), book.get("authors", "")),
            ("publisher", tr("Publisher"), book.get("publisher", "")),
            ("isbn", "ISBN", book.get("isbn", "")),
            ("call_number", tr("Call Number"), book.get("call_number", "")),
        ]
        
        entries = {}
        for i, (key, label, default) in enumerate(fields):
            row = ctk.CTkFrame(dialog, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=8)
            
            ctk.CTkLabel(row, text=label, width=100, anchor="w").pack(side="left")
            entry = ctk.CTkEntry(row, width=400)
            entry.insert(0, str(default) if default else "")
            entry.pack(side="left", padx=10)
            entries[key] = entry
            
            # Add ISBN Lookup button next to ISBN field
            if key == "isbn":
                isbn_lookup_btn = ctk.CTkButton(
                    row,
                    text=tr("ISBN Lookup"),
                    font=ctk.CTkFont(size=12, weight="bold"),
                    fg_color=C_SUCCESS,
                    hover_color=C_SUCCESS_HOVER,
                    text_color=C_SIDEBAR_TEXT,
                    width=100,
                    height=32,
                    corner_radius=8,
                    command=lambda: self._isbn_lookup(entries)
                )
                isbn_lookup_btn.pack(side="left", padx=(0, 10))
        
        # Category dropdown
        cat_row = ctk.CTkFrame(dialog, fg_color="transparent")
        cat_row.pack(fill="x", padx=20, pady=8)
        ctk.CTkLabel(cat_row, text=tr("Category"), width=100, anchor="w").pack(side="left")
        
        try:
            cats = API.list_categories()
            cat_options = [""] + [f"{c['id']} - {c['name']}" for c in cats.get("data", [])]
        except:
            cat_options = [""]
        
        cat_var = ctk.StringVar()
        current_cat = book.get("category_id")
        if current_cat:
            for opt in cat_options:
                if opt.startswith(f"{current_cat} -"):
                    cat_var.set(opt)
                    break
        
        cat_menu = ctk.CTkOptionMenu(cat_row, variable=cat_var, values=cat_options, width=400)
        cat_menu.pack(side="left", padx=10)
        entries["category_id"] = cat_var
        
        # Copies
        copies_row = ctk.CTkFrame(dialog, fg_color="transparent")
        copies_row.pack(fill="x", padx=20, pady=8)
        ctk.CTkLabel(copies_row, text=tr("Total Copies"), width=100, anchor="w").pack(side="left")
        copies_entry = ctk.CTkEntry(copies_row, width=100)
        copies_entry.insert(0, str(book.get("total_copies", 0)))
        copies_entry.pack(side="left", padx=10)
        entries["total_copies"] = copies_entry
        
        # Buttons
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=20)
        
        def save():
            try:
                data = {}
                for key, entry in entries.items():
                    if key == "category_id":
                        val = entry.get()
                        if val and " - " in val:
                            data[key] = int(val.split(" - ")[0])
                        else:
                            data[key] = None
                    elif key == "total_copies":
                        val = entry.get().strip()
                        data[key] = int(val) if val else 0
                    else:
                        val = entry.get().strip()
                        data[key] = val if val else None
                
                API.update_book(book_id, data)
                ToastNotification(self.winfo_toplevel(), tr("Book updated successfully"), "success")
                dialog.destroy()
                self.refresh()
            except Exception as e:
                messagebox.showerror(tr("Error"), str(e))
        
        ctk.CTkButton(btn_frame, text=tr("Save"), command=save,
                     fg_color=C_SUCCESS, hover_color=C_SUCCESS_HOVER).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text=tr("Cancel"), command=dialog.destroy,
                     fg_color=C_INPUT_BG, hover_color=C_BORDER).pack(side="left", padx=10)

    def _batch_delete(self):
        """Batch delete selected books with confirmation dialog."""
        if not self.selected_ids:
            messagebox.showinfo(tr("Batch Delete"), tr("Please select books to delete"))
            return
        
        # Show detailed confirmation dialog
        count = len(self.selected_ids)
        confirm_msg = f"{tr('Are you sure you want to delete')} {count} {tr('books')}?\n\n"
        confirm_msg += tr("This action cannot be undone.")
        
        if not messagebox.askyesno(tr("Confirm Batch Delete"), confirm_msg, icon="warning"):
            return
        
        # Show loading overlay
        self.loading = LoadingOverlay(self, tr("Deleting..."))
        self.loading.show()
        
        def _do_delete():
            try:
                book_ids = list(self.selected_ids)
                API.batch_delete_books(book_ids)
                self.after(0, lambda: self._show_delete_success(len(book_ids)))
            except Exception as e:
                self.after(0, lambda: self._show_delete_error(str(e)))
            finally:
                self.after(0, self._hide_loading)
        
        threading.Thread(target=_do_delete, daemon=True).start()
    
    def _show_delete_success(self, count):
        """Show success message after deletion."""
        ToastNotification(self.winfo_toplevel(), f"{count} {tr('Books deleted successfully')}", "success")
        self.selected_ids.clear()
        self.refresh()
    
    def _show_delete_error(self, error):
        """Show error message after deletion failure."""
        messagebox.showerror(tr("Error"), f"{tr('Delete failed')}: {error}")

    def _classify_uncategorized(self):
        """Batch classify uncategorized books using web search + LLM."""
        import tkinter.messagebox as messagebox

        # Confirm dialog
        if not messagebox.askyesno(tr("Classify Uncategorized"),
                                    tr("This will use web search + LLM to classify all uncategorized books. Continue?")):
            return

        self.classify_btn.configure(state="disabled", text=tr("Classifying..."))
        
        # Show loading overlay
        self.loading = LoadingOverlay(self, tr("Classifying..."))
        self.loading.show()

        def _do_classify():
            try:
                result = API.classify_uncategorized_books()
                classified = result.get("classified", 0)
                failed = result.get("failed", 0)
                total = result.get("total", 0)
                msg = result.get("message", "")
                self.after(0, lambda: self._show_classify_result(classified, failed, total, msg))
            except Exception as e:
                self.after(0, lambda: self._show_classify_error(str(e)))
            finally:
                self.after(0, self._hide_loading)

        threading.Thread(target=_do_classify, daemon=True).start()
    
    def _hide_loading(self):
        """Hide loading overlay."""
        if hasattr(self, 'loading'):
            self.loading.hide()

    def _show_classify_result(self, classified, failed, total, msg):
        import tkinter.messagebox as messagebox
        messagebox.showinfo(tr("Classification Complete"), msg)
        self.classify_btn.configure(state="normal", text=tr("Classify Uncategorized"))
        self.refresh()

    def _show_classify_error(self, error):
        import tkinter.messagebox as messagebox
        messagebox.showerror(tr("Error"), error)
        self.classify_btn.configure(state="normal", text=tr("Classify Uncategorized"))

    def _isbn_lookup(self, entries):
        """Look up book info by ISBN and auto-fill fields."""
        isbn = entries["isbn"].get().strip()
        if not isbn:
            messagebox.showinfo(tr("Info"), tr("Please enter ISBN"))
            return
        
        def _do_lookup():
            try:
                result = API.isbn_lookup(isbn)
                if result and result.get("title"):
                    self.after(0, lambda: self._fill_isbn_info(entries, result))
                else:
                    self.after(0, lambda: messagebox.showinfo(tr("Info"), tr("Book info not found")))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))
        
        threading.Thread(target=_do_lookup, daemon=True).start()
    
    def _fill_isbn_info(self, entries, info):
        """Fill form fields with ISBN lookup result."""
        if info.get("title"):
            entries["title"].delete(0, "end")
            entries["title"].insert(0, info["title"])
        if info.get("authors"):
            entries["authors"].delete(0, "end")
            entries["authors"].insert(0, info["authors"])
        if info.get("publisher"):
            entries["publisher"].delete(0, "end")
            entries["publisher"].insert(0, info["publisher"])
        ToastNotification(self.winfo_toplevel(), tr("ISBN Lookup") + " - " + tr("Success"), "success")

    def _export_books(self):
        """Export books to Excel file."""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[(tr("Excel files"), "*.xlsx")]
        )
        if not save_path:
            return

        def _do_export():
            try:
                API.export_books(save_path)
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), tr("Export successful"), "success"))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

        threading.Thread(target=_do_export, daemon=True).start()


class UserProfileView(ctk.CTkFrame):
    """User profile page with personal info, borrowing history, and settings."""

    def __init__(self, master, reader_id=1, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.reader_id = reader_id
        self.profile = {}
        self.borrowings = []
        self.overdue = []
        self._build()

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        # ── Profile Header Card ──
        header_card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        header_card.pack(fill="x", padx=24, pady=(20, 12))

        header_inner = ctk.CTkFrame(header_card, fg_color="transparent")
        header_inner.pack(fill="x", padx=24, pady=20)

        # Avatar circle (clickable)
        self.avatar_frame = ctk.CTkFrame(header_inner, fg_color=C_INPUT_BG, width=80, height=80, corner_radius=40, border_width=2, border_color=C_BORDER)
        self.avatar_frame.pack(side="left", padx=(0, 20))
        self.avatar_frame.pack_propagate(False)
        self.avatar_frame.bind("<Button-1>", lambda e: self._upload_avatar())
        self.avatar_frame.configure(cursor="hand2")

        self.avatar_label = ctk.CTkLabel(
            self.avatar_frame, text="👤",
            font=ctk.CTkFont(size=36), text_color=C_TEXT_SEC
        )
        self.avatar_label.place(relx=0.5, rely=0.5, anchor="center")
        self.avatar_label.bind("<Button-1>", lambda e: self._upload_avatar())

        self.avatar_image = None  # Will hold PhotoImage if avatar is loaded

        # Name + identity
        info_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        info_frame.pack(side="left", fill="x", expand=True)

        self.name_label = ctk.CTkLabel(
            info_frame, text="",
            font=ctk.CTkFont(size=22, weight="bold"), text_color=C_TEXT, anchor="w"
        )
        self.name_label.pack(anchor="w")

        self.identity_label = ctk.CTkLabel(
            info_frame, text="",
            font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC, anchor="w"
        )
        self.identity_label.pack(anchor="w", pady=(4, 0))

        # Status badge
        self.status_badge = ctk.CTkLabel(
            info_frame, text="",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=C_SIDEBAR_TEXT,
            fg_color=C_SUCCESS, corner_radius=8
        )
        self.status_badge.pack(anchor="w", pady=(8, 0))

        # ── Stats Cards Row ──
        stats_row = ctk.CTkFrame(scroll, fg_color="transparent")
        stats_row.pack(fill="x", padx=24, pady=8)

        self.stat_total = self._make_stat_card(stats_row, "0", tr("Total Borrowed"), "#6366f1")
        self.stat_active = self._make_stat_card(stats_row, "0", tr("Currently Borrowed"), "#22c55e")
        self.stat_overdue = self._make_stat_card(stats_row, "0", tr("Overdue"), "#ef4444")
        self.stat_returned = self._make_stat_card(stats_row, "0", tr("Returned"), "#6b7280")

        # ── Tabs ──
        tab_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        tab_bar.pack(fill="x", padx=24, pady=(16, 0))

        self.tab_borrow = ctk.CTkButton(
            tab_bar, text=tr("Borrowing Records"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT, text_color=C_SIDEBAR_TEXT,
            height=36, corner_radius=10,
            command=lambda: self._switch_tab("borrow")
        )
        self.tab_borrow.pack(side="left", padx=(0, 8))

        self.tab_overdue = ctk.CTkButton(
            tab_bar, text=tr("Overdue Books"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_INPUT_BG, text_color=C_SIDEBAR_TEXT,
            height=36, corner_radius=10,
            command=lambda: self._switch_tab("overdue")
        )
        self.tab_overdue.pack(side="left", padx=(0, 8))

        self.tab_settings = ctk.CTkButton(
            tab_bar, text=tr("Account Settings"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_INPUT_BG, text_color=C_SIDEBAR_TEXT,
            height=36, corner_radius=10,
            command=lambda: self._switch_tab("settings")
        )
        self.tab_settings.pack(side="left")

        # Tab content containers
        self.borrow_frame = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        self.overdue_frame = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        self.settings_frame = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)

        self._build_borrow_tab()
        self._build_overdue_tab()
        self._build_settings_tab()

        self._current_tab = "borrow"
        self._switch_tab("borrow")

    def _make_stat_card(self, parent, value, label, color):
        card = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        card.pack(side="left", fill="x", expand=True, padx=4)

        ctk.CTkFrame(card, fg_color=color, height=3, corner_radius=0).pack(fill="x")

        val_label = ctk.CTkLabel(
            card, text=value,
            font=ctk.CTkFont(size=24, weight="bold"), text_color=C_TEXT
        )
        val_label.pack(pady=(12, 0))

        ctk.CTkLabel(
            card, text=label,
            font=ctk.CTkFont(size=11), text_color=C_TEXT_SEC
        ).pack(pady=(0, 12))

        return val_label

    def _build_borrow_tab(self):
        inner = ctk.CTkFrame(self.borrow_frame, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=20, pady=16)

        self.borrow_tree = ttk.Treeview(
            inner, columns=("id", "title", "borrow_date", "due_date", "return_date", "status"),
            show="headings", height=12
        )
        headings = {
            "id": "ID", "title": tr("Title"), "borrow_date": tr("Borrow Date"),
            "due_date": tr("Due Date"), "return_date": tr("Return Date"), "status": tr("Status")
        }
        widths = {"id": 50, "title": 260, "borrow_date": 100, "due_date": 100, "return_date": 100, "status": 80}
        for col in ("id", "title", "borrow_date", "due_date", "return_date", "status"):
            self.borrow_tree.heading(col, text=headings[col])
            self.borrow_tree.column(col, width=widths[col])

        vsb = ttk.Scrollbar(inner, orient="vertical", command=self.borrow_tree.yview)
        self.borrow_tree.configure(yscrollcommand=vsb.set)
        self.borrow_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.borrow_empty = ctk.CTkLabel(
            inner, text=tr("No borrowing records"),
            font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC
        )

    def _build_overdue_tab(self):
        inner = ctk.CTkFrame(self.overdue_frame, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=20, pady=16)

        self.overdue_tree = ttk.Treeview(
            inner, columns=("id", "title", "borrow_date", "due_date", "overdue_days", "status"),
            show="headings", height=8
        )
        headings = {
            "id": "ID", "title": tr("Title"), "borrow_date": tr("Borrow Date"),
            "due_date": tr("Due Date"), "overdue_days": tr("Overdue Days"), "status": tr("Status")
        }
        widths = {"id": 50, "title": 260, "borrow_date": 100, "due_date": 100, "overdue_days": 100, "status": 80}
        for col in ("id", "title", "borrow_date", "due_date", "overdue_days", "status"):
            self.overdue_tree.heading(col, text=headings[col])
            self.overdue_tree.column(col, width=widths[col])

        vsb = ttk.Scrollbar(inner, orient="vertical", command=self.overdue_tree.yview)
        self.overdue_tree.configure(yscrollcommand=vsb.set)
        self.overdue_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.overdue_empty = ctk.CTkLabel(
            inner, text=tr("No overdue books"),
            font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC
        )

    def _build_settings_tab(self):
        inner = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        inner.pack(fill="x", padx=20, pady=20)

        # Info fields (read-only)
        fields = [
            (tr("Name"), "name"),
            (tr("Card No."), "card_number"),
            (tr("Identity"), "identity_type"),
            (tr("Phone"), "phone"),
            (tr("Registered"), "register_date"),
            (tr("Max Borrow"), "max_borrow"),
        ]

        for label_text, key in fields:
            row = ctk.CTkFrame(inner, fg_color="transparent")
            row.pack(fill="x", pady=6)

            ctk.CTkLabel(
                row, text=label_text,
                font=ctk.CTkFont(size=13, weight="bold"), text_color=C_TEXT_SEC,
                width=120, anchor="w"
            ).pack(side="left", padx=(0, 12))

            lbl = ctk.CTkLabel(
                row, text="",
                font=ctk.CTkFont(size=13), text_color=C_TEXT, anchor="w"
            )
            lbl.pack(side="left", fill="x", expand=True)
            setattr(self, f"settings_{key}", lbl)

        # Separator
        sep = ctk.CTkFrame(inner, fg_color=C_BORDER, height=1, corner_radius=0)
        sep.pack(fill="x", pady=20)

        # Account actions
        actions_frame = ctk.CTkFrame(inner, fg_color="transparent")
        actions_frame.pack(fill="x", pady=8)

        ctk.CTkButton(
            actions_frame,
            text=tr("Register New Account"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT, hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=40, corner_radius=10,
            command=self._register_account
        ).pack(side="left", padx=(0, 12))

        ctk.CTkButton(
            actions_frame,
            text=tr("Logout"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#ef4444", hover_color="#dc2626",
            text_color=C_SIDEBAR_TEXT,
            height=40, corner_radius=10,
            command=self._logout_account
        ).pack(side="left")

    def _upload_avatar(self):
        """Open file dialog to select avatar image."""
        import tkinter.filedialog as filedialog
        filepath = filedialog.askopenfilename(
            title=tr("Select Avatar"),
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.webp")]
        )
        if not filepath:
            return

        try:
            from PIL import Image, ImageTk
            img = Image.open(filepath)
            img = img.resize((76, 76), Image.LANCZOS)
            # Make circular mask
            mask = Image.new("L", (76, 76), 0)
            from PIL import ImageDraw
            draw = ImageDraw.Draw(mask)
            draw.ellipse((0, 0, 76, 76), fill=255)
            img.putalpha(mask)

            self.avatar_image = ImageTk.PhotoImage(img)
            self.avatar_label.configure(text="", image=self.avatar_image)

            # Upload to backend
            API.upload_avatar(self.reader_id, filepath)
        except Exception as e:
            import tkinter.messagebox as messagebox
            messagebox.showerror(tr("Error"), f"{tr('Failed to load image')}: {e}")

    def _register_account(self):
        """Open registration dialog."""
        dialog = ctk.CTkToplevel(self)
        dialog.title(tr("Register New Account"))
        dialog.geometry("400x380")
        dialog.resizable(False, False)
        dialog.grab_set()

        ctk.CTkLabel(
            dialog, text=tr("Register New Account"),
            font=ctk.CTkFont(size=18, weight="bold"), text_color=C_TEXT
        ).pack(pady=(20, 16))

        fields_data = {}
        for label_text, key, placeholder in [
            (tr("Name"), "name", tr("Enter name")),
            (tr("Card No."), "card_number", tr("Enter card number")),
            (tr("Identity"), "identity_type", tr("e.g. Student / Staff")),
            (tr("Phone"), "phone", tr("Enter phone number")),
        ]:
            row = ctk.CTkFrame(dialog, fg_color="transparent")
            row.pack(fill="x", padx=24, pady=4)

            ctk.CTkLabel(
                row, text=label_text,
                font=ctk.CTkFont(size=13, weight="bold"), text_color=C_TEXT,
                width=100, anchor="w"
            ).pack(side="left")

            entry = ctk.CTkEntry(row, placeholder_text=placeholder, font=ctk.CTkFont(size=13))
            entry.pack(side="left", fill="x", expand=True, padx=(8, 0))
            fields_data[key] = entry

        def _do_register():
            name = fields_data["name"].get().strip()
            card_number = fields_data["card_number"].get().strip()
            if not name or not card_number:
                import tkinter.messagebox as messagebox
                messagebox.showwarning(tr("Error"), tr("Name and Card No. are required"))
                return

            try:
                result = API.create_reader({
                    "name": name,
                    "card_number": card_number,
                    "identity_type": fields_data["identity_type"].get().strip() or None,
                    "phone": fields_data["phone"].get().strip() or None,
                    "card_status": "正常",
                    "max_borrow": 10,
                })
                dialog.destroy()
                ToastNotification(self.winfo_toplevel(), tr("Account registered successfully!"), "success")
                # Switch to the new account
                self.refresh(reader_id=result.get("id"))
            except Exception as e:
                import tkinter.messagebox as messagebox
                messagebox.showerror(tr("Error"), str(e))

        ctk.CTkButton(
            dialog, text=tr("Register"),
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=C_ACCENT, hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=42, corner_radius=10,
            command=_do_register
        ).pack(pady=(20, 16))

    def _logout_account(self):
        """Switch back to dashboard."""
        import tkinter.messagebox as messagebox
        if messagebox.askyesno(tr("Logout"), tr("Are you sure you want to logout?")):
            app = self.winfo_toplevel()
            app.switch_view("dashboard")

    def _switch_tab(self, tab):
        self._current_tab = tab
        for frame in (self.borrow_frame, self.overdue_frame, self.settings_frame):
            frame.pack_forget()
        for btn in (self.tab_borrow, self.tab_overdue, self.tab_settings):
            btn.configure(fg_color=C_INPUT_BG)

        if tab == "borrow":
            self.borrow_frame.pack(fill="both", expand=True, padx=24, pady=8)
            self.tab_borrow.configure(fg_color=C_ACCENT)
        elif tab == "overdue":
            self.overdue_frame.pack(fill="both", expand=True, padx=24, pady=8)
            self.tab_overdue.configure(fg_color=C_ACCENT)
        elif tab == "settings":
            self.settings_frame.pack(fill="x", padx=24, pady=8)
            self.tab_settings.configure(fg_color=C_ACCENT)

    def refresh(self, reader_id=None):
        if reader_id is not None:
            self.reader_id = reader_id

        try:
            self.profile = API.get_reader_profile(self.reader_id)
            self.borrowings = API.get_reader_borrowings(self.reader_id).get("data", [])
            self.overdue = API.get_reader_overdue(self.reader_id).get("data", [])
            self.after(0, self._update_ui)
        except Exception as e:
            print(f"UserProfile refresh error: {e}")

    def _update_ui(self):
        p = self.profile
        stats = p.get("stats", {})

        # Header
        self.name_label.configure(text=p.get("name", ""))
        identity = p.get("identity_type", "")
        card = p.get("card_number", "")
        self.identity_label.configure(text=f"{identity}  |  {tr('Card No.')}: {card}" if identity else f"{tr('Card No.')}: {card}")

        status = p.get("card_status", "")
        self.status_badge.configure(text=status)
        if status == "正常":
            self.status_badge.configure(fg_color=C_SUCCESS)
        elif status == "挂失":
            self.status_badge.configure(fg_color="#f59e0b")
        else:
            self.status_badge.configure(fg_color="#ef4444")

        # Stats
        self.stat_total.configure(text=str(stats.get("total_borrowed", 0)))
        self.stat_active.configure(text=str(stats.get("active", 0)))
        self.stat_overdue.configure(text=str(stats.get("overdue", 0)))
        self.stat_returned.configure(text=str(stats.get("returned", 0)))

        # Borrowing tree
        for item in self.borrow_tree.get_children():
            self.borrow_tree.delete(item)
        if self.borrowings:
            self.borrow_empty.pack_forget()
            for b in self.borrowings:
                self.borrow_tree.insert("", "end", values=(
                    b.get("id", ""), b.get("book_title", ""),
                    b.get("borrow_date", ""), b.get("due_date", ""),
                    b.get("return_date", "") or "-", b.get("status", "")
                ))
        else:
            self.borrow_empty.pack(pady=40)

        # Overdue tree
        for item in self.overdue_tree.get_children():
            self.overdue_tree.delete(item)
        if self.overdue:
            self.overdue_empty.pack_forget()
            for b in self.overdue:
                self.overdue_tree.insert("", "end", values=(
                    b.get("id", ""), b.get("book_title", ""),
                    b.get("borrow_date", ""), b.get("due_date", ""),
                    b.get("overdue_days", 0), b.get("status", "")
                ))
        else:
            self.overdue_empty.pack(pady=40)

        # Settings fields
        self.settings_name.configure(text=p.get("name", ""))
        self.settings_card_number.configure(text=p.get("card_number", ""))
        self.settings_identity_type.configure(text=p.get("identity_type", "") or "-")
        self.settings_phone.configure(text=p.get("phone", "") or "-")
        self.settings_register_date.configure(text=p.get("register_date", "") or "-")
        self.settings_max_borrow.configure(text=str(p.get("max_borrow", 10)))


class ExcelImportView(ctk.CTkFrame):
    """Excel import page view with LLM auto-classification."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.filepath = None
        self._build()
    
    def _build(self):
        # Title
        ctk.CTkLabel(
            self,
            text=tr("Import Books from Excel"),
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=C_TEXT
        ).pack(pady=(24, 8))
        
        # Description
        ctk.CTkLabel(
            self,
            text=tr("Select an Excel file. AI will automatically classify each book."),
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        ).pack(pady=(0, 20))
        
        # File selection card
        file_card = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        file_card.pack(fill="x", padx=24, pady=10)
        
        file_inner = ctk.CTkFrame(file_card, fg_color="transparent")
        file_inner.pack(fill="x", padx=20, pady=16)
        
        self.file_label = ctk.CTkLabel(
            file_inner,
            text=tr("No file selected"),
            font=ctk.CTkFont(size=14),
            text_color=C_TEXT_SEC,
            anchor="w"
        )
        self.file_label.pack(side="left", fill="x", expand=True)
        
        ctk.CTkButton(
            file_inner,
            text=tr("Select File"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=120,
            height=36,
            corner_radius=10,
            command=self._select_file
        ).pack(side="right", padx=(12, 0))
        
        # Preview card
        preview_card = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        preview_card.pack(fill="both", expand=True, padx=24, pady=10)
        
        preview_header = ctk.CTkFrame(preview_card, fg_color="transparent")
        preview_header.pack(fill="x", padx=20, pady=(16, 8))
        
        ctk.CTkLabel(
            preview_header,
            text=tr("Preview"),
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=C_TEXT
        ).pack(side="left")
        
        self.preview_text = ctk.CTkTextbox(
            preview_card,
            font=ctk.CTkFont(size=12, family="Consolas"),
            text_color=C_TEXT,
            fg_color="transparent",
            border_width=0,
            wrap="word"
        )
        self.preview_text.pack(fill="both", expand=True, padx=20, pady=(0, 16))
        self.preview_text.insert("1.0", tr("Select a file to preview..."))
        self.preview_text.configure(state="disabled")
        
        # Status + Buttons bar
        bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        bottom_bar.pack(fill="x", padx=24, pady=(10, 24))
        
        self.status_label = ctk.CTkLabel(
            bottom_bar,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        )
        self.status_label.pack(side="left")
        
        btn_frame = ctk.CTkFrame(bottom_bar, fg_color="transparent")
        btn_frame.pack(side="right")
        
        ctk.CTkButton(
            btn_frame,
            text=tr("Back to Books"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            width=120,
            height=36,
            corner_radius=10,
            command=self._go_back
        ).pack(side="left", padx=(0, 10))
        
        self.import_btn = ctk.CTkButton(
            btn_frame,
            text=tr("Import with AI"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=140,
            height=36,
            corner_radius=10,
            command=self._import_with_ai,
            state="disabled"
        )
        self.import_btn.pack(side="right")
    
    def refresh(self):
        """Reset the view when navigated to."""
        pass
    
    def _select_file(self):
        filepath = filedialog.askopenfilename(
            title=tr("Select Excel File"),
            filetypes=[
                (tr("Excel files"), "*.xlsx *.xls"),
                (tr("All files"), "*.*")
            ]
        )
        if filepath:
            self.filepath = filepath
            filename = os.path.basename(filepath)
            self.file_label.configure(text=filename, text_color=C_TEXT)
            self.import_btn.configure(state="normal")
            self._preview_file()
    
    def _preview_file(self):
        """Read Excel and show preview."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
                if all(cell is None for cell in row):
                    continue
                rows.append(f"Row {i}: {row[0]}")
                if len(rows) >= 10:
                    break
            
            wb.close()
            
            preview = f"Headers: {', '.join(str(h) for h in headers)}\n\n"
            preview += f"Books found: {len(rows)}\n\n"
            preview += "First 10 books:\n" + "\n".join(rows)
            
            self.preview_text.configure(state="normal")
            self.preview_text.delete("1.0", "end")
            self.preview_text.insert("1.0", preview)
            self.preview_text.configure(state="disabled")
            
        except Exception as e:
            self.preview_text.configure(state="normal")
            self.preview_text.delete("1.0", "end")
            self.preview_text.insert("1.0", f"Error reading file: {e}")
            self.preview_text.configure(state="disabled")
    
    def _import_with_ai(self):
        if not self.filepath:
            return
        
        self.import_btn.configure(state="disabled", text=tr("Importing..."))
        self.status_label.configure(text=tr("Processing..."), text_color=C_TEXT_SEC)
        
        threading.Thread(target=self._do_import, daemon=True).start()
    
    def _do_import(self):
        try:
            # Use regular import (Excel already has category_name column)
            result = API.import_books(self.filepath)
            
            imported = result.get("imported", 0)
            errors = result.get("errors", [])
            message = result.get("message", "")
            
            result_text = f"{message}\n\n"
            
            if errors:
                result_text += f"Errors ({len(errors)}):\n"
                for err in errors[:20]:
                    result_text += f"  {err}\n"
                if len(errors) > 20:
                    result_text += f"  ... and {len(errors) - 20} more\n"
            
            self.after(0, lambda: self._show_result(result_text))
            
        except Exception as e:
            self.after(0, lambda: self._show_result(f"Import failed: {e}"))
    
    def _show_result(self, result_text):
        self.preview_text.configure(state="normal")
        self.preview_text.delete("1.0", "end")
        self.preview_text.insert("1.0", result_text)
        self.preview_text.configure(state="disabled")
        
        self.status_label.configure(text=tr("Import complete"), text_color=C_SUCCESS)
        self.import_btn.configure(state="normal", text=tr("Import with AI"))
    
    def _go_back(self):
        app = self.winfo_toplevel()
        app.switch_view("books")


class ReadersView(ctk.CTkFrame):
    """Readers management view."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.readers = []
        self._build()
    
    def _build(self):
        toolbar = ctk.CTkFrame(self, fg_color="transparent")
        toolbar.pack(fill="x", padx=24, pady=(20, 0))
        
        add_btn = ctk.CTkButton(
            toolbar,
            text=tr("+ Add Reader"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._add_reader
        )
        add_btn.pack(side="right")
        add_hover_effect(add_btn, hover_color=C_ACCENT_HOVER, original_color=C_ACCENT)
        
        reset_btn = ctk.CTkButton(
            toolbar,
            text=tr("Reset Password"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_WARN,
            hover_color="#e8c85e",
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._reset_password
        )
        reset_btn.pack(side="right", padx=(0, 8))
        add_hover_effect(reset_btn, hover_color="#e8c85e", original_color=C_WARN)

        # Edit Identity button
        edit_btn = ctk.CTkButton(
            toolbar,
            text=tr("Edit Identity"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_SUCCESS,
            hover_color=C_SUCCESS_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._edit_identity
        )
        edit_btn.pack(side="right", padx=(0, 8))
        add_hover_effect(edit_btn, hover_color=C_SUCCESS_HOVER, original_color=C_SUCCESS)

        # Freeze/Unfreeze button
        freeze_btn = ctk.CTkButton(
            toolbar,
            text=tr("Freeze"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_WARN,
            hover_color="#e8c85e",
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._toggle_freeze
        )
        freeze_btn.pack(side="right", padx=(0, 8))
        add_hover_effect(freeze_btn, hover_color="#e8c85e", original_color=C_WARN)

        # Delete button
        delete_btn = ctk.CTkButton(
            toolbar,
            text=tr("Delete"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_DANGER,
            hover_color=C_DANGER_HOVER,
            text_color="#ffffff",
            height=36,
            corner_radius=10,
            command=self._delete_reader
        )
        delete_btn.pack(side="right", padx=(0, 8))
        add_hover_effect(delete_btn, hover_color=C_DANGER_HOVER, original_color=C_DANGER)

        # Export button
        ctk.CTkButton(
            toolbar,
            text=tr("Export"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_SUCCESS,
            hover_color=C_SUCCESS_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._export_readers
        ).pack(side="right", padx=(0, 8))
        
        table_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=(16, 24))
        
        columns = ("id", "card_number", "name", "identity_type", "phone", "card_status")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        headings = {"id": "ID", "card_number": tr("Card No."), "name": tr("Name"),
                    "identity_type": tr("Identity"), "phone": tr("Phone"), "card_status": tr("Status")}
        widths = {"id": 60, "card_number": 100, "name": 150, "identity_type": 120,
                  "phone": 120, "card_status": 80}
        
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col])
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
    
    def refresh(self):
        try:
            data = API.list_readers()
            self.readers = data.get("data", [])
            for item in self.tree.get_children():
                self.tree.delete(item)
            for r in self.readers:
                self.tree.insert("", "end", values=(
                    r.get("id", ""),
                    r.get("card_number", ""),
                    r.get("name", ""),
                    r.get("identity_type", "") or "—",
                    r.get("phone", "") or "—",
                    r.get("card_status", ""),
                ))
        except Exception as e:
            print(f"Readers refresh error: {e}")
    
    def _add_reader(self):
        """Switch to reader import page view."""
        app = self.winfo_toplevel()
        app.switch_view("reader_import")
    
    def _reset_password(self):
        """Reset password for selected reader."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(tr("Warning"), tr("Please select a reader"))
            return
        
        item = self.tree.item(selection[0])
        reader_id = item["values"][0]
        reader_name = item["values"][2]
        
        # Create dialog
        dialog = ctk.CTkToplevel(self)
        dialog.title(tr("Reset Password"))
        dialog.geometry("400x200")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        
        ctk.CTkLabel(
            dialog,
            text=f"{tr('Reset password for')}: {reader_name}",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(20, 10))
        
        ctk.CTkLabel(
            dialog,
            text=tr("New Password"),
            font=ctk.CTkFont(size=13)
        ).pack(pady=(0, 5))
        
        password_entry = ctk.CTkEntry(
            dialog,
            show="*",
            width=250,
            height=36,
            corner_radius=8
        )
        password_entry.pack(pady=(0, 15))
        
        def do_reset():
            new_password = password_entry.get()
            if not new_password:
                messagebox.showwarning(tr("Warning"), tr("Password cannot be empty"))
                return
            
            if len(new_password) < 6:
                messagebox.showwarning(tr("Warning"), tr("Password must be at least 6 characters"))
                return
            
            try:
                API.reset_reader_password(reader_id, new_password)
                ToastNotification(self.winfo_toplevel(), tr("Password reset successfully"), "success")
                dialog.destroy()
            except Exception as e:
                messagebox.showerror(tr("Error"), str(e))
        
        ctk.CTkButton(
            dialog,
            text=tr("Confirm"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=120,
            height=36,
            corner_radius=8,
            command=do_reset
        ).pack()

    def _edit_identity(self):
        """Edit identity type for selected reader."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(tr("Warning"), tr("Please select a reader"))
            return

        item = self.tree.item(selection[0])
        reader_id = item["values"][0]
        reader_name = item["values"][2]
        current_identity = item["values"][3] if item["values"][3] != "—" else ""

        dialog = ctk.CTkToplevel(self)
        dialog.title(tr("Edit Identity"))
        dialog.geometry("400x220")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()

        ctk.CTkLabel(
            dialog, text=f"{tr('Edit identity for')}: {reader_name}",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(20, 10))

        ctk.CTkLabel(dialog, text=tr("Identity Type"), font=ctk.CTkFont(size=13)).pack(pady=(0, 5))

        identity_var = ctk.StringVar(value=current_identity)
        identity_combo = ctk.CTkComboBox(
            dialog, values=["学生", "教师", "访客"],
            variable=identity_var, width=250, height=36, corner_radius=8
        )
        identity_combo.pack(pady=(0, 15))

        def do_save():
            new_identity = identity_var.get()

            def _do_update():
                try:
                    API.update_reader(reader_id, {"identity_type": new_identity})
                    self.after(0, lambda: ToastNotification(self.winfo_toplevel(), tr("Identity updated"), "success"))
                    self.after(0, self.refresh)
                except Exception as e:
                    self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

            threading.Thread(target=_do_update, daemon=True).start()
            dialog.destroy()

        ctk.CTkButton(
            dialog, text=tr("Save"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT, hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT, width=120, height=36, corner_radius=8,
            command=do_save
        ).pack()

    def _toggle_freeze(self):
        """Freeze or unfreeze selected reader."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(tr("Warning"), tr("Please select a reader"))
            return

        item = self.tree.item(selection[0])
        reader_id = item["values"][0]
        reader_name = item["values"][2]
        current_status = item["values"][5]

        new_status = "冻结" if current_status == "正常" else "正常"
        action_text = tr("Freeze") if new_status == "冻结" else tr("Unfreeze")

        if messagebox.askyesno(
            f"{action_text} {tr('Reader')}",
            f"{action_text} \"{reader_name}\"?\n({tr('Current status')}: {current_status} → {new_status})"
        ):
            def _do_toggle():
                try:
                    API.update_reader(reader_id, {"card_status": new_status})
                    self.after(0, lambda: ToastNotification(self.winfo_toplevel(), f"{action_text} {tr('successful')}", "success"))
                    self.after(0, self.refresh)
                except Exception as e:
                    self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

            threading.Thread(target=_do_toggle, daemon=True).start()

    def _delete_reader(self):
        """Delete selected reader after confirmation."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(tr("Warning"), tr("Please select a reader"))
            return

        item = self.tree.item(selection[0])
        reader_id = item["values"][0]
        reader_name = item["values"][2]

        if not messagebox.askyesno(tr("Delete Reader"), f"{tr('Are you sure you want to delete')} \"{reader_name}\"?"):
            return

        def _do_delete():
            try:
                API.delete_reader(reader_id)
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), tr("Reader deleted"), "success"))
                self.after(0, self.refresh)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

        threading.Thread(target=_do_delete, daemon=True).start()

    def _export_readers(self):
        """Export readers to Excel file."""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[(tr("Excel files"), "*.xlsx")]
        )
        if not save_path:
            return

        def _do_export():
            try:
                API.export_readers(save_path)
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), tr("Export successful"), "success"))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

        threading.Thread(target=_do_export, daemon=True).start()


class ReaderImportView(ctk.CTkFrame):
    """Reader Excel import page view."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.filepath = None
        self._build()
    
    def _build(self):
        # Title
        ctk.CTkLabel(
            self,
            text=tr("Reader Import"),
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=C_TEXT
        ).pack(pady=(24, 8))
        
        # Description
        ctk.CTkLabel(
            self,
            text=tr("Import readers from Excel"),
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        ).pack(pady=(0, 20))
        
        # File selection card
        file_card = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        file_card.pack(fill="x", padx=24, pady=10)
        
        file_inner = ctk.CTkFrame(file_card, fg_color="transparent")
        file_inner.pack(fill="x", padx=20, pady=16)
        
        self.file_label = ctk.CTkLabel(
            file_inner,
            text=tr("No file selected"),
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC,
            anchor="w"
        )
        self.file_label.pack(side="left", fill="x", expand=True)
        
        ctk.CTkButton(
            file_inner,
            text=tr("Select File"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=120,
            height=36,
            corner_radius=10,
            command=self._select_file
        ).pack(side="right")
        
        # Preview card
        preview_card = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        preview_card.pack(fill="both", expand=True, padx=24, pady=10)
        
        ctk.CTkLabel(
            preview_card,
            text=tr("Preview"),
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", padx=20, pady=(16, 8))
        
        self.preview_text = ctk.CTkTextbox(
            preview_card,
            font=ctk.CTkFont(size=12, family="Consolas"),
            fg_color=C_INPUT_BG,
            border_width=0,
            corner_radius=10,
            wrap="word"
        )
        self.preview_text.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.preview_text.insert("1.0", tr("Select a file to preview..."))
        self.preview_text.configure(state="disabled")
        
        # Status label
        self.status_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        )
        self.status_label.pack(pady=(10, 0))
        
        # Bottom buttons bar
        bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        bottom_bar.pack(fill="x", padx=24, pady=(10, 24))
        
        btn_frame = ctk.CTkFrame(bottom_bar, fg_color="transparent")
        btn_frame.pack(side="right")
        
        ctk.CTkButton(
            btn_frame,
            text=tr("Back to Readers"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            width=120,
            height=36,
            corner_radius=10,
            command=self._go_back
        ).pack(side="left", padx=(0, 10))
        
        self.import_btn = ctk.CTkButton(
            btn_frame,
            text=tr("Import"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=100,
            height=36,
            corner_radius=10,
            state="disabled",
            command=self._do_import
        )
        self.import_btn.pack(side="right")
    
    def refresh(self):
        """Reset the view when navigated to."""
        self.filepath = None
        self.file_label.configure(text=tr("No file selected"))
        self.preview_text.configure(state="normal")
        self.preview_text.delete("1.0", "end")
        self.preview_text.insert("1.0", tr("Select a file to preview..."))
        self.preview_text.configure(state="disabled")
        self.status_label.configure(text="")
        self.import_btn.configure(state="disabled")
    
    def _go_back(self):
        app = self.winfo_toplevel()
        app.switch_view("readers")
    
    def _select_file(self):
        filepath = filedialog.askopenfilename(
            title=tr("Select Excel File"),
            filetypes=[(tr("Excel files"), "*.xlsx *.xls"), (tr("All files"), "*.*")]
        )
        if filepath:
            self.filepath = filepath
            self.file_label.configure(text=os.path.basename(filepath), text_color=C_TEXT)
            self.import_btn.configure(state="normal")
            self._preview_file()
    
    def _preview_file(self):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            ws = wb.active
            lines = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 15:
                    lines.append("...")
                    break
                vals = [str(v) if v is not None else "" for v in row]
                lines.append(" | ".join(vals))
            wb.close()
            
            self.preview_text.configure(state="normal")
            self.preview_text.delete("1.0", "end")
            self.preview_text.insert("1.0", "\n".join(lines))
            self.preview_text.configure(state="disabled")
        except Exception as e:
            self.preview_text.configure(state="normal")
            self.preview_text.delete("1.0", "end")
            self.preview_text.insert("1.0", f"{tr('Error reading file')}: {e}")
            self.preview_text.configure(state="disabled")
    
    def _do_import(self):
        if not self.filepath:
            return
        
        self.import_btn.configure(state="disabled", text=tr("Importing..."))
        self.status_label.configure(text=tr("Processing..."), text_color=C_TEXT_SEC)
        
        threading.Thread(target=self._do_import_thread, daemon=True).start()
    
    def _do_import_thread(self):
        try:
            result = API.import_readers(self.filepath)
            imported = result.get("imported", 0)
            skipped = result.get("skipped", 0)
            errors = result.get("errors", [])
            
            lines = [f"{tr('Import complete')}: {tr('Imported')} {imported}, {tr('Skipped')} {skipped}"]
            if errors:
                lines.append(f"\n{tr('Errors')}:")
                for err in errors[:10]:
                    lines.append(f"  Row {err.get('row', '?')}: {err.get('error', '')}")
            
            self.after(0, lambda: self._show_result("\n".join(lines), True))
        except Exception as e:
            self.after(0, lambda: self._show_result(f"{tr('Import failed')}: {e}", False))
    
    def _show_result(self, text, success):
        self.status_label.configure(text=text, text_color=C_SUCCESS if success else "#ef4444")
        self.import_btn.configure(state="normal", text=tr("Import"))


class AuditLogView(ctk.CTkFrame):
    """Audit log viewer with filters and pagination."""

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.current_page = 1
        self.page_size = 20
        self.total_pages = 1
        self._build()

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        # Title
        ctk.CTkLabel(scroll, text="审计日志",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", padx=24, pady=(20, 4))
        ctk.CTkLabel(scroll, text="查看所有操作记录，确保系统安全与合规",
                     font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC).pack(anchor="w", padx=24, pady=(0, 16))

        # ── Filters ──
        filter_frame = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        filter_frame.pack(fill="x", padx=24, pady=(0, 16))

        filter_inner = ctk.CTkFrame(filter_frame, fg_color="transparent")
        filter_inner.pack(fill="x", padx=16, pady=12)

        # Row 1: User type, Action, Resource type
        row1 = ctk.CTkFrame(filter_inner, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(row1, text="用户类型:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C_TEXT).pack(side="left", padx=(0, 8))
        self.user_type_var = ctk.StringVar(value="")
        user_type_menu = ctk.CTkOptionMenu(row1, variable=self.user_type_var,
                                            values=["", "admin", "reader"],
                                            font=ctk.CTkFont(size=12), width=120, height=32)
        user_type_menu.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="操作:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C_TEXT).pack(side="left", padx=(0, 8))
        self.action_var = ctk.StringVar(value="")
        action_menu = ctk.CTkOptionMenu(row1, variable=self.action_var,
                                         values=["", "create", "update", "delete", "login", "borrow", "return_request", "confirm_return", "reject_return"],
                                         font=ctk.CTkFont(size=12), width=140, height=32)
        action_menu.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="资源类型:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C_TEXT).pack(side="left", padx=(0, 8))
        self.resource_type_var = ctk.StringVar(value="")
        resource_type_menu = ctk.CTkOptionMenu(row1, variable=self.resource_type_var,
                                                values=["", "book", "reader", "borrowing", "admin"],
                                                font=ctk.CTkFont(size=12), width=120, height=32)
        resource_type_menu.pack(side="left", padx=(0, 16))

        # Row 2: Date range
        row2 = ctk.CTkFrame(filter_inner, fg_color="transparent")
        row2.pack(fill="x")

        ctk.CTkLabel(row2, text="开始日期:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C_TEXT).pack(side="left", padx=(0, 8))
        self.start_date_entry = ctk.CTkEntry(row2, placeholder_text="YYYY-MM-DD",
                                              font=ctk.CTkFont(size=12), width=120, height=32)
        self.start_date_entry.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row2, text="结束日期:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C_TEXT).pack(side="left", padx=(0, 8))
        self.end_date_entry = ctk.CTkEntry(row2, placeholder_text="YYYY-MM-DD",
                                            font=ctk.CTkFont(size=12), width=120, height=32)
        self.end_date_entry.pack(side="left", padx=(0, 16))

        # Buttons
        btn_frame = ctk.CTkFrame(row2, fg_color="transparent")
        btn_frame.pack(side="right")

        ctk.CTkButton(btn_frame, text="查询", font=ctk.CTkFont(size=12, weight="bold"),
                      fg_color=C_ACCENT, hover_color=C_ACCENT_HOVER,
                      text_color=C_SIDEBAR_TEXT, width=80, height=32,
                      command=self._apply_filters).pack(side="left", padx=(0, 8))

        ctk.CTkButton(btn_frame, text="重置", font=ctk.CTkFont(size=12),
                      fg_color=C_INPUT_BG, hover_color=C_BORDER,
                      text_color=C_TEXT, width=80, height=32,
                      command=self._reset_filters).pack(side="left")

        # ── Table ──
        table_frame = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        # Table header
        header = ctk.CTkFrame(table_frame, fg_color=C_INPUT_BG, corner_radius=0, height=40)
        header.pack(fill="x")
        header.pack_propagate(False)

        columns = [("时间", 0.18), ("用户", 0.12), ("操作", 0.12), ("资源类型", 0.12), ("资源ID", 0.10), ("详情", 0.36)]
        for col_text, weight in columns:
            lbl = ctk.CTkLabel(header, text=col_text, font=ctk.CTkFont(size=12, weight="bold"),
                               text_color=C_TEXT, anchor="w")
            lbl.pack(side="left", padx=8, fill="x", expand=True)
            lbl.pack_propagate(False)

        # Table content
        self.table_content = ctk.CTkFrame(table_frame, fg_color="transparent")
        self.table_content.pack(fill="both", expand=True, padx=0, pady=0)

        # ── Pagination ──
        pagination_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        pagination_frame.pack(fill="x", padx=24, pady=(0, 20))

        self.prev_btn = ctk.CTkButton(pagination_frame, text="上一页", font=ctk.CTkFont(size=12),
                                       fg_color=C_INPUT_BG, hover_color=C_BORDER,
                                       text_color=C_TEXT, width=80, height=32,
                                       command=self._prev_page)
        self.prev_btn.pack(side="left")

        self.page_label = ctk.CTkLabel(pagination_frame, text="第 1 页 / 共 1 页",
                                        font=ctk.CTkFont(size=12), text_color=C_TEXT_SEC)
        self.page_label.pack(side="left", padx=16)

        self.next_btn = ctk.CTkButton(pagination_frame, text="下一页", font=ctk.CTkFont(size=12),
                                       fg_color=C_INPUT_BG, hover_color=C_BORDER,
                                       text_color=C_TEXT, width=80, height=32,
                                       command=self._next_page)
        self.next_btn.pack(side="left")

        self.total_label = ctk.CTkLabel(pagination_frame, text="共 0 条记录",
                                         font=ctk.CTkFont(size=12), text_color=C_TEXT_SEC)
        self.total_label.pack(side="right")

    def refresh(self):
        """Refresh audit logs from API."""
        try:
            user_type = self.user_type_var.get() if self.user_type_var.get() else None
            action = self.action_var.get() if self.action_var.get() else None
            resource_type = self.resource_type_var.get() if self.resource_type_var.get() else None
            start_date = self.start_date_entry.get().strip() if self.start_date_entry.get().strip() else None
            end_date = self.end_date_entry.get().strip() if self.end_date_entry.get().strip() else None

            result = API.list_audit_logs(
                user_type=user_type,
                action=action,
                resource_type=resource_type,
                start_date=start_date,
                end_date=end_date,
                page=self.current_page,
                page_size=self.page_size
            )

            logs = result.get("data", [])
            total = result.get("total", 0)
            self.total_pages = max(1, (total + self.page_size - 1) // self.page_size)

            self._update_table(logs)
            self._update_pagination(total)
        except Exception as e:
            print(f"Error loading audit logs: {e}")
            self._update_table([])
            self._update_pagination(0)

    def _update_table(self, logs):
        """Update table with log data."""
        # Clear existing rows
        for widget in self.table_content.winfo_children():
            widget.destroy()

        if not logs:
            ctk.CTkLabel(self.table_content, text="暂无审计日志记录",
                         font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC).pack(pady=40)
            return

        for log in logs:
            row = ctk.CTkFrame(self.table_content, fg_color="transparent", height=48)
            row.pack(fill="x", padx=0, pady=0)
            row.pack_propagate(False)

            # Time
            created_at = log.get("created_at", "")
            if created_at:
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    time_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    time_str = created_at
            else:
                time_str = ""
            ctk.CTkLabel(row, text=time_str, font=ctk.CTkFont(size=11),
                         text_color=C_TEXT, anchor="w").pack(side="left", padx=8, fill="x", expand=True)

            # User
            user_id = log.get("user_id")
            user_type = log.get("user_type", "")
            user_str = f"{user_type}#{user_id}" if user_id else user_type
            ctk.CTkLabel(row, text=user_str, font=ctk.CTkFont(size=11),
                         text_color=C_TEXT, anchor="w").pack(side="left", padx=8, fill="x", expand=True)

            # Action
            action = log.get("action", "")
            ctk.CTkLabel(row, text=action, font=ctk.CTkFont(size=11),
                         text_color=C_TEXT, anchor="w").pack(side="left", padx=8, fill="x", expand=True)

            # Resource type
            resource_type = log.get("resource_type", "")
            ctk.CTkLabel(row, text=resource_type, font=ctk.CTkFont(size=11),
                         text_color=C_TEXT, anchor="w").pack(side="left", padx=8, fill="x", expand=True)

            # Resource ID
            resource_id = log.get("resource_id", "")
            ctk.CTkLabel(row, text=str(resource_id) if resource_id else "", font=ctk.CTkFont(size=11),
                         text_color=C_TEXT, anchor="w").pack(side="left", padx=8, fill="x", expand=True)

            # Details
            details = log.get("details", "")
            if details:
                try:
                    details_dict = json.loads(details)
                    details_str = json.dumps(details_dict, ensure_ascii=False, indent=2)
                except:
                    details_str = str(details)
            else:
                details_str = ""
            ctk.CTkLabel(row, text=details_str, font=ctk.CTkFont(size=10),
                         text_color=C_TEXT_SEC, anchor="w", justify="left").pack(side="left", padx=8, fill="x", expand=True)

    def _update_pagination(self, total):
        """Update pagination controls."""
        self.page_label.configure(text=f"第 {self.current_page} 页 / 共 {self.total_pages} 页")
        self.total_label.configure(text=f"共 {total} 条记录")
        self.prev_btn.configure(state="normal" if self.current_page > 1 else "disabled")
        self.next_btn.configure(state="normal" if self.current_page < self.total_pages else "disabled")

    def _apply_filters(self):
        """Apply filters and refresh."""
        self.current_page = 1
        self.refresh()

    def _reset_filters(self):
        """Reset all filters."""
        self.user_type_var.set("")
        self.action_var.set("")
        self.resource_type_var.set("")
        self.start_date_entry.delete(0, "end")
        self.end_date_entry.delete(0, "end")
        self.current_page = 1
        self.refresh()

    def _prev_page(self):
        """Go to previous page."""
        if self.current_page > 1:
            self.current_page -= 1
            self.refresh()

    def _next_page(self):
        """Go to next page."""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.refresh()


class BorrowingView(ctk.CTkFrame):
    """Borrowing records view with stats, tab filtering, and return confirmation."""

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.all_records = []
        self._build()

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        ctk.CTkLabel(scroll, text=tr("Borrowing Management"),
                     font=ctk.CTkFont(size=20, weight="bold")).pack(anchor="w", padx=24, pady=(20, 4))
        ctk.CTkLabel(scroll, text=tr("Manage book borrowing and returns"),
                     font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC).pack(anchor="w", padx=24, pady=(0, 16))

        # ── Export Button ──
        export_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        export_bar.pack(fill="x", padx=24, pady=(0, 8))

        ctk.CTkButton(
            export_bar,
            text=tr("Export"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_SUCCESS,
            hover_color=C_SUCCESS_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._export_borrowings
        ).pack(side="right")

        # ── Stats Cards Row ──
        stats_row = ctk.CTkFrame(scroll, fg_color="transparent")
        stats_row.pack(fill="x", padx=24, pady=8)

        self.stat_active = self._make_stat_card(stats_row, "0", tr("Borrowed"), "#22c55e")
        self.stat_near = self._make_stat_card(stats_row, "0", tr("Due Soon"), "#f59e0b")
        self.stat_overdue = self._make_stat_card(stats_row, "0", tr("Overdue"), "#ef4444")
        self.stat_pending = self._make_stat_card(stats_row, "0", tr("Pending Confirm"), "#8b5cf6")
        self.stat_returned = self._make_stat_card(stats_row, "0", tr("Returned"), "#6b7280")

        # ── Tab Bar ──
        tab_bar = ctk.CTkFrame(scroll, fg_color="transparent")
        tab_bar.pack(fill="x", padx=24, pady=(16, 0))

        self.tabs = {}
        tab_labels = [
            ("all", tr("All")),
            ("借出", tr("Borrowed")),
            ("临期未还", tr("Due Soon")),
            ("逾期未还", tr("Overdue")),
            ("待确认", tr("Pending Confirm")),
            ("已归还", tr("Returned")),
        ]
        for i, (key, label) in enumerate(tab_labels):
            btn = ctk.CTkButton(
                tab_bar, text=label,
                font=ctk.CTkFont(size=13, weight="bold"),
                fg_color=C_ACCENT if i == 0 else C_INPUT_BG,
                text_color=C_SIDEBAR_TEXT,
                height=36, corner_radius=10,
                command=lambda k=key: self._switch_tab(k)
            )
            btn.pack(side="left", padx=(0, 8) if i < len(tab_labels) - 1 else 0)
            self.tabs[key] = btn

        # ── Table ──
        table_frame = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=12)

        columns = ("id", "reader_name", "book_title", "borrow_date", "due_date", "return_date", "classified_status", "action")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)

        headings = {
            "id": "ID", "reader_name": tr("Reader"), "book_title": tr("Book"),
            "borrow_date": tr("Borrow Date"), "due_date": tr("Due Date"),
            "return_date": tr("Return Date"), "classified_status": tr("Status"),
            "action": tr("Action")
        }
        widths = {
            "id": 60, "reader_name": 120, "book_title": 200,
            "borrow_date": 100, "due_date": 100, "return_date": 100,
            "classified_status": 100, "action": 100
        }

        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col])

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_double_click)

        self._current_tab = "all"

    def _make_stat_card(self, parent, value, label, color):
        card = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=12, border_width=1, border_color=C_BORDER)
        card.pack(side="left", fill="x", expand=True, padx=4)

        ctk.CTkFrame(card, fg_color=color, height=3, corner_radius=0).pack(fill="x")

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=16, pady=12)

        val_label = ctk.CTkLabel(
            inner, text=str(value),
            font=ctk.CTkFont(size=24, weight="bold"), text_color=C_TEXT
        )
        val_label.pack(anchor="w")

        ctk.CTkLabel(
            inner, text=label,
            font=ctk.CTkFont(size=12), text_color=C_TEXT_SEC
        ).pack(anchor="w", pady=(4, 0))

        return {"card": card, "value": val_label}

    def _switch_tab(self, tab_id):
        self._current_tab = tab_id
        for key, btn in self.tabs.items():
            btn.configure(fg_color=C_ACCENT if key == tab_id else C_INPUT_BG)
        self._refresh_table()

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        filtered = self.all_records
        if self._current_tab != "all":
            filtered = [r for r in self.all_records if r.get("classified_status") == self._current_tab]

        for r in filtered:
            action = ""
            if r.get("classified_status") == "待确认":
                action = tr("Confirm")

            self.tree.insert("", "end", values=(
                r.get("id", ""),
                r.get("reader_name", "") or "—",
                r.get("book_title", "") or "—",
                r.get("borrow_date", "") or "—",
                r.get("due_date", "") or "—",
                r.get("return_date", "") or "—",
                r.get("classified_status", r.get("status", "")),
                action,
            ))

    def refresh(self):
        try:
            data = API.list_borrowing()
            self.all_records = data.get("data", [])

            # Update stats
            stats = {"借出": 0, "临期未还": 0, "逾期未还": 0, "已归还": 0, "待确认": 0}
            for r in self.all_records:
                cs = r.get("classified_status", r.get("status", ""))
                stats[cs] = stats.get(cs, 0) + 1

            self.stat_active["value"].configure(text=str(stats["借出"]))
            self.stat_near["value"].configure(text=str(stats["临期未还"]))
            self.stat_overdue["value"].configure(text=str(stats["逾期未还"]))
            self.stat_pending["value"].configure(text=str(stats["待确认"]))
            self.stat_returned["value"].configure(text=str(stats["已归还"]))

            self._refresh_table()
        except Exception as e:
            print(f"Borrowing refresh error: {e}")

    def _on_double_click(self, event=None):
        """Handle double-click to confirm pending return."""
        selection = self.tree.selection()
        if not selection:
            return
        item = self.tree.item(selection[0])
        values = item["values"]
        record_id = values[0]
        status = values[6]

        if status == tr("Pending Confirm") or status == "待确认":
            if messagebox.askyesno(tr("Confirm Return"),
                                   f"{tr('Confirm this return')} (ID: {record_id})?"):
                try:
                    result = API.confirm_return(record_id)
                    ToastNotification(self.winfo_toplevel(), result.get("message", tr("Return confirmed")), "success")
                    self.refresh()
                except Exception as e:
                    messagebox.showerror(tr("Error"), str(e))

    def _new_borrow(self):
        messagebox.showinfo(tr("New Borrow"), tr("New borrow dialog coming soon!"))

    def _export_borrowings(self):
        """Export borrowing records to Excel file."""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[(tr("Excel files"), "*.xlsx")]
        )
        if not save_path:
            return

        def _do_export():
            try:
                API.export_borrowings(save_path)
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), tr("Export successful"), "success"))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

        threading.Thread(target=_do_export, daemon=True).start()


class ModelsView(ctk.CTkFrame):
    """LLM Models management view."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self._build()
    
    def _build(self):
        toolbar = ctk.CTkFrame(self, fg_color="transparent")
        toolbar.pack(fill="x", padx=24, pady=(20, 0))
        
        ctk.CTkButton(
            toolbar,
            text=tr("+ Add Model"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._add_model
        ).pack(side="right")
        
        table_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        table_frame.pack(fill="both", expand=True, padx=24, pady=(16, 24))
        
        columns = ("id", "name", "provider", "model_name", "temperature", "max_tokens", "is_active", "action")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        headings = {"id": "ID", "name": tr("Name"), "provider": tr("Provider"),
                    "model_name": tr("Model"), "temperature": tr("Temp"), 
                    "max_tokens": tr("Max Tokens"), "is_active": tr("Active"),
                    "action": tr("Action")}
        widths = {"id": 60, "name": 150, "provider": 100, "model_name": 150,
                  "temperature": 70, "max_tokens": 80, "is_active": 60, "action": 80}
        
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col])
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        
        # Bind double-click to edit
        self.tree.bind("<Double-1>", self._on_double_click)
        
        # Bind right-click context menu for delete
        self.tree.bind("<Button-3>", self._show_context_menu)
        self.context_menu = None
    
    def refresh(self):
        try:
            data = API.list_models()
            for item in self.tree.get_children():
                self.tree.delete(item)
            for m in data:
                self.tree.insert("", "end", values=(
                    m.get("id", ""),
                    m.get("name", ""),
                    m.get("provider", ""),
                    m.get("model_name", ""),
                    m.get("temperature", 0.7),
                    m.get("max_tokens", 4096),
                    tr("Yes") if m.get("is_active") else tr("No"),
                    tr("Delete")
                ))
        except Exception as e:
            print(f"Models refresh error: {e}")
    
    def _on_double_click(self, event):
        """Handle double-click to edit model."""
        selection = self.tree.selection()
        if not selection:
            return
        item = self.tree.item(selection[0])
        values = item["values"]
        model_id = values[0]
        
        # Find the model data
        try:
            models = API.list_models()
            for m in models:
                if m["id"] == model_id:
                    self._edit_model(m)
                    return
        except Exception as e:
            print(f"Edit model error: {e}")
    
    def _add_model(self):
        """Switch to model edit page view."""
        app = self.winfo_toplevel()
        app.switch_view("model_edit")
    
    def _edit_model(self, model_data):
        """Switch to model edit page with existing model data."""
        app = self.winfo_toplevel()
        app.switch_view("model_edit", model_data=model_data)
    
    def _delete_model(self, event=None):
        """Delete selected model."""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        values = item["values"]
        model_id = values[0]
        model_name = values[1]
        
        if messagebox.askyesno(tr("Confirm Delete"), f"{tr('Delete model')} '{model_name}'?"):
            try:
                API.delete_model(model_id)
                self.refresh()
            except Exception as e:
                messagebox.showerror(tr("Error"), str(e))
    
    def _show_context_menu(self, event):
        """Show right-click context menu."""
        # Select the item under cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            
            # Create context menu
            if self.context_menu:
                self.context_menu.destroy()
            
            self.context_menu = tk.Menu(self.tree, tearoff=0)
            self.context_menu.add_command(label=tr("Edit"), command=lambda: self._on_double_click(None))
            self.context_menu.add_separator()
            self.context_menu.add_command(label=tr("Delete"), command=self._delete_model)
            self.context_menu.post(event.x_root, event.y_root)


class ModelEditView(ctk.CTkFrame):
    """Model add/edit page view with support for mainstream LLM providers."""
    
    PROVIDERS = [
        ("openai", "OpenAI", "https://api.openai.com/v1", ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"]),
        ("anthropic", "Anthropic", "https://api.anthropic.com", ["claude-sonnet-4-20250514", "claude-3-5-sonnet-20241022", "claude-3-haiku-20240307"]),
        ("ollama", "Ollama (Local)", "http://localhost:11434", []),
        ("azure", "Azure OpenAI", "https://YOUR-RESOURCE.openai.azure.com", ["gpt-4o", "gpt-4-turbo"]),
        ("google", "Google Gemini", "https://generativelanguage.googleapis.com", ["gemini-2.0-flash", "gemini-1.5-pro"]),
        ("deepseek", "DeepSeek", "https://api.deepseek.com", ["deepseek-chat", "deepseek-coder"]),
        ("zhipu", "Zhipu AI (智谱)", "https://open.bigmodel.cn/api/paas/v4", ["glm-4", "glm-4-flash", "glm-4v"]),
        ("custom", "Custom (OpenAI-Compatible)", "", []),
    ]
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.model_data = None
        self._build()
    
    def _build(self):
        # Title
        self.title_label = ctk.CTkLabel(
            self,
            text=tr("Add Model"),
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=C_TEXT
        )
        self.title_label.pack(pady=(24, 8))
        
        # Description
        self.desc_label = ctk.CTkLabel(
            self,
            text=tr("Configure LLM model connection"),
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        )
        self.desc_label.pack(pady=(0, 20))
        
        # Scrollable form
        canvas = ctk.CTkScrollableFrame(self, fg_color="transparent")
        canvas.pack(fill="both", expand=True, padx=24, pady=10)
        
        # Provider selection
        ctk.CTkLabel(
            canvas,
            text=tr("Provider"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(10, 5))
        
        provider_names = [p[1] for p in self.PROVIDERS]
        self.provider_var = ctk.StringVar(value=provider_names[0])
        self.provider_menu = ctk.CTkOptionMenu(
            canvas,
            variable=self.provider_var,
            values=provider_names,
            fg_color=C_CARD,
            button_color=C_ACCENT,
            button_hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            corner_radius=10,
            command=self._on_provider_change
        )
        self.provider_menu.pack(fill="x", pady=(0, 10))
        
        # Model name
        ctk.CTkLabel(
            canvas,
            text=tr("Model Name"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(10, 5))
        
        self.model_name_var = ctk.StringVar()
        self.model_name_entry = ctk.CTkEntry(
            canvas,
            textvariable=self.model_name_var,
            placeholder_text=tr("e.g., gpt-4o"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=10,
            height=36
        )
        self.model_name_entry.pack(fill="x", pady=(0, 10))
        
        # Model quick select
        self.model_quick_var = ctk.StringVar(value=tr("Select from list..."))
        self.model_quick_menu = ctk.CTkOptionMenu(
            canvas,
            variable=self.model_quick_var,
            values=[tr("Select from list...")],
            fg_color=C_CARD,
            button_color=C_ACCENT,
            button_hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            corner_radius=10,
            command=self._on_model_select
        )
        self.model_quick_menu.pack(fill="x", pady=(0, 10))
        
        # Ollama fetch status
        self.ollama_status = ctk.CTkLabel(
            canvas,
            text="",
            font=ctk.CTkFont(size=12),
            text_color=C_TEXT_SEC
        )
        self.ollama_status.pack(pady=(0, 5))
        
        # Display name
        ctk.CTkLabel(
            canvas,
            text=tr("Display Name"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(10, 5))
        
        self.name_var = ctk.StringVar()
        self.name_entry = ctk.CTkEntry(
            canvas,
            textvariable=self.name_var,
            placeholder_text=tr("e.g., My GPT-4"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=10,
            height=36
        )
        self.name_entry.pack(fill="x", pady=(0, 10))
        
        # API Base URL
        ctk.CTkLabel(
            canvas,
            text=tr("API Base URL"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(10, 5))
        
        self.api_base_var = ctk.StringVar()
        self.api_base_entry = ctk.CTkEntry(
            canvas,
            textvariable=self.api_base_var,
            placeholder_text=tr("e.g., https://api.openai.com/v1"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=10,
            height=36
        )
        self.api_base_entry.pack(fill="x", pady=(0, 10))
        
        # API Key
        ctk.CTkLabel(
            canvas,
            text=tr("API Key"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(10, 5))
        
        self.api_key_var = ctk.StringVar()
        self.api_key_entry = ctk.CTkEntry(
            canvas,
            textvariable=self.api_key_var,
            placeholder_text=tr("sk-..."),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=10,
            height=36,
            show="*"
        )
        self.api_key_entry.pack(fill="x", pady=(0, 10))
        
        # Temperature and Max Tokens row
        row = ctk.CTkFrame(canvas, fg_color="transparent")
        row.pack(fill="x", pady=(10, 10))
        
        # Temperature
        left = ctk.CTkFrame(row, fg_color="transparent")
        left.pack(side="left", fill="x", expand=True)
        
        ctk.CTkLabel(
            left,
            text=tr("Temperature"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(0, 5))
        
        self.temp_var = ctk.StringVar(value="0.7")
        ctk.CTkEntry(
            left,
            textvariable=self.temp_var,
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=10,
            height=36,
            width=100
        ).pack(fill="x")
        
        # Max Tokens
        right = ctk.CTkFrame(row, fg_color="transparent")
        right.pack(side="right", fill="x", expand=True, padx=(20, 0))
        
        ctk.CTkLabel(
            right,
            text=tr("Max Tokens"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(fill="x", pady=(0, 5))
        
        self.max_tokens_var = ctk.StringVar(value="4096")
        ctk.CTkEntry(
            right,
            textvariable=self.max_tokens_var,
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=10,
            height=36,
            width=100
        ).pack(fill="x")
        
        # Active toggle
        self.active_var = ctk.BooleanVar(value=True)
        active_check = ctk.CTkCheckBox(
            canvas,
            text=tr("Enable this model"),
            variable=self.active_var,
            font=ctk.CTkFont(size=13),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            corner_radius=10
        )
        active_check.pack(anchor="w", pady=(10, 10))
        
        # Test connection button
        self.test_btn = ctk.CTkButton(
            canvas,
            text=tr("Test Connection"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            border_width=1,
            border_color=C_BUTTON_BORDER,
            corner_radius=10,
            height=36,
            command=self._test_connection
        )
        self.test_btn.pack(fill="x", pady=(10, 0))
        
        # Test result label
        self.test_result_label = ctk.CTkLabel(
            canvas,
            text="",
            font=ctk.CTkFont(size=12),
            text_color=C_TEXT_SEC
        )
        self.test_result_label.pack(pady=(5, 0))
        
        # Bottom buttons bar
        bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        bottom_bar.pack(fill="x", padx=24, pady=(10, 24))
        
        self.save_status = ctk.CTkLabel(
            bottom_bar,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=C_TEXT_SEC
        )
        self.save_status.pack(side="left")
        
        btn_frame = ctk.CTkFrame(bottom_bar, fg_color="transparent")
        btn_frame.pack(side="right")
        
        ctk.CTkButton(
            btn_frame,
            text=tr("Back to Models"),
            font=ctk.CTkFont(size=13),
            fg_color=C_INPUT_BG,
            hover_color=C_BORDER,
            text_color=C_SIDEBAR_TEXT,
            width=120,
            height=36,
            corner_radius=10,
            command=self._go_back
        ).pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(
            btn_frame,
            text=tr("Save"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            width=100,
            height=36,
            corner_radius=10,
            command=self._save
        ).pack(side="right")
    
    def refresh(self, model_data=None):
        """Reset the view when navigated to. model_data=None means add mode."""
        self.model_data = model_data
        
        if model_data:
            self.title_label.configure(text=tr("Edit Model"))
            self.desc_label.configure(text=tr("Modify model configuration"))
            self._load_model_data(model_data)
        else:
            self.title_label.configure(text=tr("Add Model"))
            self.desc_label.configure(text=tr("Configure LLM model connection"))
            # Reset form
            self.name_var.set("")
            self.model_name_var.set("")
            self.api_base_var.set("")
            self.api_key_var.set("")
            self.temp_var.set("0.7")
            self.max_tokens_var.set("4096")
            self.active_var.set(True)
            self.provider_var.set(self.PROVIDERS[0][1])
            self.model_quick_var.set(tr("Select from list..."))
            self.ollama_status.configure(text="")
            self.test_result_label.configure(text="")
            self.save_status.configure(text="")
            self._on_provider_change(self.PROVIDERS[0][1])
    
    def _go_back(self):
        app = self.winfo_toplevel()
        app.switch_view("models")
    
    def _get_provider_key(self):
        name = self.provider_var.get()
        for key, display, _, _ in self.PROVIDERS:
            if display == name:
                return key
        return "openai"
    
    def _on_provider_change(self, value):
        for p_key, p_name, default_base, default_models in self.PROVIDERS:
            if p_name == value:
                if not self.api_base_var.get() or self.api_base_var.get() in [p[2] for p in self.PROVIDERS if p[1] != value]:
                    self.api_base_var.set(default_base)
                
                if p_key == "ollama":
                    # Auto-fetch Ollama models
                    self.ollama_status.configure(text=tr("Fetching Ollama models..."), text_color=C_TEXT_SEC)
                    threading.Thread(target=self._fetch_ollama_models, daemon=True).start()
                else:
                    self.ollama_status.configure(text="")
                    if default_models:
                        self.model_quick_menu.configure(values=[tr("Select from list...")] + default_models)
                    else:
                        self.model_quick_menu.configure(values=[tr("Select from list...")])
                    self.model_quick_var.set(tr("Select from list..."))
                break
    
    def _fetch_ollama_models(self):
        """Fetch available models from local Ollama instance."""
        try:
            api_base = self.api_base_var.get() or "http://localhost:11434"
            models = API.fetch_ollama_models(api_base)
            model_names = [m.get("name", "") for m in models if m.get("name")]
            
            if model_names:
                self.after(0, lambda: self._set_ollama_models(model_names))
            else:
                self.after(0, lambda: self.ollama_status.configure(
                    text=tr("No models found. Make sure Ollama is running."), text_color="#ef4444"))
        except Exception as e:
            err_msg = str(e)
            self.after(0, lambda: self.ollama_status.configure(
                text=f"{tr('Failed to connect to Ollama')}: {err_msg}", text_color="#ef4444"))
    
    def _set_ollama_models(self, model_names):
        self.model_quick_menu.configure(values=[tr("Select from list...")] + model_names)
        self.model_quick_var.set(tr("Select from list..."))
        self.ollama_status.configure(
            text=f"{tr('Found')} {len(model_names)} {tr('model(s)')}", text_color=C_SUCCESS)
    
    def _on_model_select(self, value):
        if value != tr("Select from list..."):
            self.model_name_var.set(value)
            if not self.name_var.get():
                provider = self.provider_var.get()
                self.name_var.set(f"{provider} {value}")
    
    def _load_model_data(self, data):
        provider_key = data.get("provider", "openai")
        for p_key, p_name, _, _ in self.PROVIDERS:
            if p_key == provider_key:
                self.provider_var.set(p_name)
                break
        
        self.model_name_var.set(data.get("model_name", ""))
        self.name_var.set(data.get("name", ""))
        self.api_base_var.set(data.get("api_base", ""))
        self.api_key_var.set(data.get("api_key", ""))
        self.temp_var.set(str(data.get("temperature", 0.7)))
        self.max_tokens_var.set(str(data.get("max_tokens", 4096)))
        self.active_var.set(bool(data.get("is_active", True)))
        
        # If ollama, fetch models
        if provider_key == "ollama":
            self.ollama_status.configure(text=tr("Fetching Ollama models..."), text_color=C_TEXT_SEC)
            threading.Thread(target=self._fetch_ollama_models, daemon=True).start()
    
    def _test_connection(self):
        self.test_btn.configure(state="disabled", text=tr("Testing..."))
        self.test_result_label.configure(text="", text_color=C_TEXT_SEC)
        threading.Thread(target=self._do_test, daemon=True).start()
    
    def _do_test(self):
        """Test model connection. For new models, clean up temporary model after test."""
        try:
            data = self._collect_data()
            if not data.get("name") or not data.get("model_name"):
                self.after(0, lambda: self._show_test_result(False, tr("Please fill in name and model name")))
                return

            is_new_model = not (self.model_data and self.model_data.get("id"))

            if is_new_model:
                # Create temporary model for testing
                result = API.create_model(data)
                model_id = result.get("id")
            else:
                # Update existing model and test
                API.update_model(self.model_data["id"], data)
                model_id = self.model_data["id"]

            try:
                result = API.test_model(model_id)

                if result.get("success"):
                    resp = result.get("response", "")[:100]
                    self.after(0, lambda: self._show_test_result(True, f"{tr('Connection successful!')}\n{resp}"))
                else:
                    self.after(0, lambda: self._show_test_result(False, result.get("message", tr("Connection failed"))))
            finally:
                # Clean up temporary model for new models
                if is_new_model and model_id:
                    try:
                        API.delete_model(model_id)
                    except Exception:
                        pass  # Ignore cleanup errors
        except Exception as e:
            self.after(0, lambda: self._show_test_result(False, str(e)))
    
    def _show_test_result(self, success, message):
        self.test_btn.configure(state="normal", text=tr("Test Connection"))
        color = C_SUCCESS if success else "#ef4444"
        self.test_result_label.configure(text=message, text_color=color)
    
    def _collect_data(self):
        return {
            "name": self.name_var.get() or f"{self.provider_var.get()} {self.model_name_var.get()}",
            "provider": self._get_provider_key(),
            "model_name": self.model_name_var.get(),
            "api_base": self.api_base_var.get(),
            "api_key": self.api_key_var.get(),
            "temperature": float(self.temp_var.get()) if self.temp_var.get() else 0.7,
            "max_tokens": int(self.max_tokens_var.get()) if self.max_tokens_var.get() else 4096,
            "is_active": self.active_var.get(),
        }
    
    def _save(self):
        data = self._collect_data()
        
        if not data["name"] or not data["model_name"]:
            self._show_toast(tr("Please fill in display name and model name"), False)
            return
        
        try:
            # Check for duplicate name
            existing = API.list_models()
            for m in existing:
                if m["name"] == data["name"]:
                    if self.model_data and m["id"] == self.model_data.get("id"):
                        continue  # Same model being edited, skip
                    self._show_toast(tr("Model name already exists"), False)
                    return
            
            if self.model_data and self.model_data.get("id"):
                API.update_model(self.model_data["id"], data)
            else:
                API.create_model(data)
            
            self._show_toast(tr("Saved successfully!"), True)
            
        except Exception as e:
            self._show_toast(f"{tr('Error')}: {e}", False)
    
    def _show_toast(self, message, success):
        """Show a floating toast notification centered upper-middle of the window."""
        # Remove existing toast if any
        if hasattr(self, "_toast_frame") and self._toast_frame:
            try:
                self._toast_frame.destroy()
            except Exception:
                pass
        
        self._toast_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=12,
                                          border_width=1, border_color=C_BORDER)
        self._toast_frame.place(relx=0.5, rely=0.15, anchor="center")
        
        color = C_SUCCESS if success else "#ef4444"
        ctk.CTkLabel(
            self._toast_frame,
            text=message,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=color
        ).pack(padx=24, pady=12)
        
        # Auto-dismiss after 2 seconds
        self.after(2000, self._dismiss_toast)
    
    def _dismiss_toast(self):
        if hasattr(self, "_toast_frame") and self._toast_frame:
            try:
                self._toast_frame.destroy()
            except Exception:
                pass
            self._toast_frame = None


class AIAssistantView(ctk.CTkFrame):
    """AI Assistant chat view - TRAE style with Agent tool calling."""
    
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.current_conv_id = None
        self.current_model_id = None
        self._build()
    
    def _build(self):
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=24, pady=(20, 0))
        
        ctk.CTkLabel(
            header,
            text=tr("AI Assistant"),
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=C_TEXT
        ).pack(side="left")
        
        # Model selector
        self.model_var = ctk.StringVar(value=tr("Select Model"))
        self.model_menu = ctk.CTkOptionMenu(
            header,
            variable=self.model_var,
            values=[tr("Loading...")],
            fg_color=C_CARD,
            button_color=C_ACCENT,
            button_hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            corner_radius=10,
            width=180,
            command=self._on_model_change
        )
        self.model_menu.pack(side="right")
        
        ctk.CTkButton(
            header,
            text=tr("New Chat"),
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=36,
            corner_radius=10,
            command=self._new_chat
        ).pack(side="right", padx=(0, 12))
        
        # Main chat area
        chat_container = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        chat_container.pack(fill="both", expand=True, padx=24, pady=(16, 0))
        
        # Conversation list
        conv_frame = ctk.CTkFrame(chat_container, fg_color=C_INPUT_BG, corner_radius=12, width=220)
        conv_frame.pack(side="left", fill="y")
        conv_frame.pack_propagate(False)
        
        ctk.CTkLabel(
            conv_frame,
            text=tr("Conversations"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT
        ).pack(fill="x", padx=16, pady=(12, 8))
        
        self.conv_list = ctk.CTkScrollableFrame(conv_frame, fg_color="transparent")
        self.conv_list.pack(fill="both", expand=True, padx=8)
        
        # Messages area
        msg_frame = ctk.CTkFrame(chat_container, fg_color="transparent")
        msg_frame.pack(side="left", fill="both", expand=True, padx=(8, 0))
        
        self.messages_box = ctk.CTkScrollableFrame(msg_frame, fg_color="transparent", corner_radius=12)
        self.messages_box.pack(fill="both", expand=True, padx=12, pady=12)
        
        # Input area
        input_frame = ctk.CTkFrame(self, fg_color="transparent")
        input_frame.pack(fill="x", padx=24, pady=(16, 24))
        
        self.chat_input = ctk.CTkEntry(
            input_frame,
            placeholder_text=tr("Enter your instruction here..."),
            fg_color=C_CARD,
            border_width=1,
            border_color=C_BORDER,
            corner_radius=16,
            height=48,
            font=ctk.CTkFont(size=14)
        )
        self.chat_input.pack(side="left", fill="x", expand=True)
        self.chat_input.bind("<Return>", lambda e: self._send_message())
        
        ctk.CTkButton(
            input_frame,
            text=tr("Send"),
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=48,
            width=80,
            corner_radius=16,
            command=self._send_message
        ).pack(side="right", padx=(12, 0))
        
        # Quick actions
        quick_actions = ctk.CTkFrame(self, fg_color="transparent")
        quick_actions.pack(fill="x", padx=24, pady=(0, 16))
        
        actions = [tr("List Books"), tr("Search Books"), tr("List Readers"), tr("Statistics")]
        for action in actions:
            ctk.CTkButton(
                quick_actions,
                text=action,
                font=ctk.CTkFont(size=12),
                fg_color=C_CARD,
                text_color=C_SIDEBAR_TEXT,
                hover_color=C_INPUT_BG,
                border_width=1,
                border_color=C_BORDER,
                height=32,
                corner_radius=8,
                command=lambda a=action: self._quick_action(a)
            ).pack(side="left", padx=(0, 8))
    
    def _on_model_change(self, value):
        """Update model_id when model selection changes."""
        try:
            models = API.list_models()
            for m in models:
                if m["name"] == value:
                    self.current_model_id = m["id"]
                    break
        except Exception:
            pass
    
    def refresh(self):
        try:
            models = API.list_models()
            if models:
                self.model_menu.configure(values=[m["name"] for m in models])
                self.model_var.set(models[0]["name"])
                self.current_model_id = models[0]["id"]
            
            # Load agent conversations
            convs = API.list_agent_conversations()
            for w in self.conv_list.winfo_children():
                w.destroy()
            
            for conv in convs:
                title = conv.get("title", tr("New Chat"))[:20]
                row = ctk.CTkFrame(self.conv_list, fg_color="transparent")
                row.pack(fill="x", pady=2)

                btn = ctk.CTkButton(
                    row,
                    text=title,
                    font=ctk.CTkFont(size=12),
                    fg_color="transparent",
                    text_color=C_SIDEBAR_TEXT,
                    hover_color=C_BORDER,
                    anchor="w",
                    height=36,
                    corner_radius=8,
                    command=lambda c=conv: self._select_conv(c)
                )
                btn.pack(side="left", fill="x", expand=True)

                del_btn = ctk.CTkButton(
                    row,
                    text="✕",
                    font=ctk.CTkFont(size=12),
                    fg_color="transparent",
                    text_color=C_TEXT_SEC,
                    hover_color=C_DANGER,
                    hover_text_color="#ffffff",
                    width=28,
                    height=28,
                    corner_radius=6,
                    command=lambda cid=conv["id"]: self._delete_conv(cid)
                )
                del_btn.pack(side="right", padx=(4, 0))
            
            if not convs:
                ctk.CTkLabel(
                    self.conv_list,
                    text=tr("No conversations"),
                    font=ctk.CTkFont(size=12),
                    text_color=C_TEXT_SEC
                ).pack(pady=8)
                
        except Exception as e:
            print(f"AI refresh error: {e}")
    
    def _new_chat(self):
        self.current_conv_id = None
        for w in self.messages_box.winfo_children():
            w.destroy()
        ctk.CTkLabel(
            self.messages_box,
            text=tr("Start a new conversation!"),
            font=ctk.CTkFont(size=14),
            text_color=C_TEXT_SEC
        ).pack(pady=40)
    
    def _select_conv(self, conv):
        try:
            self.current_conv_id = conv["id"]
            conv_detail = API.get_agent_conversation(conv["id"])
            for w in self.messages_box.winfo_children():
                w.destroy()
            for msg in conv_detail.get("messages", []):
                self._add_message(msg.get("role"), msg.get("content"))
        except Exception as e:
            print(f"Select conversation error: {e}")

    def _delete_conv(self, conv_id):
        """Delete a conversation after confirmation."""
        if not messagebox.askyesno(tr("Confirm Delete"), tr("Are you sure you want to delete this conversation?")):
            return

        def do_delete():
            try:
                API.delete_agent_conversation(conv_id)
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), tr("Conversation deleted"), "success"))
                self.after(0, self.refresh)
                # If deleted conversation was selected, clear messages
                if self.current_conv_id == conv_id:
                    self.after(0, self._new_chat)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))

        threading.Thread(target=do_delete, daemon=True).start()
    
    def _send_message(self):
        text = self.chat_input.get().strip()
        if not text:
            return

        if not self.current_model_id:
            messagebox.showwarning(tr("Error"), tr("Please select a model first"))
            return

        self.chat_input.delete(0, "end")
        self._add_message("user", text)

        # Create streaming message bubble
        bubble = ctk.CTkFrame(self.messages_box, fg_color=C_INPUT_BG, corner_radius=12)
        bubble.pack(anchor="w", padx=8, pady=4, fill="x")

        msg_label = ctk.CTkLabel(
            bubble,
            text="",
            font=ctk.CTkFont(size=14),
            text_color=C_TEXT,
            anchor="w",
            wraplength=600,
            justify="left"
        )
        msg_label.pack(padx=12, pady=10)

        def stream_response():
            full_response = ""
            try:
                # Use streaming API
                for chunk_json in API.agent_chat_stream(
                    message=text,
                    model_id=self.current_model_id,
                    conversation_id=self.current_conv_id
                ):
                    import json
                    data = json.loads(chunk_json)

                    if data.get("type") == "chunk":
                        chunk = data.get("content", "")
                        full_response += chunk
                        # Update UI with accumulated text (typing effect)
                        self.after(0, lambda t=full_response: msg_label.configure(text=t))
                        # Auto-scroll to bottom
                        self.after(0, lambda: self.messages_box._parent_canvas.yview_moveto(1.0))

                    elif data.get("type") == "done":
                        self.current_conv_id = data.get("conversation_id")
                        # Final update
                        self.after(0, lambda t=full_response: msg_label.configure(text=t))
                        self.after(0, self.refresh)

                    elif data.get("type") == "error":
                        error_msg = data.get("content", "Unknown error")
                        self.after(0, lambda t=error_msg: msg_label.configure(text=t, text_color="#dc2626"))

            except Exception as e:
                error_text = f"Error: {str(e)}"
                self.after(0, lambda: msg_label.configure(text=error_text, text_color="#dc2626"))

        threading.Thread(target=stream_response, daemon=True).start()
    
    def _add_message(self, role, content):
        is_user = role == "user"
        bg_color = "#eef2ff" if is_user else C_INPUT_BG

        bubble = ctk.CTkFrame(self.messages_box, fg_color=bg_color, corner_radius=12)
        bubble.pack(anchor="e" if is_user else "w", padx=8, pady=4, fill="x")

        # Try to parse and render as table or chart
        parsed = self._try_parse_table(content)
        if parsed:
            if parsed.get("type") == "statistics":
                self._render_statistics(bubble, parsed)
            else:
                self._render_table(bubble, parsed)
        else:
            ctk.CTkLabel(
                bubble,
                text=content,
                font=ctk.CTkFont(size=14),
                text_color=C_TEXT,
                anchor="w",
                wraplength=600,
                justify="left"
            ).pack(padx=12, pady=10)
    
    def _try_parse_table(self, content):
        """Try to parse agent response into structured table data.
        Returns dict with 'headers', 'rows', 'title' or None.
        """
        text = content.strip()

        # Parse "Found N books:" format
        if text.startswith("Found ") and " books:" in text:
            return self._parse_book_list(text)

        # Parse "Readers (N):" format
        if text.startswith("Readers (") and "):" in text:
            return self._parse_reader_list(text)

        # Parse "Borrowing Records (N):" format
        if text.startswith("Borrowing Records ("):
            return self._parse_borrowing_list(text)

        # Parse "Library Statistics" format
        if "Library Statistics" in text or "library statistics" in text.lower():
            return self._parse_statistics(text)

        return None

    def _parse_statistics(self, text):
        """Parse statistics response into chart data."""
        lines = text.split("\n")
        total_books = 0
        total_readers = 0
        borrowed = 0
        categories = {}

        in_category = False
        for line in lines:
            line = line.strip()
            if "Total books:" in line:
                total_books = int(line.split("Total books:")[1].strip())
            elif "Total readers:" in line:
                total_readers = int(line.split("Total readers:")[1].strip())
            elif "Currently borrowed:" in line:
                borrowed = int(line.split("Currently borrowed:")[1].strip())
            elif "By category:" in line:
                in_category = True
            elif in_category and ":" in line:
                parts = line.split(":")
                if len(parts) == 2:
                    cat_name = parts[0].strip()
                    cat_count = int(parts[1].strip())
                    categories[cat_name] = cat_count

        return {
            "type": "statistics",
            "total_books": total_books,
            "total_readers": total_readers,
            "borrowed": borrowed,
            "categories": categories
        }
    
    def _parse_book_list(self, text):
        """Parse book list response into table data."""
        # Extract title line
        title_match = text.split("\n")[0]  # "Found 14 books:"

        entries = text.split("\n\n")
        rows = []
        for entry in entries:
            lines = entry.strip().split("\n")
            if not lines:
                continue
            first = lines[0]
            # Extract ID and title: "[1] Introduction to Algorithms"
            title = first
            book_id = ""
            if first.startswith("["):
                bracket_end = first.index("]")
                book_id = first[1:bracket_end].strip()
                title = first[bracket_end+1:].strip()

            info = {}
            for line in lines[1:]:
                line = line.strip()
                if "Author:" in line:
                    info["authors"] = line.split("Author:")[1].split("|")[0].strip()
                if "ISBN:" in line:
                    info["isbn"] = line.split("ISBN:")[1].split("|")[0].strip()
                if "Category:" in line:
                    info["category"] = line.split("Category:")[1].split("|")[0].strip()
                if "Call#:" in line:
                    info["call_number"] = line.split("Call#:")[1].split("|")[0].strip()
                if "Publisher:" in line:
                    info["publisher"] = line.split("Publisher:")[1].split("|")[0].strip()
                if "Year:" in line:
                    info["year"] = line.split("Year:")[1].split("|")[0].strip()
                if "Copies:" in line:
                    copies_part = line.split("Copies:")[1].strip()
                    info["copies"] = copies_part.split("Available:")[0].strip()
                    if "Available:" in copies_part:
                        info["available"] = copies_part.split("Available:")[1].strip()

            rows.append({
                "id": book_id,
                "title": title,
                **info
            })

        if not rows:
            return None

        return {
            "title": title_match,
            "headers": ["#", "书名", "作者", "ISBN", "分类", "索书号", "出版社", "年份", "馆藏", "可借"],
            "rows": rows,
            "type": "books"
        }
    
    def _parse_reader_list(self, text):
        """Parse reader list response into table data."""
        title_match = text.split("\n")[0]
        entries = text.split("\n")[1:]  # Skip title line
        rows = []
        for line in entries:
            line = line.strip()
            if not line or not line.startswith("["):
                continue
            # "[1] 张三 (card: 12345, type: 身份证)"
            bracket_end = line.index("]")
            reader_id = line[1:bracket_end].strip()
            rest = line[bracket_end+1:].strip()
            name = rest.split("(")[0].strip() if "(" in rest else rest
            card = ""
            rtype = ""
            if "card:" in rest:
                card = rest.split("card:")[1].split(",")[0].strip()
            if "type:" in rest:
                rtype = rest.split("type:")[1].split(")")[0].strip()
            rows.append({"id": reader_id, "name": name, "card": card, "type": rtype})
        
        if not rows:
            return None
        
        return {
            "title": title_match,
            "headers": ["#", "姓名", "卡号", "类型"],
            "rows": rows,
            "type": "readers"
        }
    
    def _parse_borrowing_list(self, text):
        """Parse borrowing records response into table data."""
        title_match = text.split("\n")[0]
        rows = []
        lines = text.split("\n")
        i = 1
        while i < len(lines):
            line = lines[i].strip()
            if line.startswith("["):
                # "  [1] 张三 borrowed '三体'"
                bracket_end = line.index("]")
                rec_id = line[1:bracket_end].strip()
                rest = line[bracket_end+1:].strip()
                reader_name = rest.split(" borrowed ")[0].strip() if " borrowed " in rest else ""
                book_name = ""
                if "borrowed '" in rest:
                    book_name = rest.split("borrowed '")[1].split("'")[0]
                
                # Next line has details
                detail = ""
                if i + 1 < len(lines):
                    detail = lines[i+1].strip()
                    i += 1
                
                copy_id = ""
                borrow_date = ""
                due_date = ""
                status = ""
                if "Copy#" in detail:
                    copy_id = detail.split("Copy#")[1].split("|")[0].strip()
                if "Borrowed:" in detail:
                    borrow_date = detail.split("Borrowed:")[1].split("|")[0].strip()
                if "Due:" in detail:
                    due_date = detail.split("Due:")[1].split("|")[0].strip()
                if "Status:" in detail:
                    status = detail.split("Status:")[1].strip()
                
                rows.append({
                    "id": rec_id,
                    "reader": reader_name,
                    "book": book_name,
                    "copy": copy_id,
                    "borrow_date": borrow_date,
                    "due_date": due_date,
                    "status": status
                })
            i += 1
        
        if not rows:
            return None
        
        return {
            "title": title_match,
            "headers": ["#", "读者", "图书", "副本", "借阅日期", "到期日期", "状态"],
            "rows": rows,
            "type": "borrowing"
        }
    
    def _render_table(self, parent, data):
        """Render parsed data as a table in the chat bubble."""
        # Title
        ctk.CTkLabel(
            parent,
            text=data["title"],
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(padx=12, pady=(10, 4), anchor="w")
        
        # Table frame
        table_frame = ctk.CTkFrame(parent, fg_color="transparent")
        table_frame.pack(padx=12, pady=4, fill="x")
        
        headers = data["headers"]
        rows = data["rows"]
        dtype = data["type"]
        
        # Header row
        header_frame = ctk.CTkFrame(table_frame, fg_color="#f0f0f5", corner_radius=6)
        header_frame.pack(fill="x", pady=(0, 2))
        
        # Calculate column widths based on content
        col_widths = self._calc_col_widths(headers, rows, dtype)
        
        for j, h in enumerate(headers):
            ctk.CTkLabel(
                header_frame,
                text=h,
                font=ctk.CTkFont(size=11, weight="bold"),
                text_color=C_TEXT,
                width=col_widths[j],
                anchor="w"
            ).pack(side="left", padx=4, pady=4)
        
        # Data rows (limit to 20 for performance)
        max_rows = min(len(rows), 20)
        for i in range(max_rows):
            row = rows[i]
            row_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
            row_frame.pack(fill="x", pady=1)
            
            for j, h in enumerate(headers):
                key = self._header_to_key(h, dtype)
                val = row.get(key, "")
                ctk.CTkLabel(
                    row_frame,
                    text=str(val),
                    font=ctk.CTkFont(size=11),
                    text_color=C_TEXT_SEC,
                    width=col_widths[j],
                    anchor="w"
                ).pack(side="left", padx=4, pady=2)
        
        if len(rows) > 20:
            ctk.CTkLabel(
                parent,
                text=f"... 共 {len(rows)} 条，仅显示前 20 条",
                font=ctk.CTkFont(size=11),
                text_color=C_TEXT_SEC,
                anchor="w"
            ).pack(padx=12, pady=(2, 10), anchor="w")

    def _render_statistics(self, parent, data):
        """Render statistics as summary cards + canvas line chart."""
        # Title
        ctk.CTkLabel(
            parent,
            text="Library Statistics",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(padx=12, pady=(10, 4), anchor="w")

        # Summary cards row
        cards_frame = ctk.CTkFrame(parent, fg_color="transparent")
        cards_frame.pack(fill="x", padx=12, pady=4)

        stats = [
            ("Total Books", str(data["total_books"]), "#6366f1"),
            ("Total Readers", str(data["total_readers"]), "#22c55e"),
            ("Borrowed", str(data["borrowed"]), "#f59e0b"),
        ]

        for label, value, color in stats:
            card = ctk.CTkFrame(cards_frame, fg_color="#f8f9fa", corner_radius=10, border_width=1, border_color="#e5e7eb")
            card.pack(side="left", fill="x", expand=True, padx=4)

            # Color strip
            ctk.CTkFrame(card, fg_color=color, height=3, corner_radius=0).pack(fill="x")

            ctk.CTkLabel(
                card, text=value,
                font=ctk.CTkFont(size=22, weight="bold"),
                text_color=C_TEXT
            ).pack(pady=(8, 0))
            ctk.CTkLabel(
                card, text=label,
                font=ctk.CTkFont(size=11),
                text_color=C_TEXT_SEC
            ).pack(pady=(0, 8))

        # Canvas line chart for categories
        categories = data.get("categories", {})
        if categories:
            chart_frame = ctk.CTkFrame(parent, fg_color="#fafafa", corner_radius=10, border_width=1, border_color="#e5e7eb")
            chart_frame.pack(fill="x", padx=12, pady=(8, 10))

            ctk.CTkLabel(
                chart_frame,
                text="Books by Category",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=C_TEXT,
                anchor="w"
            ).pack(padx=12, pady=(10, 4), anchor="w")

            # Draw chart on canvas
            canvas = tk.Canvas(chart_frame, bg="#fafafa", height=220, highlightthickness=0)
            canvas.pack(fill="x", padx=12, pady=(0, 10))

            # Delay drawing until canvas is rendered
            self.after(50, lambda c=canvas, cats=categories: self._draw_line_chart(c, cats))

    def _draw_line_chart(self, canvas, categories):
        """Draw a line chart on tkinter Canvas showing category distribution."""
        canvas.update_idletasks()
        w = canvas.winfo_width()
        h = canvas.winfo_height()

        if w < 50:
            w = 500
        if h < 50:
            h = 200

        # Margins
        margin_left = 50
        margin_right = 20
        margin_top = 20
        margin_bottom = 40

        chart_w = w - margin_left - margin_right
        chart_h = h - margin_top - margin_bottom

        cats = list(categories.keys())
        vals = list(categories.values())
        n = len(cats)

        if n == 0:
            return

        max_val = max(vals) if vals else 1
        if max_val == 0:
            max_val = 1

        # Chart colors
        line_color = "#6366f1"
        dot_color = "#6366f1"
        grid_color = "#e5e7eb"
        text_color = "#6b7280"

        # Draw horizontal grid lines and Y-axis labels
        num_grid = 4
        for i in range(num_grid + 1):
            y = margin_top + chart_h - (i / num_grid) * chart_h
            val = round((i / num_grid) * max_val)
            canvas.create_line(margin_left, y, w - margin_right, y, fill=grid_color, width=1)
            canvas.create_text(margin_left - 8, y, text=str(val), fill=text_color, font=("Arial", 9), anchor="e")

        # Calculate data points
        points = []
        for i in range(n):
            x = margin_left + (i + 0.5) * (chart_w / n)
            y = margin_top + chart_h - (vals[i] / max_val) * chart_h
            points.append((x, y))

        # Draw area fill under line
        if len(points) >= 2:
            area_coords = []
            for px, py in points:
                area_coords.extend([px, py])
            area_coords.extend([points[-1][0], margin_top + chart_h])
            area_coords.extend([points[0][0], margin_top + chart_h])
            canvas.create_polygon(area_coords, fill="#6366f1", outline="", stipple="gray25")

        # Draw line
        for i in range(len(points) - 1):
            canvas.create_line(points[i][0], points[i][1], points[i+1][0], points[i+1][1],
                               fill=line_color, width=2.5)

        # Draw dots and X-axis labels
        for i, (px, py) in enumerate(points):
            # Dot
            canvas.create_oval(px - 4, py - 4, px + 4, py + 4, fill=dot_color, outline="white", width=2)
            # Value on top of dot
            canvas.create_text(px, py - 10, text=str(vals[i]), fill=line_color, font=("Arial", 9, "bold"))
            # X-axis label
            label = cats[i]
            if len(label) > 8:
                label = label[:8] + ".."
            canvas.create_text(px, margin_top + chart_h + 16, text=label, fill=text_color, font=("Arial", 9), anchor="n")
    
    def _header_to_key(self, header, dtype):
        """Map table header to row data key."""
        mapping = {
            "books": {
                "#": "id", "书名": "title", "作者": "authors", "ISBN": "isbn",
                "分类": "category", "索书号": "call_number", "出版社": "publisher",
                "年份": "year", "馆藏": "copies", "可借": "available"
            },
            "readers": {
                "#": "id", "姓名": "name", "卡号": "card", "类型": "type"
            },
            "borrowing": {
                "#": "id", "读者": "reader", "图书": "book", "副本": "copy",
                "借阅日期": "borrow_date", "到期日期": "due_date", "状态": "status"
            }
        }
        return mapping.get(dtype, {}).get(header, header.lower())
    
    def _calc_col_widths(self, headers, rows, dtype):
        """Calculate column widths based on content."""
        widths = []
        for j, h in enumerate(headers):
            key = self._header_to_key(h, dtype)
            max_len = len(str(h))
            for row in rows[:20]:
                val_len = len(str(row.get(key, "")))
                if val_len > max_len:
                    max_len = val_len
            # Convert char count to pixels (approx 8px per char for Chinese)
            widths.append(min(max(max_len * 8 + 16, 50), 200))
        return widths
    
    def _quick_action(self, action):
        prompts = {
            "List Books": tr("列出所有图书"),
            "Search Books": tr("搜索关于人工智能的图书"),
            "List Readers": tr("列出所有读者"),
            "Statistics": tr("显示图书馆统计信息"),
        }
        self.chat_input.insert(0, prompts.get(action, action))


class SettingsView(ctk.CTkFrame):
    """Settings view with appearance, language, admin, and LLM configuration."""

    def __init__(self, master, app_ref=None, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.app_ref = app_ref
        self._build()

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        ctk.CTkLabel(scroll, text=tr("Settings"),
                     font=ctk.CTkFont(size=24, weight="bold")).pack(anchor="w", padx=24, pady=(20, 4))
        ctk.CTkLabel(scroll, text=tr("System administration and configuration"),
                     font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC).pack(anchor="w", padx=24, pady=(0, 16))

        # ── Appearance Section ───────────────────────────────
        appearance_card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        appearance_card.pack(fill="x", padx=24, pady=8)

        ctk.CTkLabel(appearance_card, text=tr("Appearance"),
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT
                     ).pack(anchor="w", padx=20, pady=(16, 8))

        # Theme row
        theme_row = ctk.CTkFrame(appearance_card, fg_color="transparent")
        theme_row.pack(fill="x", padx=20, pady=8)
        ctk.CTkLabel(theme_row, text=tr("Theme"), font=ctk.CTkFont(size=14)).pack(side="left")
        self.theme_var = ctk.StringVar(value=tr("System"))
        theme_menu = ctk.CTkOptionMenu(theme_row, variable=self.theme_var,
                                       values=[tr("System"), tr("Light"), tr("Dark")],
                                       fg_color=C_INPUT_BG, button_color=C_ACCENT,
                                       text_color=C_TEXT,
                                       corner_radius=10, height=36,
                                       command=self._set_theme)
        theme_menu.pack(side="right")

        # Language row
        lang_row = ctk.CTkFrame(appearance_card, fg_color="transparent")
        lang_row.pack(fill="x", padx=20, pady=(0, 16))
        ctk.CTkLabel(lang_row, text=tr("Language"), font=ctk.CTkFont(size=14)).pack(side="left")
        self.lang_var = ctk.StringVar(value="中文" if get_language() == "zh" else "English")
        lang_menu = ctk.CTkOptionMenu(lang_row, variable=self.lang_var,
                                     values=["中文", "English"],
                                     fg_color=C_INPUT_BG, button_color=C_ACCENT,
                                     text_color=C_TEXT,
                                     corner_radius=10, height=36,
                                     command=self._set_language)
        lang_menu.pack(side="right")

        # ── Administrators Section ───────────────────────────
        admins_card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        admins_card.pack(fill="x", padx=24, pady=8)

        ctk.CTkLabel(admins_card, text=tr("Administrators"),
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT
                     ).pack(anchor="w", padx=20, pady=(16, 8))

        self.admin_list = ctk.CTkFrame(admins_card, fg_color="transparent")
        self.admin_list.pack(fill="x", padx=20, pady=4)

        ctk.CTkButton(admins_card, text=tr("+ Add Admin"), width=120, height=36,
                      fg_color=C_ACCENT, hover_color=C_ACCENT_HOVER,
                      text_color=C_SIDEBAR_TEXT, corner_radius=10,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=self._add_admin).pack(anchor="w", padx=20, pady=(8, 16))

        # ── LLM Configuration Section ─────────────────────────
        llm_card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        llm_card.pack(fill="x", padx=24, pady=8)

        ctk.CTkLabel(llm_card, text=tr("LLM Configuration"),
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT
                     ).pack(anchor="w", padx=20, pady=(16, 8))

        # Provider row
        prov_row = ctk.CTkFrame(llm_card, fg_color="transparent")
        prov_row.pack(fill="x", padx=20, pady=6)
        ctk.CTkLabel(prov_row, text=tr("Provider"), font=ctk.CTkFont(size=14), width=100, anchor="w").pack(side="left")
        self.provider_var = ctk.StringVar(value="ollama")
        providers = ["ollama", "openai", "anthropic", "azure", "google", "deepseek", "groq", "mistral", "openrouter", "github"]
        self.provider_menu = ctk.CTkOptionMenu(prov_row, variable=self.provider_var, values=providers,
                                              fg_color=C_INPUT_BG, button_color=C_ACCENT,
                                              text_color=C_TEXT,
                                              corner_radius=10, height=36,
                                              command=self._on_provider_change)
        self.provider_menu.pack(side="right")

        # Type row
        type_row = ctk.CTkFrame(llm_card, fg_color="transparent")
        type_row.pack(fill="x", padx=20, pady=6)
        ctk.CTkLabel(type_row, text=tr("Type"), font=ctk.CTkFont(size=14), width=100, anchor="w").pack(side="left")
        self.type_label = ctk.CTkLabel(type_row, text=tr("Local"), font=ctk.CTkFont(size=14), text_color=C_SUCCESS)
        self.type_label.pack(side="left")

        # LLM Fields
        def llm_field(parent, label, attr, default="", show=None):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=6)
            ctk.CTkLabel(row, text=tr(label), font=ctk.CTkFont(size=14), width=100, anchor="w").pack(side="left")
            e = ctk.CTkEntry(row, height=40, corner_radius=10, fg_color=C_INPUT_BG, border_color=C_BORDER)
            if default:
                e.insert(0, default)
            e.pack(side="left", fill="x", expand=True, padx=(8, 0))
            setattr(self, attr, e)

        llm_field(llm_card, "Display Name", "llm_name", "My Model")
        llm_field(llm_card, "Model", "llm_model_name", "gemma4:latest")
        llm_field(llm_card, "API Base URL", "llm_api_base", "http://localhost:11434")
        llm_field(llm_card, "API Key", "llm_api_key", show="*")

        # Temperature and Max Tokens row
        temp_row = ctk.CTkFrame(llm_card, fg_color="transparent")
        temp_row.pack(fill="x", padx=20, pady=6)
        ctk.CTkLabel(temp_row, text=tr("Temperature"), font=ctk.CTkFont(size=14), width=100, anchor="w").pack(side="left")
        self.llm_temp_var = ctk.StringVar(value="0.7")
        ctk.CTkEntry(temp_row, height=40, corner_radius=10, width=100, textvariable=self.llm_temp_var,
                    fg_color=C_INPUT_BG, border_color=C_BORDER).pack(side="left", padx=(8, 24))
        ctk.CTkLabel(temp_row, text=tr("Max Tokens"), font=ctk.CTkFont(size=14), anchor="w").pack(side="left")
        self.llm_tk_var = ctk.StringVar(value="4096")
        ctk.CTkEntry(temp_row, height=40, corner_radius=10, width=100, textvariable=self.llm_tk_var,
                    fg_color=C_INPUT_BG, border_color=C_BORDER).pack(side="left", padx=(8, 0))

        # Saved Models row
        sel_row = ctk.CTkFrame(llm_card, fg_color="transparent")
        sel_row.pack(fill="x", padx=20, pady=6)
        ctk.CTkLabel(sel_row, text=tr("Saved Models"), font=ctk.CTkFont(size=14), width=100, anchor="w").pack(side="left")
        self.model_selector = ctk.CTkOptionMenu(sel_row, values=[tr("(New model)")],
                                                fg_color=C_INPUT_BG, button_color=C_ACCENT,
                                                text_color=C_SIDEBAR_TEXT,
                                                corner_radius=10, height=36,
                                                command=self._on_model_selected)
        self.model_selector.pack(side="right", expand=True, fill="x", padx=(8, 0))

        # LLM Buttons row
        lbtn_row = ctk.CTkFrame(llm_card, fg_color="transparent")
        lbtn_row.pack(fill="x", padx=20, pady=(12, 16))
        ctk.CTkButton(lbtn_row, text=tr("Save"), width=100, height=40,
                      fg_color=C_ACCENT, hover_color=C_ACCENT_HOVER,
                      text_color=C_SIDEBAR_TEXT, corner_radius=10,
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._save_llm_model).pack(side="left", padx=(0, 8))
        ctk.CTkButton(lbtn_row, text=tr("Test Connection"), width=130, height=40,
                      fg_color=C_SUCCESS, hover_color=C_SUCCESS_HOVER,
                      text_color=C_SIDEBAR_TEXT, corner_radius=10,
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._test_llm_connection).pack(side="left", padx=(0, 8))
        ctk.CTkButton(lbtn_row, text=tr("Delete"), width=100, height=40,
                      fg_color=C_DANGER, hover_color=C_DANGER_HOVER,
                      text_color=C_SIDEBAR_TEXT, corner_radius=10,
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._delete_llm_model).pack(side="left")

        # ── Data Management Section ───────────────────────────
        data_card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        data_card.pack(fill="x", padx=24, pady=8)

        ctk.CTkLabel(data_card, text=tr("数据管理"),
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT
                     ).pack(anchor="w", padx=20, pady=(16, 8))

        # Backup button
        ctk.CTkButton(data_card, text=tr("创建备份"), width=120, height=36,
                      fg_color=C_SUCCESS, hover_color=C_SUCCESS_HOVER,
                      text_color=C_SIDEBAR_TEXT, corner_radius=10,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=self._create_backup).pack(anchor="w", padx=20, pady=(0, 12))

        # Backup list
        self.backup_list_frame = ctk.CTkFrame(data_card, fg_color="transparent")
        self.backup_list_frame.pack(fill="x", padx=20, pady=(0, 16))

        # ── About Section ───────────────────────────────────
        about_card = ctk.CTkFrame(scroll, fg_color=C_CARD, corner_radius=16, border_width=1, border_color=C_BORDER)
        about_card.pack(fill="x", padx=24, pady=8)

        ctk.CTkLabel(about_card, text=tr("About"),
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=C_TEXT
                     ).pack(anchor="w", padx=20, pady=(16, 8))

        about_items = [
            (tr("LitManager"), tr("Library Management System")),
            (tr("Version"), "0.1.0"),
            (tr("Backend"), tr("FastAPI + SQLite")),
            (tr("Frontend"), tr("Modern Tkinter GUI")),
        ]
        for label, value in about_items:
            row = ctk.CTkFrame(about_card, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=4)
            ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=13, weight="bold"),
                         width=100, anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=value, font=ctk.CTkFont(size=13),
                         anchor="w").pack(side="left", padx=(8, 0))

        # Refresh button
        ctk.CTkButton(scroll, text=tr("Refresh Dashboard Data"), fg_color=C_ACCENT,
                      text_color=C_SIDEBAR_TEXT,
                      corner_radius=12, height=44, width=200,
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._refresh_all).pack(pady=20)

        self._update_api_key_visibility()

    def _set_theme(self, choice):
        mapping = {"System": "System", "Light": "Light", "Dark": "Dark"}
        ctk.set_appearance_mode(mapping.get(choice, "System"))

    def _set_language(self, choice):
        lang = "zh" if choice == "中文" else "en"
        set_language(lang)
        # Defer rebuild to avoid destroying widgets during callback
        if self.app_ref:
            self.app_ref.after(100, self.app_ref.rebuild_all_views)

    def _on_provider_change(self, choice):
        local_provs = ["ollama"]
        is_local = choice in local_provs
        self.type_label.configure(text=tr("Local") if is_local else tr("Cloud"),
                                  text_color=C_SUCCESS if is_local else C_ACCENT)
        self._update_api_key_visibility()
        defaults = {
            "ollama": ("http://localhost:11434", "gemma4:latest"),
            "openai": ("https://api.openai.com/v1", "gpt-4o"),
            "anthropic": ("https://api.anthropic.com/v1", "claude-sonnet-4-20250514"),
            "azure": ("https://{name}.openai.azure.com", "my-deployment"),
            "google": ("https://generativelanguage.googleapis.com/v1beta", "gemini-2.0-flash"),
            "deepseek": ("https://api.deepseek.com/v1", "deepseek-chat"),
            "groq": ("https://api.groq.com/openai/v1", "llama3-70b-8192"),
            "mistral": ("https://api.mistral.ai/v1", "mistral-large-latest"),
            "openrouter": ("https://openrouter.ai/api/v1", "anthropic/claude-3.5-sonnet"),
            "github": ("https://models.inference.ai.azure.com", "gpt-4o"),
        }
        if choice in defaults:
            base, model = defaults[choice]
            self.llm_api_base.delete(0, "end")
            self.llm_api_base.insert(0, base)
            self.llm_model_name.delete(0, "end")
            self.llm_model_name.insert(0, model)

    def _update_api_key_visibility(self):
        is_local = self.provider_var.get() in ["ollama"]
        if hasattr(self, "llm_api_key"):
            if is_local:
                self.llm_api_key.configure(show="", placeholder_text="(not needed for local)")
            else:
                self.llm_api_key.configure(show="*", placeholder_text="Enter API key")

    def _clear_llm_form(self):
        for attr in ["llm_name", "llm_model_name", "llm_api_base", "llm_api_key"]:
            if hasattr(self, attr):
                getattr(self, attr).delete(0, "end")
        if hasattr(self, "llm_temp_var"):
            self.llm_temp_var.set("0.7")
        if hasattr(self, "llm_tk_var"):
            self.llm_tk_var.set("4096")
        if hasattr(self, "provider_menu"):
            self.provider_var.set("ollama")
            self._on_provider_change("ollama")

    def _on_model_selected(self, choice):
        if choice == tr("(New model)"):
            self._clear_llm_form()
            return
        try:
            for m in API.list_models():
                if m.get("name") == choice:
                    self._load_model(m)
                    break
        except Exception as e:
            print(f"Model selection error: {e}")

    def _load_model(self, m):
        self.llm_name.delete(0, "end")
        self.llm_name.insert(0, m.get("name", ""))
        self.llm_model_name.delete(0, "end")
        self.llm_model_name.insert(0, m.get("model_name", ""))
        self.llm_api_base.delete(0, "end")
        self.llm_api_base.insert(0, m.get("api_base") or "")
        self.llm_api_key.delete(0, "end")
        if m.get("api_key"):
            self.llm_api_key.insert(0, "****")
        if hasattr(self, "llm_temp_var"):
            self.llm_temp_var.set(str(m.get("temperature", 0.7)))
        if hasattr(self, "llm_tk_var"):
            self.llm_tk_var.set(str(m.get("max_tokens", 4096)))
        self.provider_var.set(m.get("provider", "openai"))
        self._on_provider_change(m.get("provider", "openai"))

    def _refresh_model_selector(self):
        try:
            models = API.list_models()
            names = [m["name"] for m in models]
            names.insert(0, tr("(New model)"))
            self.model_selector.configure(values=names)
            self.model_selector.set(tr("(New model)"))
        except Exception as e:
            print(f"Refresh model selector error: {e}")

    def _save_llm_model(self):
        name = self.llm_name.get().strip()
        model_name = self.llm_model_name.get().strip()
        if not name or not model_name:
            messagebox.showwarning(tr("Validation"), tr("Name and Model Name are required."))
            return
        api_key = self.llm_api_key.get().strip() if hasattr(self, "llm_api_key") else ""
        if api_key == "****":
            api_key = ""
        data = {
            "name": name,
            "provider": self.provider_var.get(),
            "model_name": model_name,
            "api_base": self.llm_api_base.get().strip() or None,
            "api_key": api_key or None,
            "temperature": float(self.llm_temp_var.get() or "0.7"),
            "max_tokens": int(self.llm_tk_var.get() or "4096"),
        }
        try:
            API.create_model(data)
            ToastNotification(self.winfo_toplevel(), f"{tr('Model saved')}: {name}", "success")
            self._refresh_model_selector()
        except Exception as e:
            messagebox.showerror(tr("Error"), str(e))

    def _test_llm_connection(self):
        name = self.llm_name.get().strip()
        model_name = self.llm_model_name.get().strip()
        if not name or not model_name:
            messagebox.showwarning(tr("Validation"), tr("Fill model name and ID first."))
            return
        api_key = self.llm_api_key.get().strip() if hasattr(self, "llm_api_key") else ""
        if api_key == "****":
            api_key = ""
        data = {
            "name": name,
            "provider": self.provider_var.get(),
            "model_name": model_name,
            "api_base": self.llm_api_base.get().strip() or None,
            "api_key": api_key or None,
            "temperature": float(self.llm_temp_var.get() or "0.7"),
            "max_tokens": int(self.llm_tk_var.get() or "4096"),
        }
        try:
            r = API.create_model(data)
            mid = r.get("id")
            try:
                import httpx
                with httpx.Client(base_url="http://127.0.0.1:8000", timeout=30) as h:
                    chat = h.post("/api/conversations/chat",
                                  json={"model_id": mid, "message": "Reply with just OK."})
                    reply_data = chat.json()
                    reply = reply_data.get("reply", "")
                if "OK" in reply or reply_data.get("tokens_out", 0) > 0:
                    ToastNotification(self.winfo_toplevel(), f"{tr('Connection OK')}! {tr('Reply')}: {reply[:60]}", "success")
                else:
                    messagebox.showinfo(tr("Result"), f"{tr('Reply')}: {reply[:60]}")
            finally:
                # Clean up temporary model created for testing
                try:
                    API.delete_model(mid)
                except Exception:
                    pass
        except Exception as e:
            messagebox.showerror(tr("Error"), f"{tr('Connection failed')}: {e}")

    def _delete_llm_model(self):
        choice = self.model_selector.get()
        if choice == tr("(New model)"):
            return
        dlg = ConfirmDialog(self, tr("Delete"), f'{tr("Delete model")} {choice}?')
        self.wait_window(dlg)
        if dlg.result:
            try:
                for m in API.list_models():
                    if m["name"] == choice:
                        API.delete_model(m["id"])
                        self._clear_llm_form()
                        self._refresh_model_selector()
                        break
            except Exception as e:
                messagebox.showerror(tr("Error"), str(e))

    def refresh(self):
        """Refresh settings data."""
        try:
            # Refresh admin list
            for w in self.admin_list.winfo_children():
                w.destroy()
            data = API.list_admins()
            admins = data.get("data", [])
            if admins:
                for a in admins:
                    row = ctk.CTkFrame(self.admin_list, fg_color="transparent")
                    row.pack(fill="x", pady=2)
                    ctk.CTkLabel(row, text=f"  {a.get('username', '')}",
                                 font=ctk.CTkFont(size=13)).pack(side="left")
                    ctk.CTkLabel(row, text=f"  {a.get('role', '')}",
                                 font=ctk.CTkFont(size=13),
                                 text_color=C_TEXT_SEC).pack(side="left")
            else:
                ctk.CTkLabel(self.admin_list, text=tr("No admins configured"),
                             font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC).pack(pady=8)

            # Refresh model selector
            self._refresh_model_selector()
        except Exception as e:
            print(f"Settings refresh error: {e}")

    def _add_admin(self):
        modal = ModalWindow(self, tr("Add Admin"), width=440, height=380)
        entries = {}
        for key, label in [("username", tr("Username *")), ("password", tr("Password *")),
                           ("name", tr("Name")), ("role", tr("Role"))]:
            row = ctk.CTkFrame(modal.content, fg_color="transparent")
            row.pack(fill="x", pady=8, padx=16)
            ctk.CTkLabel(row, text=label, width=100, anchor="w",
                         font=ctk.CTkFont(size=14)).pack(side="left")
            e = ctk.CTkEntry(row, height=40, corner_radius=10,
                             show="*" if key == "password" else "",
                             fg_color=C_INPUT_BG, border_color=C_BORDER)
            e.pack(side="left", fill="x", expand=True, padx=(8, 0))
            entries[key] = e

        def save():
            data = {}
            for k, e in entries.items():
                val = e.get().strip()
                data[k] = val or None
            if not data.get("username") or not data.get("password"):
                messagebox.showwarning(tr("Validation"), tr("Username and Password are required."))
                return
            try:
                API.create_admin(data)
                modal.dismiss()
                self.refresh()
            except Exception as ex:
                messagebox.showerror(tr("Error"), str(ex))

        ctk.CTkButton(modal.content, text=tr("Save"), fg_color=C_ACCENT,
                      hover_color=C_ACCENT_HOVER,
                      text_color=C_SIDEBAR_TEXT, corner_radius=10, height=44,
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=save).pack(pady=20)

    def _refresh_all(self):
        """Refresh all views."""
        try:
            # Find the app and refresh all views
            root = self.master
            while root and not hasattr(root, "views"):
                root = root.master
            if root and hasattr(root, "views"):
                for view in root.views.values():
                    if hasattr(view, "refresh"):
                        view.refresh()
            messagebox.showinfo(tr("Refresh"), tr("All data refreshed!"))
        except Exception as e:
            print(f"Refresh all error: {e}")

    def _create_backup(self):
        """Create a database backup in a background thread."""
        def _do_backup():
            try:
                result = API.create_backup()
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), result.get("message", "备份成功"), "success"))
                self.after(0, self._refresh_backup_list)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))
        threading.Thread(target=_do_backup, daemon=True).start()

    def _refresh_backup_list(self):
        """Refresh the backup list from the API."""
        try:
            result = API.list_backups()
            backups = result.get("data", [])
            # Clear existing widgets
            for w in self.backup_list_frame.winfo_children():
                w.destroy()
            if not backups:
                ctk.CTkLabel(self.backup_list_frame, text=tr("暂无备份文件"),
                             font=ctk.CTkFont(size=13), text_color=C_TEXT_SEC).pack(pady=8)
                return
            for b in backups:
                row = ctk.CTkFrame(self.backup_list_frame, fg_color=C_INPUT_BG, corner_radius=10)
                row.pack(fill="x", pady=4)
                filename = b.get("filename", "")
                size_kb = b.get("size", 0) / 1024
                created = b.get("created", "")
                info_text = f"{filename}  ({size_kb:.1f} KB)  {created[:19]}"
                ctk.CTkLabel(row, text=info_text, font=ctk.CTkFont(size=12),
                             text_color=C_TEXT, anchor="w").pack(side="left", padx=12, pady=8)
                btn_frame = ctk.CTkFrame(row, fg_color="transparent")
                btn_frame.pack(side="right", padx=8)
                ctk.CTkButton(btn_frame, text=tr("恢复"), width=60, height=28,
                              fg_color=C_WARN, hover_color="#e8c75e",
                              text_color=C_SIDEBAR_TEXT, corner_radius=8,
                              font=ctk.CTkFont(size=12, weight="bold"),
                              command=lambda f=filename: self._restore_backup(f)).pack(side="left", padx=4)
                ctk.CTkButton(btn_frame, text=tr("删除"), width=60, height=28,
                              fg_color=C_DANGER, hover_color=C_DANGER_HOVER,
                              text_color=C_SIDEBAR_TEXT, corner_radius=8,
                              font=ctk.CTkFont(size=12, weight="bold"),
                              command=lambda f=filename: self._delete_backup(f)).pack(side="left", padx=4)
        except Exception as e:
            print(f"Refresh backup list error: {e}")

    def _restore_backup(self, filename):
        """Restore database from a backup with confirmation."""
        dlg = ConfirmDialog(self, tr("恢复确认"), f'{tr("确定要从备份恢复吗?")}\n{filename}')
        self.wait_window(dlg)
        if not dlg.result:
            return
        def _do_restore():
            try:
                result = API.restore_backup(filename)
                self.after(0, lambda: ToastNotification(self.winfo_toplevel(), result.get("message", "恢复成功"), "success"))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(tr("Error"), str(e)))
        threading.Thread(target=_do_restore, daemon=True).start()

    def _delete_backup(self, filename):
        """Delete a backup file with confirmation."""
        dlg = ConfirmDialog(self, tr("删除确认"), f'{tr("确定要删除备份吗?")}\n{filename}')
        self.wait_window(dlg)
        if not dlg.result:
            return
        try:
            API.delete_backup(filename)
            ToastNotification(self.winfo_toplevel(), tr("删除成功"), "success")
            self._refresh_backup_list()
        except Exception as e:
            messagebox.showerror(tr("Error"), str(e))


class ConfirmDialog(ctk.CTkToplevel):
    """Confirmation dialog."""

    def __init__(self, master, title, message):
        super().__init__(master)
        self.title(title)
        self.geometry(f"360x160+{master.winfo_x()+120}+{master.winfo_y()+120}")
        self.transient(master)
        self.grab_set()
        self.resizable(False, False)
        self.result = False

        ctk.CTkLabel(self, text=message, font=ctk.CTkFont(size=14),
                     wraplength=320).pack(pady=20, padx=20)

        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(pady=10)
        ctk.CTkButton(row, text=tr("Cancel"), width=100, height=36,
                      fg_color=C_INPUT_BG, text_color=C_SIDEBAR_TEXT, hover_color=C_BORDER,
                      corner_radius=10, command=self._cancel).pack(side="left", padx=8)
        ctk.CTkButton(row, text=tr("Confirm"), width=100, height=36,
                      fg_color=C_DANGER, hover_color=C_DANGER_HOVER,
                      text_color=C_SIDEBAR_TEXT,
                      corner_radius=10, command=self._confirm).pack(side="left", padx=8)

    def _confirm(self):
        self.result = True
        self.grab_release()
        self.destroy()

    def _cancel(self):
        self.grab_release()
        self.destroy()


class ModalWindow(ctk.CTkToplevel):
    """Modal window for forms."""

    def __init__(self, master, title, width=480, height=520):
        super().__init__(master)
        self.title(title)
        self.geometry(f"{width}x{height}+{master.winfo_x()+80}+{master.winfo_y()+60}")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        # Nav bar
        nav = ctk.CTkFrame(self, fg_color="#fafafa", corner_radius=0, height=50)
        nav.pack(fill="x")
        nav.pack_propagate(False)
        ctk.CTkLabel(nav, text=title, font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=C_TEXT).pack(side="center", pady=12)
        close_btn = ctk.CTkButton(nav, text="X", width=36, height=32,
                                  fg_color="transparent", text_color=C_SIDEBAR_TEXT,
                                  hover_color=C_DANGER, corner_radius=16,
                                  command=self.dismiss, font=ctk.CTkFont(size=14, weight="bold"))
        close_btn.pack(side="right", padx=12, pady=8)

        self.content = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.content.pack(fill="both", expand=True, padx=20, pady=16)

        self.protocol("WM_DELETE_WINDOW", self.dismiss)
        self.bind("<Escape>", lambda e: self.dismiss())

    def dismiss(self, result=None):
        self._result = result
        self.grab_release()
        self.destroy()


class LoginView(ctk.CTkFrame):
    """Login interface with consistent TRAE-style design."""

    def __init__(self, master, on_login_success, **kwargs):
        super().__init__(master, fg_color=C_CONTENT, **kwargs)
        self.on_login_success = on_login_success
        self._mode = "login"  # "login" or "register"
        self._build()
        self._try_start_backend()

    def _try_start_backend(self):
        """Start backend server in background if not already running."""
        def _check_and_start():
            try:
                API.health()
                return  # Already running
            except Exception:
                pass

            # Not running — start it
            try:
                import subprocess
                backend_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backend")
                # Use a log file to capture errors
                log_path = os.path.join(backend_dir, "_server.log")
                with open(log_path, "w") as log_f:
                    log_f.write("Starting backend...\n")
                    log_f.flush()
                    proc = subprocess.Popen(
                        [sys.executable, "-m", "uvicorn", "app.main:app",
                         "--host", "127.0.0.1", "--port", "8000"],
                        cwd=backend_dir,
                        stdout=log_f,
                        stderr=subprocess.STDOUT,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
                    )
                # Wait for backend to be ready (up to 15 seconds)
                import time
                ready = False
                for i in range(30):
                    time.sleep(0.5)
                    try:
                        API.health()
                        ready = True
                        break
                    except Exception:
                        continue
                if not ready:
                    print("Backend failed to start within 15 seconds")
            except Exception as e:
                print(f"Backend start failed: {e}")

        threading.Thread(target=_check_and_start, daemon=True).start()

    def _build(self):
        # Fixed container (no scrollbar)
        center = ctk.CTkFrame(self, fg_color="transparent")
        center.place(relx=0.5, rely=0.5, anchor="center")

        # Logo/Title
        ctk.CTkLabel(
            center,
            text="LitManager",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color=C_TEXT
        ).pack(pady=(0, 8))

        ctk.CTkLabel(
            center,
            text=tr("Library Management System"),
            font=ctk.CTkFont(size=14),
            text_color=C_TEXT_SEC
        ).pack(pady=(0, 30))

        # Login card (no border)
        login_card = ctk.CTkFrame(
            center,
            fg_color=C_CARD,
            corner_radius=16,
            width=380,
            height=380
        )
        login_card.pack(pady=10)
        login_card.pack_propagate(False)

        inner = ctk.CTkFrame(login_card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=40, pady=24)

        if self._mode == "login":
            self._build_login_form(inner)
        else:
            self._build_register_form(inner)

        # Language switch
        lang_frame = ctk.CTkFrame(center, fg_color="transparent")
        lang_frame.pack(pady=(16, 0))

        ctk.CTkLabel(
            lang_frame,
            text=tr("Language") + ":",
            font=ctk.CTkFont(size=12),
            text_color=C_TEXT_SEC
        ).pack(side="left", padx=(0, 8))

        self.lang_zh_btn = ctk.CTkButton(
            lang_frame,
            text="中文",
            font=ctk.CTkFont(size=12),
            fg_color=C_INPUT_BG if get_language() == "zh" else "transparent",
            text_color=C_TEXT,
            width=60,
            height=28,
            corner_radius=6,
            command=lambda: self._switch_lang("zh")
        )
        self.lang_zh_btn.pack(side="left", padx=4)

        self.lang_en_btn = ctk.CTkButton(
            lang_frame,
            text="English",
            font=ctk.CTkFont(size=12),
            fg_color=C_INPUT_BG if get_language() == "en" else "transparent",
            text_color=C_TEXT,
            width=70,
            height=28,
            corner_radius=6,
            command=lambda: self._switch_lang("en")
        )
        self.lang_en_btn.pack(side="left", padx=4)

    def _build_login_form(self, inner):
        # Username
        ctk.CTkLabel(
            inner,
            text=tr("Username"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(anchor="w", pady=(0, 6))

        self.username_entry = ctk.CTkEntry(
            inner,
            placeholder_text=tr("Enter username"),
            font=ctk.CTkFont(size=14),
            height=42,
            corner_radius=10
        )
        self.username_entry.pack(fill="x", pady=(0, 16))
        self.username_entry.bind("<Return>", lambda e: self.password_entry.focus())

        # Password
        ctk.CTkLabel(
            inner,
            text=tr("Password"),
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=C_TEXT,
            anchor="w"
        ).pack(anchor="w", pady=(0, 6))

        # Password entry with show/hide toggle
        pwd_frame = ctk.CTkFrame(inner, fg_color=C_INPUT_BG, corner_radius=10, height=42)
        pwd_frame.pack(fill="x", pady=(0, 12))
        pwd_frame.pack_propagate(False)

        self.password_entry = ctk.CTkEntry(
            pwd_frame,
            placeholder_text=tr("Enter password"),
            font=ctk.CTkFont(size=14),
            height=42,
            corner_radius=10,
            show="*",
            fg_color="transparent",
            border_width=0
        )
        self.password_entry.pack(side="left", fill="x", expand=True, padx=(12, 0))
        self.password_entry.bind("<Return>", lambda e: self._do_login())

        self.show_pwd_btn = ctk.CTkButton(
            pwd_frame,
            text="👁",
            font=ctk.CTkFont(size=14),
            fg_color="transparent",
            text_color=C_TEXT_SEC,
            hover_color=C_ACCENT_HOVER,
            width=36,
            height=36,
            command=self._toggle_password_visibility
        )
        self.show_pwd_btn.pack(side="right", padx=(0, 8))
        self._password_visible = False

        # Remember me + Forgot password row
        options_row = ctk.CTkFrame(inner, fg_color="transparent")
        options_row.pack(fill="x", pady=(0, 16))

        self.remember_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            options_row,
            text=tr("Remember me"),
            variable=self.remember_var,
            font=ctk.CTkFont(size=12),
            text_color=C_TEXT_SEC,
            fg_color=C_INPUT_BG,
            hover_color=C_ACCENT_HOVER,
            border_color=C_TEXT_SEC,
            border_width=2,
            width=120
        ).pack(side="left")

        ctk.CTkButton(
            options_row,
            text=tr("Forgot password?"),
            font=ctk.CTkFont(size=12),
            fg_color="transparent",
            text_color=C_TEXT_SEC,
            hover_color=C_INPUT_BG,
            width=120,
            height=28,
            command=self._forgot_password
        ).pack(side="right")

        # Login button
        self.login_btn = ctk.CTkButton(
            inner,
            text=tr("Login"),
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=44,
            corner_radius=10,
            command=self._do_login
        )
        self.login_btn.pack(fill="x")

        # Error message
        self.error_label = ctk.CTkLabel(
            inner,
            text="",
            font=ctk.CTkFont(size=12),
            text_color="#ef4444",
            anchor="center"
        )
        self.error_label.pack(pady=(8, 0))

        # Register link
        ctk.CTkButton(
            inner,
            text=tr("No account? Register now"),
            font=ctk.CTkFont(size=12),
            fg_color="transparent",
            text_color=C_TEXT_SEC,
            hover_color=C_INPUT_BG,
            height=28,
            command=self._switch_to_register
        ).pack(pady=(8, 0))

    def _build_register_form(self, inner):
        # Title
        ctk.CTkLabel(
            inner,
            text=tr("Register Account"),
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=C_TEXT,
            anchor="center"
        ).pack(pady=(0, 16))

        fields = {}
        for label_text, key, placeholder in [
            (tr("Username"), "username", tr("Enter username")),
            (tr("Password"), "password", tr("Enter password")),
            (tr("Confirm Password"), "confirm", tr("Confirm password")),
            (tr("Name"), "name", tr("Enter name (optional)")),
        ]:
            ctk.CTkLabel(
                inner,
                text=label_text,
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=C_TEXT,
                anchor="w"
            ).pack(anchor="w", pady=(0, 4))

            entry = ctk.CTkEntry(
                inner,
                placeholder_text=placeholder,
                font=ctk.CTkFont(size=13),
                height=36,
                corner_radius=8
            )
            if key in ("password", "confirm"):
                entry.configure(show="*")
            entry.pack(fill="x", pady=(0, 10))
            fields[key] = entry

        self.reg_fields = fields

        # Error label
        self.error_label = ctk.CTkLabel(
            inner,
            text="",
            font=ctk.CTkFont(size=11),
            text_color="#ef4444",
            anchor="center"
        )
        self.error_label.pack(pady=(0, 8))

        # Register button
        self.login_btn = ctk.CTkButton(
            inner,
            text=tr("Register"),
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=C_ACCENT,
            hover_color=C_ACCENT_HOVER,
            text_color=C_SIDEBAR_TEXT,
            height=40,
            corner_radius=10,
            command=self._do_register
        )
        self.login_btn.pack(fill="x", pady=(0, 8))

        # Back to login link
        ctk.CTkButton(
            inner,
            text=tr("Back to login"),
            font=ctk.CTkFont(size=12),
            fg_color="transparent",
            text_color=C_TEXT_SEC,
            hover_color=C_INPUT_BG,
            height=28,
            command=self._switch_to_login
        ).pack()

    def _switch_to_register(self):
        self._mode = "register"
        self.destroy()
        view = LoginView(self.master, self.on_login_success)
        view.pack(fill="both", expand=True)
        view.focus_username()

    def _switch_to_login(self):
        self._mode = "login"
        self.destroy()
        view = LoginView(self.master, self.on_login_success)
        view.pack(fill="both", expand=True)
        view.focus_username()

    def _switch_lang(self, lang):
        set_language(lang)
        # Rebuild the login view
        self.destroy()
        login_view = LoginView(self.master, self.on_login_success)
        login_view.pack(fill="both", expand=True)
        login_view.focus_username()

    def _do_login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()

        if not username or not password:
            self.error_label.configure(text=tr("Please enter username and password"))
            return

        self.login_btn.configure(state="disabled", text=tr("Logging in..."))
        self.error_label.configure(text="")

        def _login():
            import time
            max_retries = 8
            last_err = None
            callback = self.on_login_success  # capture callback reference
            
            # 先尝试管理员登录
            for attempt in range(max_retries):
                try:
                    result = API.login(username, password)
                    # 管理员登录成功
                    result['user_type'] = 'admin'
                    self.after(0, lambda r=result: callback(r))
                    return
                except Exception as e:
                    last_err = str(e)
                    # 如果是认证错误，尝试读者登录
                    if "401" in last_err or "用户名或密码错误" in last_err:
                        # 尝试读者登录
                        try:
                            reader_result = API.reader_login(username, password)
                            # 读者登录成功
                            reader_result['user_type'] = 'reader'
                            self.after(0, lambda r=reader_result: callback(r))
                            return
                        except Exception as reader_err:
                            reader_err_str = str(reader_err)
                            if "401" in reader_err_str or "读者不存在" in reader_err_str or "密码错误" in reader_err_str:
                                self.after(0, lambda: self._show_error(tr("Invalid username or password")))
                                return
                            # 其他错误继续重试
                    # Connection error — backend may still be starting, wait and retry
                    if attempt < max_retries - 1:
                        time.sleep(1.5)
                    else:
                        err_msg = tr("Cannot connect to server. Please ensure the backend is running.")
                        self.after(0, lambda: self._show_error(err_msg))

        threading.Thread(target=_login, daemon=True).start()

    def _do_register(self):
        uname = self.reg_fields["username"].get().strip()
        pwd = self.reg_fields["password"].get().strip()
        confirm = self.reg_fields["confirm"].get().strip()
        name = self.reg_fields["name"].get().strip() or None

        if not uname or not pwd:
            self.error_label.configure(text=tr("Username and password are required"))
            return
        if pwd != confirm:
            self.error_label.configure(text=tr("Passwords do not match"))
            return

        self.login_btn.configure(state="disabled", text=tr("Registering..."))
        self.error_label.configure(text="")

        def _register():
            try:
                client = APIClient()
                client.create_admin({
                    "username": uname,
                    "password": pwd,
                    "name": name,
                    "role": "普通管理员",
                })
                self.after(0, lambda: self._switch_to_login())
            except Exception as e:
                err_msg = str(e)
                if "已存在" in err_msg or "exists" in err_msg.lower():
                    err_msg = tr("Username already exists")
                self.after(0, lambda: self._show_error(err_msg))

        threading.Thread(target=_register, daemon=True).start()

    def _show_error(self, msg):
        self.error_label.configure(text=msg)
        self.login_btn.configure(state="normal", text=tr("Login") if self._mode == "login" else tr("Register"))

    def _toggle_password_visibility(self):
        """Toggle password visibility between hidden (*) and visible."""
        self._password_visible = not self._password_visible
        if self._password_visible:
            self.password_entry.configure(show="")
            self.show_pwd_btn.configure(text="")
        else:
            self.password_entry.configure(show="*")
            self.show_pwd_btn.configure(text="👁")

    def _forgot_password(self):
        messagebox.showinfo(
            tr("Forgot Password"),
            tr("Please contact the system administrator to reset your password.")
        )

    def focus_username(self):
        """Focus on username entry."""
        if self._mode == "login":
            self.username_entry.focus()
        else:
            self.reg_fields["username"].focus()


if __name__ == "__main__":
    # Set default language to Chinese
    set_language("zh")

    root = ctk.CTk()
    root.title("LitManager - Login")
    root.geometry("500x650")
    root.minsize(400, 580)

    def on_login_success(user_info):
        """Called after successful login. Route to different UI based on user type."""
        # Must quit mainloop BEFORE destroying the window
        root.quit()
        root.destroy()
        
        user_type = user_info.get('user_type', 'admin')
        
        if user_type == 'reader':
            # 普通用户 - 跳转到读者界面
            from reader_gui import start_reader_app
            start_reader_app(user_info)
        else:
            # 管理员 - 跳转到管理员界面
            app = ModernApp(admin_info=user_info)
            app.mainloop()

    login_view = LoginView(root, on_login_success)
    login_view.pack(fill="both", expand=True)
    login_view.focus_username()

    root.mainloop()
