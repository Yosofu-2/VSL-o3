"""Generate 5000 random books Excel file for import testing."""
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

# ── Data pools ──

FIRST_NAMES_CN = [
    "张", "王", "李", "赵", "陈", "杨", "刘", "黄", "周", "吴",
    "徐", "孙", "马", "朱", "胡", "郭", "何", "高", "林", "罗",
]
LAST_NAMES_CN = [
    "伟", "芳", "娜", "秀英", "敏", "静", "丽", "强", "磊", "军",
    "洋", "勇", "艳", "杰", "娟", "涛", "明", "超", "秀兰", "霞",
    "平", "刚", "桂英", "华", "建国", "文", "辉", "玉兰", "雪", "婷",
]

CN_BOOK_TITLES = [
    "红楼梦", "西游记", "水浒传", "三国演义", "儒林外史", "聊斋志异",
    "金瓶梅", "封神演义", "东周列国志", "隋唐演义", "说岳全传", "杨家将",
    "七侠五义", "三侠五义", "儿女英雄传", "镜花缘", "老残游记", "官场现形记",
    "二十年目睹之怪现状", "孽海花", "搜神记", "世说新语", "太平广记", "阅微草堂笔记",
    "浮生六记", "随园食单", "陶庵梦忆", "西湖梦寻", "幽梦影", "小窗幽记",
    "菜根谭", "围炉夜话", "增广贤文", "弟子规", "三字经", "百家姓",
    "千字文", "大学", "中庸", "论语", "孟子", "诗经",
    "楚辞", "尚书", "礼记", "春秋", "周易", "道德经",
    "庄子", "荀子", "韩非子", "墨子", "孙子兵法", "三十六计",
]

CN_SUBJECTS = [
    "中国文学", "外国文学", "科幻小说", "历史研究", "哲学思想", "心理学",
    "经济学", "管理学", "计算机科学", "人工智能", "数学", "物理学",
    "化学", "生物学", "医学", "法学", "教育学", "艺术设计",
    "建筑学", "音乐", "电影", "摄影", "旅游", "美食",
    "体育", "军事", "政治", "社会学", "语言学", "考古学",
]

CN_PUBLISHERS = [
    "人民文学出版社", "商务印书馆", "中华书局", "上海译文出版社", "译林出版社",
    "北京大学出版社", "清华大学出版社", "复旦大学出版社", "浙江大学出版社",
    "机械工业出版社", "电子工业出版社", "人民邮电出版社", "科学出版社",
    "高等教育出版社", "中国社会科学出版社", "三联书店", "新星出版社",
    "中信出版社", "南海出版公司", "作家出版社", "花城出版社",
    "江苏文艺出版社", "浙江文艺出版社", "上海文艺出版社", "北京十月文艺出版社",
]

EN_FIRST_NAMES = [
    "James", "John", "Robert", "Michael", "William", "David", "Richard",
    "Joseph", "Thomas", "Charles", "Mary", "Patricia", "Jennifer", "Linda",
    "Elizabeth", "Barbara", "Susan", "Jessica", "Sarah", "Karen",
]

EN_BOOK_TITLES = [
    "Pride and Prejudice", "Sense and Sensibility", "Emma", "Mansfield Park",
    "Northanger Abbey", "Persuasion", "Wuthering Heights", "Jane Eyre",
    "Great Expectations", "Oliver Twist", "A Tale of Two Cities", "David Copperfield",
    "The Picture of Dorian Gray", "Dracula", "Frankenstein", "The Time Machine",
    "The War of the Worlds", "The Invisible Man", "Treasure Island", "Kidnapped",
    "Robinson Crusoe", "Gulliver's Travels", "Moby Dick", "The Scarlet Letter",
    "The Adventures of Huckleberry Finn", "The Great Gatsby", "To Kill a Mockingbird",
    "1984", "Animal Farm", "Brave New World", "Fahrenheit 451",
    "The Catcher in the Rye", "Lord of the Flies", "Of Mice and Men",
    "The Old Man and the Sea", "A Farewell to Arms", "For Whom the Bell Tolls",
    "The Sun Also Rises", "Catch-22", "Slaughterhouse-Five",
    "One Hundred Years of Solitude", "Love in the Time of Cholera",
    "The Alchemist", "The Little Prince", "Don Quixote",
    "Crime and Punishment", "War and Peace", "Anna Karenina",
    "The Brothers Karamazov", "Dead Souls", "Les Misérables",
    "The Count of Monte Cristo", "The Three Musketeers",
    "Madame Bovary", "The Stranger", "The Plague",
    "Ulysses", "A Portrait of the Artist as a Young Man",
    "The Odyssey", "The Iliad", "Hamlet", "Macbeth",
    "Romeo and Juliet", "Othello", "King Lear",
    "The Divine Comedy", "Paradise Lost", "The Canterbury Tales",
]

EN_PUBLISHERS = [
    "Penguin Classics", "Oxford University Press", "Cambridge University Press",
    "HarperCollins", "Random House", "Simon & Schuster", "Macmillan",
    "Hachette", "Scholastic", "Wiley", "Springer", "Elsevier",
    "MIT Press", "Yale University Press", "Princeton University Press",
]

EN_SUBJECTS = [
    "Fiction", "Science Fiction", "Fantasy", "Mystery", "Thriller",
    "Romance", "Historical Fiction", "Biography", "History",
    "Philosophy", "Psychology", "Economics", "Business",
    "Computer Science", "Mathematics", "Physics", "Chemistry",
    "Biology", "Medicine", "Law", "Education", "Art",
    "Music", "Architecture", "Travel", "Cooking", "Sports",
]

GENRES = ["小说", "散文", "诗歌", "戏剧", "传记", "学术", "教材", "科普", "工具书", "绘本"]
LANGUAGES = ["中文", "英文", "中英对照"]
BINDINGS = ["平装", "精装", "线装"]

# ── Helpers ──

def gen_isbn():
    prefix = random.choice(["978", "979"])
    group = random.randint(0, 99)
    publisher = random.randint(100, 999)
    title = random.randint(1000, 9999)
    check = random.randint(0, 9)
    return f"{prefix}-{group:02d}-{publisher:03d}-{title:04d}-{check}"

def gen_year():
    return random.randint(1950, 2025)

def gen_pages():
    return random.randint(100, 800)

def gen_price():
    return round(random.uniform(15.0, 128.0), 2)

def gen_cn_author():
    return random.choice(FIRST_NAMES_CN) + random.choice(LAST_NAMES_CN)

def gen_en_author():
    return random.choice(EN_FIRST_NAMES) + " " + random.choice(["Smith", "Johnson", "Williams", "Brown", "Jones", "Davis", "Miller", "Wilson", "Moore", "Taylor"])

# ── Generate books ──

books = []
categories_set = set()

for i in range(5000):
    is_cn = random.random() < 0.6  # 60% Chinese books

    if is_cn:
        title = random.choice(CN_BOOK_TITLES)
        if random.random() < 0.3:
            suffix = random.choice(["（修订版）", "（新版）", "（典藏版）", "（插图版）", "（注释版）",
                                     "（上）", "（下）", "（全译本）", "（青少版）", f" 第{random.randint(2,5)}版"])
            title += suffix
        author = gen_cn_author()
        if random.random() < 0.15:
            author += f"，{gen_cn_author()}"
        publisher = random.choice(CN_PUBLISHERS)
        subject = random.choice(CN_SUBJECTS)
        language = random.choice(LANGUAGES)
    else:
        title = random.choice(EN_BOOK_TITLES)
        if random.random() < 0.25:
            suffix = random.choice([" (Revised Edition)", " (2nd Edition)", " (Illustrated)",
                                     " (Annotated)", " (Collector's Edition)", " (Abridged)"])
            title += suffix
        author = gen_en_author()
        if random.random() < 0.1:
            author += f", {gen_en_author()}"
        publisher = random.choice(EN_PUBLISHERS)
        subject = random.choice(EN_SUBJECTS)
        language = "英文"

    category = subject  # Use subject as category
    categories_set.add(category)

    books.append({
        "title": title,
        "authors": author,
        "publisher": publisher,
        "publication_year": gen_year(),
        "isbn": gen_isbn(),
        "category_name": category,
        "genre": random.choice(GENRES),
        "language": language,
        "pages": gen_pages(),
        "price": gen_price(),
        "binding": random.choice(BINDINGS),
        "copies": random.randint(1, 10),
        "description": f"这是一本关于{subject}的{random.choice(GENRES)}作品，适合广大读者阅读与收藏。" if is_cn
                       else f"A {random.choice(GENRES)} work on {subject}, suitable for general readers and collectors.",
    })

# ── Write Excel ──

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Books"

headers = ["title", "authors", "publisher", "publication_year", "isbn",
           "category_name", "genre", "language", "pages", "price",
           "binding", "copies", "description"]

header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, book in enumerate(books, 2):
    for col_idx, key in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=book[key])

# Column widths
widths = [35, 20, 22, 14, 22, 16, 10, 10, 8, 8, 8, 8, 50]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

# Freeze header
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:M{len(books)+1}"

output_path = r"e:\SAL o3\test_books_5000.xlsx"
wb.save(output_path)

print(f"Generated {len(books)} books")
print(f"Categories: {len(categories_set)}")
print(f"Saved to: {output_path}")

# Show category distribution
from collections import Counter
cat_counts = Counter(b["category_name"] for b in books)
print("\nTop 10 categories:")
for cat, count in cat_counts.most_common(10):
    print(f"  {cat}: {count}")
